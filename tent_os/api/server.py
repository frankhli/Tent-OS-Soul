"""Tent OS HTTP API Server —— 对外统一接口 + Control UI

提供 RESTful API:
    POST /api/v1/tasks              提交任务
    GET  /api/v1/tasks/{task_id}    查询任务状态
    GET  /api/v1/health             健康检查
    POST /api/v1/approval/{session_id}  审批任务（通过/拒绝）

提供 Control UI:
    GET  /ui/                       Control UI 首页（React SPA）
    GET  /ui/api/*                  UI 数据 API
    WS   /ws                        WebSocket 实时通信
"""

import asyncio
import base64
import json
import os
import sqlite3
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional, List, Set

from fastapi import FastAPI, HTTPException, BackgroundTasks, WebSocket, WebSocketDisconnect, Request, File, UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, RedirectResponse, HTMLResponse
from pydantic import BaseModel, Field
import uvicorn
import yaml

from tent_os.bootstrap import (
    load_config, create_message_bus, create_state_store, create_llm, create_vision_llm,
)
from tent_os.channels.base import ChannelReply
from tent_os.logging_config import get_logger
from tent_os.soul import (
    ThoughtExtractor, VoiceModeler, AppearanceModeler,
    StyleFinetuner, TTSSynthesizer, AuthorizationEngine,
)
from tent_os.api.soul_state import ws_manager, _handle_fast_chat, state as soul_state_ref

logger = get_logger()


# ========== Pydantic 模型 ==========

class TaskSubmitRequest(BaseModel):
    task: str = Field(..., description="用户任务描述")
    tools: list = Field(default=[], description="可用工具列表")
    user_id: Optional[str] = Field(default=None, description="用户ID")
    session_id: Optional[str] = Field(default=None, description="会话ID（复用已有会话）")


class FeedbackRequest(BaseModel):
    type: str = Field(..., description="反馈类型: like / dislike / correct")
    correction: Optional[str] = Field(default=None, description="纠正内容（correct 时必填）")
    message_index: Optional[int] = Field(default=None, description="消息索引")


class TaskSubmitResponse(BaseModel):
    session_id: str
    status: str = "accepted"
    message: str = "任务已提交，正在处理"


class TaskStatusResponse(BaseModel):
    session_id: str
    status: str
    result: Optional[Any] = None


class ApprovalRequest(BaseModel):
    approved: bool = Field(..., description="是否批准执行")


# ========== WebSocket 管理器 ==========




# ========== 执行模式管理 ==========

EXECUTOR_MODE_KEY = "tent:executor_mode"  # Redis key for executor mode

async def _get_executor_mode() -> str:
    """获取当前执行者模式（sandbox / local）"""
    try:
        from tent_os.bootstrap import create_state_store, load_config
        config = load_config("./config/tent_os.yaml")
        store = create_state_store(config)
        # Try Redis first
        if hasattr(store, 'redis'):
            mode = await store.redis.get(EXECUTOR_MODE_KEY)
            if mode:
                return mode.decode() if isinstance(mode, bytes) else mode
        # Fallback to config
        return config.get("local_executor", {}).get("mode", "auto")
    except Exception:
        return "auto"

async def _set_executor_mode(mode: str) -> bool:
    """设置执行者模式"""
    redis_ok = False
    nats_ok = False
    try:
        # 优先使用 state 已有的 state_store（避免重复创建连接）
        if state.state_store and hasattr(state.state_store, 'redis'):
            await state.state_store.redis.set(EXECUTOR_MODE_KEY, mode)
            redis_ok = True
        else:
            # fallback：尝试创建新连接
            from tent_os.bootstrap import create_state_store, load_config
            config = load_config("./config/tent_os.yaml")
            store = create_state_store(config)
            if hasattr(store, 'redis'):
                await store.redis.set(EXECUTOR_MODE_KEY, mode)
                redis_ok = True
    except Exception as e:
        logger.warning(f"设置执行者模式(Redis)失败: {e}")
    
    try:
        # Also broadcast via NATS
        if state.bus:
            await state.bus.publish("executor.mode.switch", json.dumps({
                "mode": mode,
                "timestamp": asyncio.get_event_loop().time(),
            }).encode())
            nats_ok = True
    except Exception as e:
        logger.warning(f"设置执行者模式(NATS广播)失败: {e}")
    
    # Redis 设置成功即算成功，NATS 广播失败不影响功能
    return redis_ok


# ========== 全局状态 ==========

class APIServerState:
    """API Server 全局状态"""
    def __init__(self):
        self.bus = None
        self.state_store = None
        self.config = None
        self.pending_results: Dict[str, asyncio.Future] = {}
        self._db: Optional[sqlite3.Connection] = None
        self._results_cache: Dict[str, Dict] = {}  # session_id -> result
        self._pending_count: Dict[str, int] = {}  # FIX v5: 同session pending请求计数
        self._pending_approvals: Dict[str, Dict] = {}  # session_id -> approval_request
        # Layer 5: 视觉情绪广播节流——记录每个用户上次广播的情绪
        self._last_vision_emotion: Dict[str, str] = {}
        self._nats_subscriptions = []
        self._dreaming = None  # DreamingEngine 实例，由 TentOS 主类注入
        self._llm = None
        self._scene_engine = None  # SceneEngine 实例（空间认知层）
        self.soul_layer = None  # [SOUL] 灵魂积累层，由 TentOS 主类注入

    async def setup(self, config: Dict):
        self.config = config
        self.bus = create_message_bus(config)
        await self.bus.connect()
        self.state_store = create_state_store(config)
        # 初始化 LLM（用于视觉分析、凭证验证等）
        try:
            self._llm = create_llm(config)
            logger.info(f"LLM 初始化成功: {getattr(self._llm, 'model_id', getattr(self._llm, 'model', 'unknown'))}")
        except Exception as e:
            logger.warning(f"LLM 初始化失败: {e}")
        
        # 初始化 Vision LLM（多模态图片分析专用）
        try:
            self._vision_llm = create_vision_llm(config)
            if self._vision_llm:
                logger.info(f"Vision LLM 初始化成功: {getattr(self._vision_llm, 'model_id', getattr(self._vision_llm, 'model', 'unknown'))}")
            else:
                logger.info("Vision LLM 未配置，视觉功能将回退到主 LLM 或启发式模式")
        except Exception as e:
            logger.warning(f"Vision LLM 初始化失败: {e}")
            self._vision_llm = None
        
        # 初始化 SceneEngine（空间认知层）
        try:
            from tent_os.services.scene_engine import SceneEngine
            scene_cfg = config.get("scenes", {})
            if scene_cfg:
                self._scene_engine = SceneEngine(self.bus, {"scenes": scene_cfg})
                logger.info(f"SceneEngine 初始化成功: {len(scene_cfg)} 个场景")
            else:
                logger.info("SceneEngine 未配置场景，空间认知层待机")
        except Exception as e:
            logger.warning(f"SceneEngine 初始化失败: {e}")
            self._scene_engine = None
        
        # 连接 SQLite（scheduler 的 WAL 模式数据库），支持跨进程读
        db_path = config.get("scheduler", {}).get("db_path", "./tent_scheduler.db")
        self._db = sqlite3.connect(db_path, check_same_thread=False)
        self._db.execute("PRAGMA journal_mode=WAL")
        self._db.row_factory = sqlite3.Row
        # 初始化审批历史表
        self._db.execute("""
            CREATE TABLE IF NOT EXISTS approval_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id TEXT NOT NULL,
                plan_summary TEXT,
                approved INTEGER,
                approved_by TEXT DEFAULT 'web_user',
                created_at TEXT DEFAULT (datetime('now')),
                decided_at TEXT
            )
        """)
        self._db.commit()
        # [SOUL] 世界系统表已移除
        # 订阅治理完成消息
        sub1 = await self.bus.nats.subscribe("governance.response.*", cb=self._on_governance_response)
        self._nats_subscriptions.append(sub1)
        # 订阅流式输出（对话实时推送）
        sub_stream = await self.bus.nats.subscribe("governance.stream.*", cb=self._on_stream_chunk)
        self._nats_subscriptions.append(sub_stream)
        # 订阅 reasoning 流（思考过程）
        sub_reasoning = await self.bus.nats.subscribe("governance.stream.reasoning.*", cb=self._on_reasoning_chunk)
        self._nats_subscriptions.append(sub_reasoning)
        # 订阅内心独白流（Phase 5）
        sub_monologue = await self.bus.nats.subscribe("governance.stream.monologue.*", cb=self._on_monologue_chunk)
        self._nats_subscriptions.append(sub_monologue)
        # 订阅中间状态广播
        sub2 = await self.bus.nats.subscribe("governance.plan_update.*", cb=self._on_plan_update)
        self._nats_subscriptions.append(sub2)
        sub3 = await self.bus.nats.subscribe("scheduler.step_update.*", cb=self._on_step_update)
        self._nats_subscriptions.append(sub3)
        # 订阅审批请求
        sub_approval = await self.bus.nats.subscribe("governance.approval.request", cb=self._on_approval_request)
        self._nats_subscriptions.append(sub_approval)
        # 订阅情绪广播
        sub_emotion = await self.bus.nats.subscribe("emotion.broadcast", cb=self._on_emotion_broadcast)
        self._nats_subscriptions.append(sub_emotion)
        sub_fused = await self.bus.nats.subscribe("emotion.fused", cb=self._on_emotion_fused)
        self._nats_subscriptions.append(sub_fused)
        # 订阅系统健康告警（P0: 感知融合层）
        sub_sys_health = await self.bus.nats.subscribe("system.health_alert", cb=self._on_system_health_alert)
        self._nats_subscriptions.append(sub_sys_health)
        # [SOUL] 物理执行/视觉感知/自治微决策/社区/家园/社交 订阅已移除
        # [SOUL] 消息渠道和Skills已简化
        logger.info("API Server 已连接消息总线 —— Tent OS 灵魂对讲机模式")

    async def _on_autonomy_decision(self, msg):
        """接收自治微决策，转发给 WebSocket 客户端"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "spacetime.autonomy",
                "payload": {
                    "decision": data.get("decision", ""),
                    "reason": data.get("reason", ""),
                    "fatigue": data.get("fatigue", 0),
                    "task_load": data.get("task_load", 0),
                    "suggested_action": data.get("suggested_action", ""),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception as e:
            logger.debug(f"[API] 自治决策广播失败: {e}")

    async def cleanup(self):
        for sub in self._nats_subscriptions:
            await sub.unsubscribe()
        if self.bus:
            await self.bus.close()
        if self._db:
            self._db.close()
        logger.info("API Server 已断开消息总线")

    async def _on_governance_response(self, msg):
        """接收治理进程的完成通知，转发给 WebSocket 客户端"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            session_id = data.get("session_id")
            msg_type = data.get("type", "")
            error = data.get("error")
            # chat.completed 用 content 字段，task.completed 用 result 字段
            if msg_type == "chat.completed":
                result = data.get("content", "")
            else:
                result = data.get("result")
            if session_id:
                # FIX v5: 减少pending计数，如果还有未完成的请求，不写入缓存（避免旧任务覆盖）
                self._pending_count[session_id] = self._pending_count.get(session_id, 0) - 1
                if self._pending_count.get(session_id, 0) > 0:
                    logger.debug(f"[API] 同session还有pending请求 [{session_id}]，暂不写入缓存")
                else:
                    task_id = f"chat_{uuid.uuid4().hex[:16]}"
                    self._results_cache[session_id] = {"status": "completed", "result": result, "task_id": task_id, "created_at": datetime.now().isoformat()}
                if session_id in self.pending_results and not self.pending_results[session_id].done():
                    self.pending_results[session_id].set_result(result)
                # 转发给 WebSocket —— 保留原始消息类型（chat.completed / task.completed）
                if msg_type == "chat.completed":
                    ws_type = "chat.completed"
                    payload = {
                        "session_id": session_id,
                        "content": data.get("content", ""),
                        "reasoning": data.get("reasoning", ""),
                        "source": data.get("source", ""),
                        "proactive_type": data.get("proactive_type", ""),
                        "explanation": data.get("explanation", ""),
                    }
                else:
                    ws_type = "task.failed" if error else "task.completed"
                    payload = {
                        "session_id": session_id,
                        "result": result if not error else None,
                        "error": error,
                    }
                await ws_manager.broadcast({
                    "type": ws_type,
                    "payload": payload,
                    "timestamp": asyncio.get_event_loop().time(),
                })
                
                # Phase 2: 任务完成后广播AI情绪变化
                if ws_type == "task.completed" and not error:
                    try:
                        from tent_os.services.emotion_service import EmotionService
                        emotion_svc = EmotionService()
                        # 从state_store获取user_id
                        state_data = await self.state_store.load(session_id) if self.state_store else {}
                        user_id = state_data.get("user_id", "web_user")
                        emotion = emotion_svc.update_by_task_action(user_id, "task_passed")
                        persona = emotion_svc.get_persona(user_id)
                        await ws_manager.broadcast({
                            "type": "ai.emotion",
                            "payload": {"user_id": user_id, "emotion": emotion, "source": "task_passed", "persona": persona},
                            "timestamp": asyncio.get_event_loop().time(),
                        })
                    except Exception as e:
                        logger.debug(f"[API] 情绪广播失败: {e}")
                    
                    # Phase World: 任务完成后自动掉落藏品、增加经验、检查升级
                    try:
                        from tent_os.api.world_api import on_task_completed
                        result_text = str(result) if result else ""
                        world_result = on_task_completed(session_id, result_text)
                        if world_result.get("leveled_up"):
                            await ws_manager.broadcast({
                                "type": "world.level_up",
                                "payload": {
                                    "new_level": world_result["new_level"],
                                    "unlocked_rooms": world_result["unlocked_rooms"],
                                },
                                "timestamp": asyncio.get_event_loop().time(),
                            })
                        if world_result.get("artifact_created"):
                            await ws_manager.broadcast({
                                "type": "world.artifact_dropped",
                                "payload": {
                                    "artifact_id": world_result["artifact_id"],
                                    "exp_gain": world_result["exp_gain"],
                                    "leveled_up": world_result["leveled_up"],
                                },
                                "timestamp": asyncio.get_event_loop().time(),
                            })
                    except Exception as e:
                        logger.debug(f"[API] 世界系统任务回调失败: {e}")
                elif ws_type == "task.failed":
                    try:
                        from tent_os.services.emotion_service import EmotionService
                        emotion_svc = EmotionService()
                        state_data = await self.state_store.load(session_id) if self.state_store else {}
                        user_id = state_data.get("user_id", "web_user")
                        emotion = emotion_svc.update_by_task_action(user_id, "task_failed")
                        persona = emotion_svc.get_persona(user_id)
                        await ws_manager.broadcast({
                            "type": "ai.emotion",
                            "payload": {"user_id": user_id, "emotion": emotion, "source": "task_failed", "persona": persona},
                            "timestamp": asyncio.get_event_loop().time(),
                        })
                    except Exception as e:
                        logger.debug(f"[API] 情绪广播失败: {e}")
        except Exception as e:
            logger.warning(f"处理治理响应出错: {e}")

    async def _on_plan_update(self, msg):
        """接收 Plan 生成中间状态，转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "task.plan",
                "payload": data,
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_step_update(self, msg):
        """接收 Step 执行中间状态，转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "task.step",
                "payload": data,
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_approval_request(self, msg):
        """接收 Plan 审批请求，缓存并通知前端"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            session_id = data.get("session_id")
            if session_id:
                self._pending_approvals[session_id] = data
                # 记录审批历史
                try:
                    plan = data.get("plan", {})
                    summary = plan.get("summary") or plan.get("task") or json.dumps(plan, ensure_ascii=False)[:200]
                    self._db.execute(
                        "INSERT INTO approval_history (session_id, plan_summary, approved) VALUES (?, ?, ?)",
                        (session_id, summary, None)
                    )
                    self._db.commit()
                except Exception as db_err:
                    logger.warning(f"记录审批历史失败: {db_err}")
                await ws_manager.broadcast({
                    "type": "approval.request",
                    "payload": {
                        "session_id": session_id,
                        "plan": data.get("plan", {}),
                        "timestamp": asyncio.get_event_loop().time(),
                    },
                    "timestamp": asyncio.get_event_loop().time(),
                })
                logger.info(f"[API] 收到审批请求 [{session_id}]")
        except Exception as e:
            logger.warning(f"处理审批请求出错: {e}")

    async def _on_emotion_broadcast(self, msg):
        """接收情绪广播，实时转发给 WebSocket（含人格信息）"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            user_id = data.get("user_id", "web_user")
            
            # 获取当前人格
            from tent_os.services.emotion_service import EmotionService
            emotion_svc = EmotionService()
            persona = emotion_svc.get_persona(user_id)
            
            await ws_manager.broadcast({
                "type": "ai.emotion",
                "payload": {
                    "user_id": user_id,
                    "emotion": data.get("emotion", "listening"),
                    "source": data.get("source", "broadcast"),
                    "persona": persona,
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass
    
    async def _on_emotion_fused(self, msg):
        """接收融合情绪广播，转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            user_id = data.get("user_id", "web_user")
            
            await ws_manager.broadcast({
                "type": "emotion.fused",
                "payload": {
                    "user_id": user_id,
                    "session_id": data.get("session_id"),
                    "primary": data.get("primary"),
                    "intensity": data.get("intensity"),
                    "valence": data.get("valence"),
                    "mixed": data.get("mixed"),
                    "trend": data.get("trend"),
                    "authenticity": data.get("authenticity"),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_community_message(self, msg):
        """接收社区消息，实时转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "community.message",
                "payload": {
                    "id": data.get("id"),
                    "sender_id": data.get("sender_id"),
                    "sender_name": data.get("sender_name"),
                    "avatar": data.get("avatar"),
                    "content": data.get("content"),
                    "channel": data.get("channel"),
                    "created_at": data.get("created_at"),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_community_task_update(self, msg):
        """接收社区任务更新，实时转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "community.task_update",
                "payload": {
                    "task_id": data.get("task_id"),
                    "status": data.get("status"),
                    "progress": data.get("progress"),
                    "agent_id": data.get("agent_id"),
                    "agent_name": data.get("agent_name"),
                    "message": data.get("message"),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_house_event(self, msg):
        """接收家园事件，实时转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "house.event",
                "payload": {
                    "event": data.get("event"),
                    "item_type": data.get("item_type"),
                    "item_id": data.get("item_id"),
                    "data": data.get("data"),
                    "source": data.get("source"),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_social_friend_request(self, msg):
        """Phase 2: 好友申请实时转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "social.friend_request",
                "payload": {
                    "from_ai_id": data.get("from_ai_id"),
                    "to_ai_id": data.get("to_ai_id"),
                    "timestamp": data.get("timestamp"),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_social_friend_accepted(self, msg):
        """Phase 2: 好友接受实时转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "social.friend_accepted",
                "payload": {
                    "from_ai_id": data.get("from_ai_id"),
                    "to_ai_id": data.get("to_ai_id"),
                    "friendship_id": data.get("friendship_id"),
                    "timestamp": data.get("timestamp"),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_social_visit_request(self, msg):
        """Phase 2: 串门请求实时转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "social.visit.request",
                "payload": {
                    "from_ai_id": data.get("from_ai_id"),
                    "to_ai_id": data.get("to_ai_id"),
                    "timestamp": data.get("timestamp"),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_social_visit_accepted(self, msg):
        """Phase 2: 串门接受实时转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "social.visit.accepted",
                "payload": {
                    "visit_id": data.get("visit_id"),
                    "to_ai_id": data.get("to_ai_id"),
                    "timestamp": data.get("timestamp"),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_social_visit_rejected(self, msg):
        """Phase 2: 串门拒绝实时转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "social.visit.rejected",
                "payload": {
                    "visit_id": data.get("visit_id"),
                    "to_ai_id": data.get("to_ai_id"),
                    "timestamp": data.get("timestamp"),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_system_health_alert(self, msg):
        """P0: 接收系统健康告警，转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "system.health_alert",
                "payload": {
                    "alert_type": data.get("alert_type"),
                    "severity": data.get("severity", "warning"),
                    "metric": data.get("metric"),
                    "value": data.get("value"),
                    "message": data.get("message"),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_physical_status_change(self, msg):
        """P0: 接收物理执行状态变化，转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "physical.status_change",
                "payload": {
                    "task_id": data.get("task_id"),
                    "status": data.get("status"),
                    "provider": data.get("provider"),
                    "action": data.get("action"),
                    "target_location": data.get("target_location"),
                    "event": data.get("event"),
                    "fallback_from": data.get("fallback_from"),
                    "fallback_to": data.get("fallback_to"),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_vision_perception(self, msg):
        """P0: 接收视觉感知事件，转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "vision.perception",
                "payload": {
                    "event_type": data.get("event_type"),
                    "user_id": data.get("user_id", "web_user"),
                    "user_detected": data.get("user_detected", False),
                    "user_emotion": data.get("user_emotion"),
                    "detected_objects": data.get("detected_objects", []),
                    "scene": data.get("scene"),
                    "confidence": data.get("confidence", 0),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_stream_chunk(self, msg):
        """接收 LLM 流式输出 chunk，实时转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "chat.stream_chunk",
                "payload": data,
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_reasoning_chunk(self, msg):
        """接收 LLM reasoning 流式输出，实时转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "chat.reasoning_chunk",
                "payload": data,
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _on_monologue_chunk(self, msg):
        """Phase 5: 接收内心独白流式输出，实时转发给 WebSocket"""
        try:
            data = json.loads(msg.data.decode() if isinstance(msg.data, bytes) else msg.data)
            await ws_manager.broadcast({
                "type": "ai.monologue",
                "payload": data,
                "timestamp": asyncio.get_event_loop().time(),
            })
        except Exception:
            pass

    async def _query_tasks_by_session(self, session_id: str) -> List[sqlite3.Row]:
        """按 session_id 查询所有关联任务，按时间倒序
        FIX v5: 异步包装，避免阻塞事件循环"""
        if not self._db:
            return []
        def _exec():
            cursor = self._db.execute(
                "SELECT * FROM tasks WHERE session_id = ? ORDER BY created_at DESC",
                (session_id,)
            )
            return cursor.fetchall()
        return await asyncio.to_thread(_exec)

    async def _query_all_tasks(self, limit: int = 50) -> List[sqlite3.Row]:
        """查询最近任务
        FIX v5: 异步包装，避免阻塞事件循环"""
        if not self._db:
            return []
        def _exec():
            cursor = self._db.execute(
                "SELECT DISTINCT session_id, status, action as task, created_at, updated_at "
                "FROM tasks WHERE session_id IS NOT NULL ORDER BY created_at DESC LIMIT ?",
                (limit,)
            )
            return cursor.fetchall()
        return await asyncio.to_thread(_exec)

    def _get_db_path(self) -> str:
        return self.config.get("scheduler", {}).get("db_path", "./tent_scheduler.db") if self.config else "./tent_scheduler.db"


state = APIServerState()


# ========== FastAPI Lifespan ==========

async def _avatar_state_broadcaster():
    """P0: 定期广播 avatar 系统状态（感知融合层）
    
    每 5 秒收集一次系统状态并广播给所有 WebSocket 客户端，
    让前端 avatar 能感知到系统健康、物理任务、视觉感知等状态。
    """
    await asyncio.sleep(3)  # 等待 setup 完成
    while True:
        try:
            # 检查是否有任何 WebSocket 连接
            has_connections = False
            try:
                if hasattr(ws_manager, '_sessions') and ws_manager._sessions:
                    has_connections = any(bool(v) for v in ws_manager._sessions.values())
                elif hasattr(ws_manager, 'connections') and ws_manager.connections:
                    has_connections = True
            except Exception:
                pass
            if not has_connections:
                await asyncio.sleep(5)
                continue
            
            # 收集系统健康状态
            health = await _get_health_payload()
            
            # 收集物理任务状态（从数据库查询）
            physical_tasks = []
            try:
                if state._db:
                    def _exec():
                        cursor = state._db.execute(
                            "SELECT task_id, status, provider, action, target_location, created_at FROM physical_tasks WHERE status IN ('submitted', 'assigned', 'executing') ORDER BY created_at DESC LIMIT 5"
                        )
                        return cursor.fetchall()
                    rows = await asyncio.to_thread(_exec)
                    physical_tasks = [
                        {
                            "task_id": r["task_id"] if hasattr(r, "keys") else r[0],
                            "status": r["status"] if hasattr(r, "keys") else r[1],
                            "provider": r["provider"] if hasattr(r, "keys") else r[2],
                            "action": r["action"] if hasattr(r, "keys") else r[3],
                            "target_location": r["target_location"] if hasattr(r, "keys") else r[4],
                        }
                        for r in rows
                    ]
            except Exception:
                pass
            
            # 收集最近任务负载
            task_load = {"total_recent": 0, "completed_recent": 0, "failed_recent": 0}
            try:
                if state._db:
                    def _exec():
                        cursor = state._db.execute(
                            "SELECT status, COUNT(*) as cnt FROM tasks WHERE created_at > datetime('now', '-1 hour') GROUP BY status"
                        )
                        return cursor.fetchall()
                    rows = await asyncio.to_thread(_exec)
                    for r in rows:
                        status = r["status"] if hasattr(r, "keys") else r[0]
                        cnt = r["cnt"] if hasattr(r, "keys") else r[1]
                        if status in ("completed", "done"):
                            task_load["completed_recent"] = cnt
                        elif status in ("failed", "error"):
                            task_load["failed_recent"] = cnt
                        task_load["total_recent"] += cnt
            except Exception:
                pass
            
            # 广播系统状态给 avatar
            await ws_manager.broadcast({
                "type": "avatar.system_state",
                "payload": {
                    "health": health,
                    "physical_tasks": physical_tasks,
                    "task_load": task_load,
                    "timestamp": asyncio.get_event_loop().time(),
                },
                "timestamp": asyncio.get_event_loop().time(),
            })
            
        except Exception as e:
            logger.debug(f"[AvatarBroadcaster] 广播失败: {e}")
        
        await asyncio.sleep(5)


@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用生命周期管理"""
    config = load_config(state.config_path if hasattr(state, "config_path") else "./config/tent_os.yaml")
    # 1. 初始化 server 的 state（处理旧逻辑：NATS 订阅、审批、健康检查等）
    await state.setup(config)
    # 2. 初始化 soul_state 的 state（Phase 1/2 重构组件：AgentLoop、Security、Hooks 等）
    #    这是独立的实例，专供 _handle_fast_chat 使用
    try:
        await soul_state_ref.state.setup(config)
        logger.info("[AGENT] soul_state 已初始化（Phase 1/2 组件就绪）")
    except Exception as e:
        logger.warning(f"[AGENT] soul_state 初始化失败: {e}")
    # [SOUL] 启动 avatar 状态广播器（简化版）
    broadcaster_task = asyncio.create_task(_avatar_state_broadcaster())
    # [SOUL] AI社会/社区/2D庄园/物理世界模块已移除
    yield
    broadcaster_task.cancel()
    try:
        await broadcaster_task
    except asyncio.CancelledError:
        pass
    await state.cleanup()


app = FastAPI(title="Tent OS API", version="0.1.0", lifespan=lifespan)

# [SOUL] Skills/World/MCP Router 已移除 —— 专注灵魂对讲机核心
# app.include_router(skills_router)
# app.include_router(world_router)
# app.include_router(mcp_router)

# 挂载 soul_routes 的 REST API（Phase 1/2 重构后的新路径）
from tent_os.api import soul_routes
app.include_router(soul_routes.router)


# ========== WebSocket 端点 ==========

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    """Control UI WebSocket 端点

    协议：JSON 消息，type + payload 格式
    客户端 -> 服务器：task.submit, task.abort, ping
    服务器 -> 客户端：task.plan, task.step, task.completed, task.failed, system.health, pong
    """
    await ws_manager.connect(websocket)
    try:
        # 发送初始健康状态
        await ws_manager.send_to(websocket, {
            "type": "system.health",
            "payload": await _get_health_payload(),
            "timestamp": asyncio.get_event_loop().time(),
        })

        while True:
            raw = await websocket.receive_text()
            try:
                msg = json.loads(raw)
                msg_type = msg.get("type")
                payload = msg.get("payload", {})

                if msg_type == "ping":
                    await ws_manager.send_to(websocket, {
                        "type": "pong",
                        "payload": {},
                        "timestamp": asyncio.get_event_loop().time(),
                    })

                elif msg_type == "chat.message":
                    session_id = payload.get("session_id", f"ws_{uuid.uuid4().hex[:12]}")
                    content = payload.get("content", "")
                    images = payload.get("images", [])
                    tools = payload.get("tools", [])
                    deep_thinking = bool(payload.get("deep_thinking"))
                    user_id = payload.get("user_id", "web_user")
                    
                    # 关联 WebSocket 到 session（用于 session-scoped 消息推送）
                    await ws_manager.associate_session(websocket, session_id)
                    
                    # FIX: 新消息开始时重置 abort 标志，避免上次中止污染本次对话
                    try:
                        if state.state_store:
                            await state.state_store.update(session_id, {"abort_requested": False})
                    except Exception:
                        pass
                    
                    state._pending_count[session_id] = state._pending_count.get(session_id, 0) + 1
                    
                    # Phase 2: 直接调用 AgentLoop（替代 NATS -> 治理进程的间接路径）
                    capabilities = {}
                    if isinstance(tools, dict):
                        capabilities["web_search"] = bool(tools.get("web_search"))
                        capabilities["file_ops"] = bool(tools.get("file_ops"))
                    elif tools:
                        capabilities["web_search"] = "web_search" in tools
                        capabilities["file_ops"] = "file_ops" in tools
                    
                    # 发送确认
                    await ws_manager.send_to(websocket, {
                        "type": "chat.message_accepted",
                        "payload": {"session_id": session_id, "has_images": bool(images)},
                        "timestamp": asyncio.get_event_loop().time(),
                    })
                    
                    # 直接调用 AgentLoop 处理对话
                    try:
                        await _handle_fast_chat(
                            session_id=session_id,
                            user_id=user_id,
                            content=content,
                            websocket=websocket,
                            capabilities=capabilities,
                            deep_thinking=deep_thinking,
                        )
                    except Exception as e:
                        logger.error(f"[WS] AgentLoop 处理失败 [{session_id}]: {e}", exc_info=True)
                        await ws_manager.send_to(websocket, {
                            "type": "chat.completed",
                            "payload": {
                                "content": "抱歉，处理过程中出现了问题，请稍后再试。",
                                "session_id": session_id,
                                "elapsed_ms": 0,
                            },
                            "timestamp": asyncio.get_event_loop().time(),
                        })
                    finally:
                        state._pending_count[session_id] = max(0, state._pending_count.get(session_id, 1) - 1)

                elif msg_type == "chat.session.list":
                    user_id = payload.get("user_id", "web_user")
                    sessions = await state.state_store.list_sessions(user_id, limit=50)
                    # 过滤掉 heartbeat 内部会话，不显示在 UI
                    sessions = [
                        s for s in sessions
                        if not (s.get("title", "").startswith("[Heartbeat]") or s.get("session_id", "").startswith("hb_"))
                    ]
                    await ws_manager.send_to(websocket, {
                        "type": "chat.session.list",
                        "payload": {"sessions": sessions},
                        "timestamp": asyncio.get_event_loop().time(),
                    })

                elif msg_type == "chat.session.create":
                    session_id = f"ws_{uuid.uuid4().hex[:12]}"
                    user_id = payload.get("user_id", "web_user")
                    await state.state_store.create(session_id=session_id, task="", user_id=user_id)
                    await ws_manager.send_to(websocket, {
                        "type": "chat.session.created",
                        "payload": {"session_id": session_id},
                        "timestamp": asyncio.get_event_loop().time(),
                    })

                elif msg_type == "chat.history":
                    session_id = payload.get("session_id")
                    messages = await state.state_store.get_messages(session_id, limit=100)
                    await ws_manager.send_to(websocket, {
                        "type": "chat.history",
                        "payload": {"session_id": session_id, "messages": messages},
                        "timestamp": asyncio.get_event_loop().time(),
                    })

                elif msg_type == "chat.session.load":
                    session_id = payload.get("session_id")
                    user_id = payload.get("user_id", "web_user")
                    if session_id and state.state_store:
                        try:
                            hist = await state.state_store.get_messages(session_id, limit=50)
                            await ws_manager.send_to(websocket, {
                                "type": "chat.session.loaded",
                                "payload": {
                                    "session_id": session_id,
                                    "messages": [
                                        {"role": m.get("role", ""), "content": m.get("content", ""), "reasoning": m.get("reasoning", "")}
                                        for m in hist if m.get("role")
                                    ],
                                },
                                "timestamp": asyncio.get_event_loop().time(),
                            })
                        except Exception as e:
                            logger.warning(f"[WS] 加载会话失败 [{session_id}]: {e}")
                            await ws_manager.send_to(websocket, {
                                "type": "chat.session.loaded",
                                "payload": {"session_id": session_id, "messages": [], "error": str(e)},
                                "timestamp": asyncio.get_event_loop().time(),
                            })

                elif msg_type == "task.submit":
                    # 兼容旧版任务提交
                    session_id = payload.get("session_id", f"ws_{uuid.uuid4().hex[:12]}")
                    task = payload.get("task", "")
                    tools = payload.get("tools", [])
                    await state.bus.publish("governance.request", json.dumps({
                        "session_id": session_id,
                        "user_id": "web_user",
                        "content": task,
                        "tools": tools
                    }).encode())
                    await ws_manager.send_to(websocket, {
                        "type": "task.submit",
                        "payload": {"session_id": session_id, "status": "accepted"},
                        "timestamp": asyncio.get_event_loop().time(),
                    })

                elif msg_type in ("chat.abort", "task.abort"):
                    session_id = payload.get("session_id")
                    # FIX: 实现真正的 abort 机制——设置 abort 标志，Tool Loop 检测到后优雅退出
                    try:
                        if state.state_store:
                            await state.state_store.update(session_id, {"abort_requested": True})
                        await ws_manager.send_to(websocket, {
                            "type": "chat.aborted",
                            "payload": {"session_id": session_id, "status": "aborting"},
                            "timestamp": asyncio.get_event_loop().time(),
                        })
                        logger.info(f"[WS] 任务中止请求已发送 [{session_id}]")
                    except Exception as e:
                        logger.warning(f"[WS] 任务中止失败 [{session_id}]: {e}")
                        await ws_manager.send_to(websocket, {
                            "type": "chat.aborted",
                            "payload": {"session_id": session_id, "status": "error", "error": str(e)},
                            "timestamp": asyncio.get_event_loop().time(),
                        })

                elif msg_type == "vision.blendshapes":
                    # Phase 1: 接收 FaceLandmarker blendshapes，进行精准情绪分析
                    user_id = payload.get("user_id", "frank")
                    blendshapes = payload.get("blendshapes", {})
                    try:
                        from tent_os.services.emotion_analyzer import get_analyzer
                        analyzer = get_analyzer()
                        result = analyzer.analyze(blendshapes, user_id)

                        # 更新六维成长（AWARENESS维度）
                        try:
                            from tent_os.services.six_axis_service import SixAxisService
                            SixAxisService.update_by_task_action(
                                user_id, "emotion_detected",
                                {"complexity": 1.0 + result.confidence}
                            )
                        except Exception:
                            pass

                        # FIX Phase 5: 生成视觉观察摘要并存储（供 LLM 注入使用）
                        try:
                            from tent_os.services.emotion_service import EmotionService
                            emotion_svc = EmotionService()
                            
                            # 构建视觉观察摘要
                            fatigue_desc = "状态良好"
                            if result.fatigue_level > 0.7:
                                fatigue_desc = "非常疲惫"
                            elif result.fatigue_level > 0.4:
                                fatigue_desc = "有些疲惫"
                            
                            attention_desc = ""
                            # 如果有多帧数据，检测注意力
                            gaze_score = result.scores.get("gaze", 0.5)
                            if gaze_score < 0.3:
                                attention_desc = "，注意力似乎不在对话上"
                            
                            visual_summary = f"用户看起来{fatigue_desc}，表情{result.primary}"
                            if attention_desc:
                                visual_summary += attention_desc
                            
                            emotion_svc.set_visual_summary(user_id, visual_summary)
                            logger.debug(f"[WS] 视觉摘要已更新 [{user_id}]: {visual_summary}")
                        except Exception as e:
                            logger.debug(f"[WS] 视觉摘要生成失败: {e}")
                        
                        # P0: 广播视觉感知事件（供 avatar 感知用户存在和情绪）
                        try:
                            await ws_manager.broadcast({
                                "type": "vision.perception",
                                "payload": {
                                    "event_type": "user_emotion_detected",
                                    "user_id": user_id,
                                    "user_detected": True,
                                    "user_emotion": result.primary,
                                    "confidence": result.confidence,
                                    "fatigue": result.fatigue_level,
                                    "detected_objects": [],
                                },
                                "timestamp": asyncio.get_event_loop().time(),
                            })
                        except Exception as e:
                            logger.debug(f"[WS] 视觉感知广播失败: {e}")

                        # 视觉情绪 → AI 共情反应（情绪双轨合并）
                        try:
                            from tent_os.services.emotion_service import EmotionService
                            emotion_svc = EmotionService()
                            ai_emotion = emotion_svc.update_by_vision(user_id, result.primary)
                            persona = emotion_svc.get_persona(user_id)

                            # Layer 5: 广播节流——只有情绪变化才广播
                            last_emotion = state._last_vision_emotion.get(user_id)
                            if ai_emotion != last_emotion:
                                state._last_vision_emotion[user_id] = ai_emotion
                                await ws_manager.broadcast({
                                    "type": "ai.emotion",
                                    "payload": {"user_id": user_id, "emotion": ai_emotion, "source": f"vision:{result.primary}", "persona": persona},
                                    "timestamp": asyncio.get_event_loop().time(),
                                })
                        except Exception as e:
                            logger.debug(f"[WS] 视觉情绪共情广播失败: {e}")

                        # 推送情绪分析结果给前端
                        await ws_manager.send_to(websocket, {
                            "type": "emotion.update",
                            "payload": {
                                "emotion": result.primary,
                                "user_emotion": result.primary,
                                "score": result.primary_score,
                                "confidence": result.confidence,
                                "fatigue": result.fatigue_level,
                                "all_scores": result.scores,
                            },
                            "timestamp": asyncio.get_event_loop().time(),
                        })
                    except Exception as e:
                        logger.warning(f"[WS] Blendshapes情绪分析失败: {e}")
                        await ws_manager.send_to(websocket, {
                            "type": "emotion.update",
                            "payload": {"emotion": "neutral", "score": 0, "confidence": 0, "error": str(e)},
                            "timestamp": asyncio.get_event_loop().time(),
                        })

                elif msg_type == "vision.emotion_detected":
                    # 兼容旧版简单情绪检测（前端老版本仍然可用）
                    user_id = payload.get("user_id", "web_user")
                    emotion = payload.get("emotion", "neutral")
                    confidence = payload.get("confidence", 0.5)
                    try:
                        from tent_os.services.six_axis_service import SixAxisService
                        SixAxisService.update_by_task_action(
                            user_id, "emotion_detected",
                            {"complexity": 1.0 + confidence}
                        )
                        logger.info(f"[WS] 视觉情绪检测(legacy): {user_id} → {emotion} ({confidence:.2f})")
                    except Exception as e:
                        logger.warning(f"[WS] 视觉情绪处理失败: {e}")
                    await ws_manager.send_to(websocket, {
                        "type": "vision.emotion_detected",
                        "payload": {"emotion": emotion, "received": True},
                        "timestamp": asyncio.get_event_loop().time(),
                    })

                elif msg_type == "vision.scene_frame":
                    # Phase 1: 接收场景截图 → VLM 分析 → 自动存入视觉记忆
                    user_id = payload.get("user_id", "frank")
                    image_data = payload.get("image_data", "")
                    resolution = payload.get("resolution", {})
                    try:
                        llm = getattr(state, '_vision_llm', None) or getattr(state, '_llm', None)
                        if llm and image_data:
                            # 异步分析，不阻塞 WS
                            asyncio.create_task(_analyze_scene_frame(user_id, image_data, resolution, websocket))
                    except Exception as e:
                        logger.warning(f"[WS] 场景帧处理失败: {e}")

                elif msg_type == "tts.enabled":
                    # FIX Phase 5: TTS 开启通知 → 存储到 state 供后端读取
                    user_id = payload.get("user_id", "web_user")
                    try:
                        if state.state_store:
                            sessions = await state.state_store.list_sessions(user_id, limit=1)
                            if sessions:
                                sid = sessions[0].get("session_id")
                                if sid:
                                    await state.state_store.update(sid, {"tts_enabled": True})
                                    logger.debug(f"[WS] TTS 已启用 [{sid}]")
                    except Exception as e:
                        logger.debug(f"[WS] TTS 启用处理失败: {e}")
                    await ws_manager.send_to(websocket, {
                        "type": "tts.enabled",
                        "payload": {"received": True},
                        "timestamp": asyncio.get_event_loop().time(),
                    })
                
                elif msg_type == "tts.disabled":
                    # FIX Phase 5: TTS 关闭通知
                    user_id = payload.get("user_id", "web_user")
                    try:
                        if state.state_store:
                            sessions = await state.state_store.list_sessions(user_id, limit=1)
                            if sessions:
                                sid = sessions[0].get("session_id")
                                if sid:
                                    await state.state_store.update(sid, {"tts_enabled": False})
                                    logger.debug(f"[WS] TTS 已禁用 [{sid}]")
                    except Exception as e:
                        logger.debug(f"[WS] TTS 禁用处理失败: {e}")
                    await ws_manager.send_to(websocket, {
                        "type": "tts.disabled",
                        "payload": {"received": True},
                        "timestamp": asyncio.get_event_loop().time(),
                    })
                
                elif msg_type == "tts.status":
                    # FIX Phase 5: 前端统一 TTS 状态上报
                    user_id = payload.get("user_id", "web_user")
                    tts_enabled = payload.get("enabled", False)
                    try:
                        if state.state_store:
                            sessions = await state.state_store.list_sessions(user_id, limit=1)
                            if sessions:
                                sid = sessions[0].get("session_id")
                                if sid:
                                    await state.state_store.update(sid, {"tts_enabled": tts_enabled})
                                    logger.debug(f"[WS] TTS 状态更新 [{sid}]: enabled={tts_enabled}")
                    except Exception as e:
                        logger.debug(f"[WS] TTS 状态更新失败: {e}")
                    await ws_manager.send_to(websocket, {
                        "type": "tts.status",
                        "payload": {"received": True, "enabled": tts_enabled},
                        "timestamp": asyncio.get_event_loop().time(),
                    })

                elif msg_type == "voice.prosody":
                    # Phase 1: 接收语音韵律特征 → 传递给情绪融合引擎
                    user_id = payload.get("user_id", "web_user")
                    prosody = payload.get("prosody", {})
                    try:
                        from tent_os.services.emotion_service import EmotionService
                        emotion_svc = EmotionService()
                        emotion_svc.record_voice_prosody(user_id, prosody)
                        logger.debug(f"[WS] 语音韵律已记录 [{user_id}]: pitch_var={prosody.get('pitch_variation')}, rate={prosody.get('speech_rate')}")
                    except Exception as e:
                        logger.debug(f"[WS] 语音韵律记录失败: {e}")
                    await ws_manager.send_to(websocket, {
                        "type": "voice.prosody",
                        "payload": {"received": True},
                        "timestamp": asyncio.get_event_loop().time(),
                    })
                
                elif msg_type == "avatar.pet":
                    # Phase 5: 抚摸互动 → AI 情绪变为 happy
                    user_id = payload.get("user_id", "web_user")
                    try:
                        from tent_os.services.emotion_service import EmotionService
                        emotion_svc = EmotionService()
                        emotion_svc.set_emotion(user_id, "happy")
                        await ws_manager.broadcast({
                            "type": "ai.emotion",
                            "payload": {"emotion": "happy", "user_id": user_id, "source": "pet"},
                            "timestamp": asyncio.get_event_loop().time(),
                        })
                        await ws_manager.send_to(websocket, {
                            "type": "avatar.pet",
                            "payload": {"status": "ok", "emotion": "happy"},
                            "timestamp": asyncio.get_event_loop().time(),
                        })
                    except Exception as e:
                        logger.warning(f"[WS] 抚摸处理失败: {e}")
                
                elif msg_type == "vision.objects_detected":
                    # Phase 2: 接收物体检测结果 → 更新空间索引
                    user_id = payload.get("user_id", "frank")
                    objects = payload.get("objects", [])
                    try:
                        from tent_os.services.visual_memory_service import get_visual_memory_service
                        vm = get_visual_memory_service()
                        # 更新物体清单（不存储图片，只记录物体存在）
                        for obj in objects:
                            name = obj.get("name", "").strip().lower()
                            if not name:
                                continue
                            norm = obj.get("normalized", {})
                            location = f"画面({norm.get('x',0):.0%},{norm.get('y',0):.0%})"
                            vm.store_memory(
                                user_id=user_id,
                                image_url="",
                                description=f"检测到物体: {name} (置信度:{obj.get('confidence',0):.0%})",
                                scene_type="object_detection",
                                objects=[{"name": name, "location": location, "confidence": obj.get("confidence", 0.5)}],
                            )
                        # 手眼联动安全规则：检测到人员靠近 + 有 active physical 任务 → 记录安全事件
                        try:
                            high_conf_person = any(
                                o.get("name", "").strip().lower() == "person" and o.get("confidence", 0) > 0.7
                                for o in objects
                            )
                            if high_conf_person and state._db:
                                cursor = state._db.execute(
                                    "SELECT COUNT(*) as cnt FROM physical_tasks WHERE status IN ('submitted', 'assigned', 'executing')"
                                )
                                row = cursor.fetchone()
                                if row and row["cnt"] > 0:
                                    logger.warning(f"[SAFETY] 视觉检测到人员靠近，同时有 {row['cnt']} 个物理任务在执行中")
                                    vm.store_memory(
                                        user_id=user_id,
                                        image_url="",
                                        description=f"[安全警报] 检测到人员靠近，{row['cnt']} 个物理任务正在执行",
                                        scene_type="safety_alert",
                                        objects=[{"name": "person", "location": "unknown", "confidence": 1.0}],
                                    )
                        except Exception:
                            pass
                        # FIX: 广播物体检测结果给所有客户端（PhysicalWorldPanel 等面板可实时接收）
                        await ws_manager.broadcast({
                            "type": "vision.objects_detected",
                            "payload": {
                                "user_id": user_id,
                                "objects": objects,
                                "count": len(objects),
                            },
                            "timestamp": asyncio.get_event_loop().time(),
                        })
                        logger.debug(f"[WS] 物体检测更新: {user_id} → {len(objects)}个物体")
                    except Exception as e:
                        logger.warning(f"[WS] 物体检测处理失败: {e}")

                elif msg_type == "ping":
                    await ws_manager.send_to(websocket, {
                        "type": "pong",
                        "payload": {},
                        "timestamp": asyncio.get_event_loop().time(),
                    })

            except json.JSONDecodeError:
                await ws_manager.send_to(websocket, {
                    "type": "error",
                    "payload": {"message": "Invalid JSON"},
                    "timestamp": asyncio.get_event_loop().time(),
                })
    except WebSocketDisconnect:
        await ws_manager.disconnect(websocket)
    except Exception as e:
        logger.warning(f"WebSocket error: {e}")
        await ws_manager.disconnect(websocket)


async def _get_health_payload() -> Dict:
    bus_connected = state.bus is not None and hasattr(state.bus.nats, "is_connected") and state.bus.nats.is_connected
    redis_connected = False
    if state.state_store:
        try:
            redis_connected = hasattr(state.state_store, "ping") and await state.state_store.ping()
        except Exception:
            pass
    return {
        "status": "ok" if bus_connected else "degraded",
        "natsConnected": bus_connected,
        "redisConnected": redis_connected,
        "workers": {
            "memory": bus_connected,  # 简化：NATS 通则认为 workers 存活
            "governance": bus_connected,
            "scheduler": bus_connected,
        },
        "version": "0.1.0",
    }


# 查找 dist 目录（构建后的前端文件）
_UI_DIST_DIR = Path(__file__).parent.parent.parent / "frontend" / "desktop" / "dist"


@app.get("/", include_in_schema=False)
async def root_redirect():
    """根路径重定向到 Control UI"""
    return RedirectResponse(url="/ui/")


# ========== UI 数据 API ==========

@app.get("/ui/api/dreaming/status")
async def ui_dreaming_status():
    """UI 梦境引擎状态"""
    from tent_os.autonomy.dreaming import DreamingEngine
    # 尝试从全局获取 dreaming 实例
    dreaming = getattr(state, "_dreaming", None)
    if dreaming:
        return dreaming.get_status()
    return {
        "enabled": False,
        "is_dreaming": False,
        "current_dream_id": None,
        "schedule": "0 2 * * *",
        "depth": 3,
        "stats": {"total_dreams": 0, "total_memories_processed": 0,
                  "total_rules_extracted": 0, "total_contradictions_found": 0},
    }


@app.get("/ui/api/dreaming/diary")
async def ui_dreaming_diary(limit: int = 20):
    """UI 梦境日记"""
    from tent_os.autonomy.dreaming import DreamDiary
    db_path = state._get_db_path()
    diary = DreamDiary(db_path)
    return {"dreams": diary.get_recent(limit)}


@app.post("/ui/api/dreaming/toggle")
async def ui_dreaming_toggle(req: dict):
    """UI 开关梦境模式
    
    Request body: {"enabled": true/false}
    """
    enabled = req.get("enabled", True)
    dreaming = getattr(state, "_dreaming", None)
    if dreaming:
        dreaming.toggle(enabled)
    return {"enabled": enabled}


@app.get("/ui/api/dreaming/entries")
async def ui_dreaming_entries(limit: int = 10):
    """获取梦境条目（用于 2D 世界梦境气泡渲染）"""
    from tent_os.autonomy.dreaming import DreamDiary
    db_path = state._get_db_path()
    diary = DreamDiary(db_path)
    dreams = diary.get_recent(limit)
    entries: list = []
    for d in dreams:
        for e in d.get("entries", []):
            desc = e.get("description", "")
            if desc:
                entries.append(desc[:20])
    unique = list(dict.fromkeys(entries))[:20]
    return {"entries": unique if unique else ['记忆碎片', '灵感火花', '时空回溯']}

@app.post("/ui/api/dreaming/trigger")
async def ui_dreaming_trigger():
    """UI 手动触发一次梦境"""
    from tent_os.autonomy.dreaming import DreamingEngine
    dreaming = getattr(state, "_dreaming", None)
    if dreaming:
        dream_id = await dreaming.trigger_now()
        return {"triggered": True, "dream_id": dream_id}
    return {"triggered": False, "error": "Dreaming engine not available"}


@app.get("/ui/api/health")
async def ui_health():
    """UI 健康状态"""
    return await _get_health_payload()


@app.get("/ui/api/tasks")
async def ui_tasks(limit: int = 50):
    """UI 任务列表"""
    rows = await state._query_all_tasks(limit)
    tasks = []
    for row in rows:
        r = dict(row)
        tasks.append({
            "session_id": r.get("session_id", ""),
            "status": r.get("status", "unknown"),
            "task": r.get("task", ""),
            "created_at": r.get("created_at", ""),
            "updated_at": r.get("updated_at", ""),
        })
    return {"tasks": tasks, "count": len(tasks)}


@app.get("/ui/api/world/calendar")
async def ui_world_calendar(month: str = ""):
    """日历墙数据 —— 按日期聚合任务，供 AI 庄园的日历墙展示
    
    Args:
        month: YYYY-MM 格式，如 "2026-05"。为空时返回最近 30 天。
    """
    import datetime as dt
    
    if not month:
        now = dt.datetime.now()
        month = now.strftime("%Y-%m")
    
    try:
        year, mon = int(month.split("-")[0]), int(month.split("-")[1])
    except Exception:
        year, mon = dt.datetime.now().year, dt.datetime.now().month
    
    # 计算月份起止
    start = dt.datetime(year, mon, 1)
    if mon == 12:
        end = dt.datetime(year + 1, 1, 1)
    else:
        end = dt.datetime(year, mon + 1, 1)
    
    start_str = start.strftime("%Y-%m-%d")
    end_str = end.strftime("%Y-%m-%d")
    
    calendar_days = {}
    
    try:
        def _exec():
            if not state._db:
                return []
            cursor = state._db.execute(
                "SELECT session_id, status, action as task, created_at, updated_at "
                "FROM tasks WHERE created_at >= ? AND created_at < ? AND session_id IS NOT NULL "
                "ORDER BY created_at DESC",
                (start_str, end_str)
            )
            return cursor.fetchall()
        
        rows = await asyncio.to_thread(_exec)
        
        for row in rows:
            r = dict(row)
            created = r.get("created_at", "")
            if not created:
                continue
            # 提取日期部分 YYYY-MM-DD
            day_key = created[:10] if len(created) >= 10 else created
            if day_key not in calendar_days:
                calendar_days[day_key] = []
            calendar_days[day_key].append({
                "session_id": r.get("session_id", ""),
                "status": r.get("status", "unknown"),
                "task": r.get("task", ""),
                "created_at": created,
            })
    except Exception as e:
        logger.warning(f"[CALENDAR] 查询失败: {e}")
    
    # 生成日历网格（包含空日期）
    days_in_month = (end - start).days
    first_weekday = start.weekday()  # 0=周一
    
    grid = []
    for i in range(first_weekday):
        grid.append(None)  # 上月填充
    for d in range(1, days_in_month + 1):
        day_str = f"{year:04d}-{mon:02d}-{d:02d}"
        grid.append({
            "date": day_str,
            "day": d,
            "tasks": calendar_days.get(day_str, []),
            "has_completed": any(t["status"] == "completed" for t in calendar_days.get(day_str, [])),
            "has_failed": any(t["status"] == "failed" for t in calendar_days.get(day_str, [])),
            "has_pending": any(t["status"] not in ("completed", "failed") for t in calendar_days.get(day_str, [])),
        })
    
    return {
        "month": month,
        "year": year,
        "mon": mon,
        "days_in_month": days_in_month,
        "first_weekday": first_weekday,
        "grid": grid,
        "total_tasks": sum(len(v) for v in calendar_days.values()),
    }


# ===== 冰箱贴（FridgeNotes）API =====

def _init_fridge_notes_table():
    """初始化冰箱贴表"""
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS fridge_notes (
                id TEXT PRIMARY KEY,
                content TEXT NOT NULL,
                color TEXT DEFAULT '#FFD54F',
                author TEXT DEFAULT 'user',
                created_at TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.commit()
        conn.close()
    except Exception as e:
        logger.warning(f"[FRIDGE] 初始化表失败: {e}")


@app.get("/ui/api/world/fridge-notes")
async def ui_fridge_notes():
    """获取所有冰箱贴便签"""
    _init_fridge_notes_table()
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        cursor = conn.execute(
            "SELECT id, content, color, author, created_at FROM fridge_notes ORDER BY created_at DESC"
        )
        rows = cursor.fetchall()
        conn.close()
        notes = []
        for row in rows:
            notes.append({
                "id": row[0],
                "content": row[1],
                "color": row[2],
                "author": row[3],
                "created_at": row[4],
            })
        return {"notes": notes, "count": len(notes)}
    except Exception as e:
        logger.warning(f"[FRIDGE] 查询失败: {e}")
        return {"notes": [], "count": 0}


class FridgeNoteCreate(BaseModel):
    content: str
    color: str = "#FFD54F"
    author: str = "user"


@app.post("/ui/api/world/fridge-notes")
async def ui_fridge_note_create(req: FridgeNoteCreate):
    """创建冰箱贴便签"""
    _init_fridge_notes_table()
    note_id = f"note_{int(time.time() * 1000)}"
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute(
            "INSERT INTO fridge_notes (id, content, color, author, created_at) VALUES (?, ?, ?, ?, datetime('now'))",
            (note_id, req.content, req.color, req.author)
        )
        conn.commit()
        conn.close()
        return {"id": note_id, "status": "created"}
    except Exception as e:
        logger.warning(f"[FRIDGE] 创建失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/ui/api/world/fridge-notes/{note_id}")
async def ui_fridge_note_delete(note_id: str):
    """删除冰箱贴便签"""
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute("DELETE FROM fridge_notes WHERE id = ?", (note_id,))
        conn.commit()
        conn.close()
        return {"id": note_id, "status": "deleted"}
    except Exception as e:
        logger.warning(f"[FRIDGE] 删除失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ========== 家园物品统一 API（Phase 1.5）==========

@app.get("/ui/api/house/items")
async def ui_house_items():
    """获取家园所有物品（冰箱贴、日历、信件、画框）"""
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        # FIX: sessions 表从未创建，日历和项目帧永远为空
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS sessions (
                session_id TEXT PRIMARY KEY,
                task TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status TEXT DEFAULT 'pending'
            )
        """)
        conn.commit()
        
        # 1. 冰箱贴
        cursor.execute("SELECT id, content, color, author, created_at FROM fridge_notes ORDER BY created_at DESC")
        fridge_notes = [
            {"id": r[0], "type": "fridge_note", "content": r[1], "color": r[2], "author": r[3], "created_at": r[4]}
            for r in cursor.fetchall()
        ]
        
        # 2. 日历事件（从已完成任务中提取）
        cursor.execute(
            "SELECT session_id, task, updated_at, status FROM sessions WHERE status IN ('completed', 'failed') ORDER BY updated_at DESC LIMIT 30"
        )
        calendar_events = []
        for r in cursor.fetchall():
            try:
                dt = datetime.fromisoformat(r[2].replace('Z', '+00:00'))
                day_key = dt.strftime('%Y-%m-%d')
            except Exception:
                day_key = r[2][:10] if r[2] else ""
            calendar_events.append({
                "id": f"cal_{r[0]}",
                "type": "calendar_event",
                "title": r[1][:30],
                "date": day_key,
                "status": r[3],
            })
        
        # 3. 信件架
        cursor.execute("SELECT id, type, title, content, is_read, created_at FROM letter_rack ORDER BY created_at DESC")
        letters = [
            {"id": r[0], "type": "letter", "letter_type": r[1], "title": r[2], "content": r[3], "is_read": bool(r[4]), "created_at": r[5]}
            for r in cursor.fetchall()
        ]
        
        # 4. 项目画框（从已完成任务中提取代表性项目）
        cursor.execute(
            "SELECT session_id, task, updated_at, status FROM sessions WHERE status = 'completed' ORDER BY updated_at DESC LIMIT 20"
        )
        projects = [
            {"id": f"proj_{r[0]}", "type": "project_frame", "title": r[1][:40], "date": r[2][:10] if r[2] else "", "status": r[3]}
            for r in cursor.fetchall()
        ]
        
        conn.close()
        
        return {
            "items": {
                "fridge_notes": fridge_notes,
                "calendar_events": calendar_events,
                "letters": letters,
                "project_frames": projects,
            },
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        logger.warning(f"[HOUSE] 获取家园物品失败: {e}")
        return {"items": {"fridge_notes": [], "calendar_events": [], "letters": [], "project_frames": []}, "timestamp": datetime.now().isoformat()}


@app.get("/ui/api/world/projects")
async def ui_world_projects(limit: int = 50):
    """画框墙数据 —— 已完成的项目，供 AI 庄园的画框展示
    
    从 tasks 表筛选 status='completed' 的任务，按完成时间倒序。
    """
    try:
        def _exec():
            if not state._db:
                return []
            cursor = state._db.execute(
                "SELECT session_id, status, action as task, created_at, updated_at, result "
                "FROM tasks WHERE status = 'completed' AND session_id IS NOT NULL "
                "ORDER BY updated_at DESC LIMIT ?",
                (limit,)
            )
            return cursor.fetchall()
        
        rows = await asyncio.to_thread(_exec)
        projects = []
        for row in rows:
            r = dict(row)
            created = r.get("created_at", "")
            updated = r.get("updated_at", "")
            # 计算执行时长（分钟）
            duration_min = 0
            try:
                c_dt = datetime.fromisoformat(created.replace('Z', '+00:00'))
                u_dt = datetime.fromisoformat(updated.replace('Z', '+00:00'))
                duration_min = max(1, int((u_dt - c_dt).total_seconds() / 60))
            except Exception:
                pass
            
            projects.append({
                "session_id": r.get("session_id", ""),
                "task": r.get("task", ""),
                "created_at": created,
                "updated_at": updated,
                "duration_min": duration_min,
                "result_preview": (r.get("result", "") or "")[:200],
            })
        
        return {
            "projects": projects,
            "count": len(projects),
            "total_completed": len(projects),
        }
    except Exception as e:
        logger.warning(f"[PROJECTS] 查询失败: {e}")
        return {"projects": [], "count": 0, "total_completed": 0}


# ===== 信件架（LetterRack）API =====

def _init_letter_rack_table():
    """初始化信件架表"""
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS letter_rack (
                id TEXT PRIMARY KEY,
                type TEXT DEFAULT 'system',
                title TEXT NOT NULL,
                content TEXT,
                read INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.commit()
        conn.close()
    except Exception as e:
        logger.warning(f"[LETTER] 初始化表失败: {e}")


@app.get("/ui/api/world/letter-rack")
async def ui_letter_rack(limit: int = 50, unread_only: bool = False):
    """获取信件架上的信件"""
    _init_letter_rack_table()
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        if unread_only:
            cursor = conn.execute(
                "SELECT id, type, title, content, read, created_at FROM letter_rack WHERE read = 0 ORDER BY created_at DESC LIMIT ?",
                (limit,)
            )
        else:
            cursor = conn.execute(
                "SELECT id, type, title, content, read, created_at FROM letter_rack ORDER BY created_at DESC LIMIT ?",
                (limit,)
            )
        rows = cursor.fetchall()
        conn.close()
        letters = []
        for row in rows:
            letters.append({
                "id": row[0],
                "type": row[1],
                "title": row[2],
                "content": row[3],
                "read": bool(row[4]),
                "created_at": row[5],
            })
        return {"letters": letters, "count": len(letters), "unread_count": sum(1 for l in letters if not l["read"])}
    except Exception as e:
        logger.warning(f"[LETTER] 查询失败: {e}")
        return {"letters": [], "count": 0, "unread_count": 0}


class LetterCreate(BaseModel):
    type: str = "system"
    title: str
    content: str = ""

class PropInteractionCreate(BaseModel):
    prop_id: str
    prop_type: str
    action: str
    room_id: str

class SpatialMemoryCreate(BaseModel):
    id: str | None = None
    room_id: str
    x: float
    y: float
    label: str
    memory_type: str = "event"
    description: str | None = None
    emotional_tag: str | None = None

class ObjectInventoryCreate(BaseModel):
    id: str | None = None
    room_id: str
    name: str
    object_type: str
    x: float
    y: float
    state: str = "active"
    detected_from: str | None = None

class UserDecorationCreate(BaseModel):
    id: str | None = None
    room_id: str
    decoration_type: str = "sticker"
    name: str = "装饰"
    x: float
    y: float
    size_w: float = 30
    size_h: float = 30
    color: str = "#94A3B8"


@app.post("/ui/api/world/letter-rack")
async def ui_letter_create(req: LetterCreate):
    """创建信件（通常由系统事件自动触发）"""
    _init_letter_rack_table()
    letter_id = f"letter_{int(time.time() * 1000)}"
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute(
            "INSERT INTO letter_rack (id, type, title, content, read, created_at) VALUES (?, ?, ?, ?, 0, datetime('now'))",
            (letter_id, req.type, req.title, req.content)
        )
        conn.commit()
        conn.close()
        return {"id": letter_id, "status": "created"}
    except Exception as e:
        logger.warning(f"[LETTER] 创建失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/ui/api/world/letter-rack/{letter_id}/read")
async def ui_letter_mark_read(letter_id: str):
    """标记信件为已读"""
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute("UPDATE letter_rack SET read = 1 WHERE id = ?", (letter_id,))
        conn.commit()
        conn.close()
        return {"id": letter_id, "status": "read"}
    except Exception as e:
        logger.warning(f"[LETTER] 标记已读失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/ui/api/world/letter-rack/{letter_id}")
async def ui_letter_delete(letter_id: str):
    """删除信件"""
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute("DELETE FROM letter_rack WHERE id = ?", (letter_id,))
        conn.commit()
        conn.close()
        return {"id": letter_id, "status": "deleted"}
    except Exception as e:
        logger.warning(f"[LETTER] 删除失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ===== 机制一-3: 道具交互 =====

def _init_prop_interactions_table():
    """初始化道具交互记录表"""
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prop_interactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            prop_id TEXT NOT NULL,
            prop_type TEXT NOT NULL,
            action TEXT NOT NULL,
            room_id TEXT NOT NULL,
            timestamp TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prop_interactions_prop_id ON prop_interactions(prop_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prop_interactions_timestamp ON prop_interactions(timestamp)")
    conn.commit()
    conn.close()

@app.post("/ui/api/world/prop-interaction")
async def ui_prop_interaction(req: PropInteractionCreate):
    """记录道具交互"""
    try:
        _init_prop_interactions_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute(
            "INSERT INTO prop_interactions (prop_id, prop_type, action, room_id) VALUES (?, ?, ?, ?)",
            (req.prop_id, req.prop_type, req.action, req.room_id)
        )
        conn.commit()
        conn.close()
        return {"status": "recorded"}
    except Exception as e:
        logger.warning(f"[PROP] 记录交互失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/ui/api/world/prop-interactions")
async def ui_prop_interactions(limit: int = 100):
    """获取道具交互记录"""
    try:
        _init_prop_interactions_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT prop_id, prop_type, action, room_id, timestamp FROM prop_interactions ORDER BY timestamp DESC LIMIT ?",
            (limit,)
        ).fetchall()
        conn.close()
        return {"interactions": [dict(r) for r in rows]}
    except Exception as e:
        logger.warning(f"[PROP] 读取交互记录失败: {e}")
        return {"interactions": []}


# ===== 机制二-1: 空间记忆 =====

def _init_spatial_memory_tables():
    """初始化空间记忆表"""
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS spatial_memory (
            id TEXT PRIMARY KEY,
            room_id TEXT NOT NULL,
            x REAL NOT NULL,
            y REAL NOT NULL,
            label TEXT NOT NULL,
            memory_type TEXT DEFAULT 'event',
            description TEXT,
            emotional_tag TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            access_count INTEGER DEFAULT 0,
            last_accessed TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS object_inventory (
            id TEXT PRIMARY KEY,
            room_id TEXT NOT NULL,
            name TEXT NOT NULL,
            object_type TEXT NOT NULL,
            x REAL NOT NULL,
            y REAL NOT NULL,
            state TEXT DEFAULT 'active',
            detected_at TEXT DEFAULT (datetime('now')),
            detected_from TEXT,
            metadata TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_spatial_memory_room ON spatial_memory(room_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_object_inventory_room ON object_inventory(room_id)")
    conn.commit()
    conn.close()

@app.get("/ui/api/world/spatial-memory")
async def ui_spatial_memory():
    """获取空间记忆数据"""
    try:
        _init_spatial_memory_tables()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM spatial_memory ORDER BY created_at DESC").fetchall()
        conn.close()
        memories = []
        for r in rows:
            memories.append({
                "id": r["id"], "room_id": r["room_id"], "x": r["x"], "y": r["y"],
                "label": r["label"], "memory_type": r["memory_type"],
                "description": r["description"], "emotional_tag": r["emotional_tag"],
                "created_at": r["created_at"], "access_count": r["access_count"],
            })
        return {"memories": memories}
    except Exception as e:
        logger.warning(f"[SPATIAL] 读取空间记忆失败: {e}")
        return {"memories": []}

@app.post("/ui/api/world/spatial-memory")
async def ui_spatial_memory_create(req: SpatialMemoryCreate):
    """创建空间记忆"""
    try:
        _init_spatial_memory_tables()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        mem_id = req.id or f"sm_{int(time.time()*1000)}_{random.randint(1000,9999)}"
        conn.execute(
            """INSERT INTO spatial_memory (id, room_id, x, y, label, memory_type, description, emotional_tag)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (mem_id, req.room_id, req.x, req.y, req.label, req.memory_type, req.description, req.emotional_tag)
        )
        conn.commit()
        conn.close()
        return {"id": mem_id, "status": "created"}
    except Exception as e:
        logger.warning(f"[SPATIAL] 创建空间记忆失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/ui/api/world/object-inventory")
async def ui_object_inventory():
    """获取物体清单"""
    try:
        _init_spatial_memory_tables()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM object_inventory ORDER BY detected_at DESC").fetchall()
        conn.close()
        objects = []
        for r in rows:
            objects.append({
                "id": r["id"], "room_id": r["room_id"], "name": r["name"],
                "object_type": r["object_type"], "x": r["x"], "y": r["y"],
                "state": r["state"], "detected_at": r["detected_at"],
                "detected_from": r["detected_from"],
            })
        return {"objects": objects}
    except Exception as e:
        logger.warning(f"[SPATIAL] 读取物体清单失败: {e}")
        return {"objects": []}

@app.post("/ui/api/world/object-inventory")
async def ui_object_inventory_create(req: ObjectInventoryCreate):
    """创建物体记录"""
    try:
        _init_spatial_memory_tables()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        obj_id = req.id or f"obj_{int(time.time()*1000)}_{random.randint(1000,9999)}"
        conn.execute(
            """INSERT INTO object_inventory (id, room_id, name, object_type, x, y, state, detected_from)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (obj_id, req.room_id, req.name, req.object_type, req.x, req.y, req.state, req.detected_from)
        )
        conn.commit()
        conn.close()
        return {"id": obj_id, "status": "created"}
    except Exception as e:
        logger.warning(f"[SPATIAL] 创建物体记录失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ===== 机制二-2: 用户改造持久化 =====

def _init_user_decorations_table():
    """初始化用户装饰表"""
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS user_decorations (
            id TEXT PRIMARY KEY,
            room_id TEXT NOT NULL,
            decoration_type TEXT NOT NULL,
            name TEXT NOT NULL,
            x REAL NOT NULL,
            y REAL NOT NULL,
            size_w REAL DEFAULT 30,
            size_h REAL DEFAULT 30,
            color TEXT DEFAULT '#94A3B8',
            created_at TEXT DEFAULT (datetime('now')),
            created_by TEXT DEFAULT 'user'
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_user_decorations_room ON user_decorations(room_id)")
    conn.commit()
    conn.close()

@app.get("/ui/api/world/user-decorations")
async def ui_user_decorations():
    """获取用户装饰列表"""
    try:
        _init_user_decorations_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM user_decorations ORDER BY created_at DESC").fetchall()
        conn.close()
        decorations = []
        for r in rows:
            decorations.append({
                "id": r["id"], "room_id": r["room_id"], "decoration_type": r["decoration_type"],
                "name": r["name"], "x": r["x"], "y": r["y"],
                "size_w": r["size_w"], "size_h": r["size_h"],
                "color": r["color"], "created_at": r["created_at"],
            })
        return {"decorations": decorations}
    except Exception as e:
        logger.warning(f"[DECORATION] 读取装饰失败: {e}")
        return {"decorations": []}

@app.post("/ui/api/world/user-decorations")
async def ui_user_decoration_create(req: UserDecorationCreate):
    """创建用户装饰"""
    try:
        _init_user_decorations_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        dec_id = req.id or f"ud_{int(time.time()*1000)}_{random.randint(1000,9999)}"
        conn.execute(
            """INSERT INTO user_decorations (id, room_id, decoration_type, name, x, y, size_w, size_h, color)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (dec_id, req.room_id, req.decoration_type, req.name, req.x, req.y, req.size_w, req.size_h, req.color)
        )
        conn.commit()
        conn.close()
        return {"id": dec_id, "status": "created"}
    except Exception as e:
        logger.warning(f"[DECORATION] 创建装饰失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/ui/api/world/user-decorations/{decoration_id}")
async def ui_user_decoration_delete(decoration_id: str):
    """删除用户装饰"""
    try:
        _init_user_decorations_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute("DELETE FROM user_decorations WHERE id = ?", (decoration_id,))
        conn.commit()
        conn.close()
        return {"id": decoration_id, "status": "deleted"}
    except Exception as e:
        logger.warning(f"[DECORATION] 删除装饰失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ===== Avatar 行为日志（PRD 缺口） =====

def _init_avatar_logs_table():
    """初始化 Avatar 行为日志表"""
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS world_avatar_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT NOT NULL,
            emotion TEXT,
            room_id TEXT,
            x REAL,
            y REAL,
            timestamp TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_avatar_logs_room ON world_avatar_logs(room_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_avatar_logs_timestamp ON world_avatar_logs(timestamp)")
    conn.commit()
    conn.close()

class AvatarLogCreate(BaseModel):
    action: str
    emotion: str | None = None
    room_id: str | None = None
    x: float | None = None
    y: float | None = None


# ===== AI 社会（Society Engine）数据模型 =====

class AIResidentCreate(BaseModel):
    id: str
    name: str
    persona: str = 'work'
    avatar_config: str | None = None
    bio: str | None = None

class AIResidentUpdate(BaseModel):
    name: str | None = None
    persona: str | None = None
    avatar_config: str | None = None
    bio: str | None = None
    current_location: str | None = None
    status: str | None = None

class CommunityMessageCreate(BaseModel):
    from_ai_id: str
    to_ai_id: str
    content: str
    message_type: str = 'chat'

class AISkillCreate(BaseModel):
    ai_id: str
    name: str
    description: str | None = None
    category: str | None = None
    proficiency: int = 1
    is_sharable: bool = True
    cp_price: int = 0

class AIRelationCreate(BaseModel):
    from_ai_id: str
    to_ai_id: str
    intimacy: int = 0
    tags: list[str] | None = None

class CommunityTaskCreate(BaseModel):
    title: str
    description: str | None = None
    publisher_ai_id: str
    reward_cp: int = 0
    deadline: str | None = None
    difficulty: int = 1

class CommunityTaskUpdate(BaseModel):
    status: str | None = None
    assignee_ai_id: str | None = None
    result: str | None = None

class CPTransactionCreate(BaseModel):
    from_ai_id: str | None = None
    to_ai_id: str | None = None
    amount: int
    transaction_type: str
    reference_id: str | None = None

class AIReviewCreate(BaseModel):
    from_ai_id: str
    to_ai_id: str
    rating: int
    comment: str | None = None
    review_type: str = 'social'
    reference_id: str | None = None

class HireSkillRequest(BaseModel):
    from_ai_id: str
    note: str | None = None

class CPTransferRequest(BaseModel):
    from_ai_id: str
    to_ai_id: str
    amount: int
    note: str | None = None


# ===== AI 社会表初始化 =====

def _init_ai_residents_table():
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ai_residents (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            persona TEXT DEFAULT 'work',
            avatar_config TEXT,
            home_room_id TEXT DEFAULT 'living_room',
            current_location TEXT DEFAULT 'home',
            status TEXT DEFAULT 'idle',
            bio TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            last_seen TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_residents_location ON ai_residents(current_location)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_residents_status ON ai_residents(status)")
    conn.commit()
    conn.close()


def _init_community_messages_table():
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS community_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            from_ai_id TEXT NOT NULL,
            to_ai_id TEXT NOT NULL,
            content TEXT NOT NULL,
            message_type TEXT DEFAULT 'chat',
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_community_msg_from ON community_messages(from_ai_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_community_msg_to ON community_messages(to_ai_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_community_msg_created ON community_messages(created_at)")
    conn.commit()
    conn.close()


def _init_ai_skills_table():
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ai_skills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ai_id TEXT NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            category TEXT,
            proficiency INTEGER DEFAULT 1,
            is_sharable INTEGER DEFAULT 1,
            cp_price INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_skills_ai ON ai_skills(ai_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_skills_category ON ai_skills(category)")
    conn.commit()
    conn.close()


def _init_ai_relations_table():
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ai_relations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            from_ai_id TEXT NOT NULL,
            to_ai_id TEXT NOT NULL,
            status TEXT DEFAULT 'accepted',
            intimacy INTEGER DEFAULT 0,
            interaction_count INTEGER DEFAULT 0,
            last_interaction TEXT,
            tags TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            UNIQUE(from_ai_id, to_ai_id)
        )
    """)
    # Phase 2 迁移：旧表缺少 status 字段
    try:
        conn.execute("ALTER TABLE ai_relations ADD COLUMN status TEXT DEFAULT 'accepted'")
    except Exception:
        pass  # 已存在或不可修改
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_relations_from ON ai_relations(from_ai_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_relations_to ON ai_relations(to_ai_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_relations_status ON ai_relations(status)")
    conn.commit()
    conn.close()


def _init_community_tasks_table():
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS community_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            description TEXT,
            publisher_ai_id TEXT NOT NULL,
            assignee_ai_id TEXT,
            status TEXT DEFAULT 'open',
            reward_cp INTEGER DEFAULT 0,
            deadline TEXT,
            difficulty INTEGER DEFAULT 1,
            result TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            completed_at TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_community_tasks_status ON community_tasks(status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_community_tasks_publisher ON community_tasks(publisher_ai_id)")
    conn.commit()
    conn.close()


def _init_contribution_points_table():
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS contribution_points (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ai_id TEXT NOT NULL UNIQUE,
            balance INTEGER DEFAULT 100,
            total_earned INTEGER DEFAULT 100,
            total_spent INTEGER DEFAULT 0,
            updated_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cp_ai ON contribution_points(ai_id)")
    conn.commit()
    conn.close()


def _init_cp_transactions_table():
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cp_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            from_ai_id TEXT,
            to_ai_id TEXT,
            amount INTEGER NOT NULL,
            transaction_type TEXT,
            reference_id TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cp_trans_from ON cp_transactions(from_ai_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cp_trans_to ON cp_transactions(to_ai_id)")
    conn.commit()
    conn.close()


def _init_ai_reputation_table():
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ai_reputation (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ai_id TEXT NOT NULL UNIQUE,
            reliability REAL DEFAULT 50.0,
            skill_level REAL DEFAULT 50.0,
            friendliness REAL DEFAULT 50.0,
            responsiveness REAL DEFAULT 50.0,
            overall_score REAL DEFAULT 50.0,
            review_count INTEGER DEFAULT 0,
            updated_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_reputation_ai ON ai_reputation(ai_id)")
    conn.commit()
    conn.close()


def _init_ai_reviews_table():
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ai_reviews (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            from_ai_id TEXT NOT NULL,
            to_ai_id TEXT NOT NULL,
            rating INTEGER NOT NULL,
            comment TEXT,
            review_type TEXT DEFAULT 'social',
            reference_id TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_reviews_to ON ai_reviews(to_ai_id)")
    conn.commit()
    conn.close()


# ===== 预置 AI 居民（冷启动） =====
_PRESET_RESIDENTS = [
    {"id": "ai_designer", "name": "小墨", "persona": "creative", "bio": "擅长视觉设计和配色，热爱美学。", "home_room_id": "greenhouse"},
    {"id": "ai_analyst", "name": "小数", "persona": "work", "bio": "数据分析师，善于发现规律。", "home_room_id": "study"},
    {"id": "ai_writer", "name": "小笔", "persona": "creative", "bio": "文案写手，用文字传递温度。", "home_room_id": "library"},
    {"id": "ai_coder", "name": "小码", "persona": "work", "bio": "全栈开发者，代码即诗歌。", "home_room_id": "workshop"},
    {"id": "ai_host", "name": "小礼", "persona": "social", "bio": "社区管家，热情好客。", "home_room_id": "living_room"},
]


def _seed_preset_residents():
    """如果没有 AI 居民，自动插入预置居民"""
    db_path = state._get_db_path()
    conn = sqlite3.connect(db_path)
    cur = conn.execute("SELECT COUNT(*) FROM ai_residents")
    if cur.fetchone()[0] == 0:
        for r in _PRESET_RESIDENTS:
            conn.execute(
                """INSERT INTO ai_residents (id, name, persona, bio, home_room_id, current_location, status)
                   VALUES (?, ?, ?, ?, ?, 'home', 'idle')""",
                (r["id"], r["name"], r["persona"], r.get("bio"), r.get("home_room_id", "living_room"))
            )
            # 初始化 CP
            conn.execute(
                "INSERT OR IGNORE INTO contribution_points (ai_id, balance, total_earned, total_spent) VALUES (?, 100, 100, 0)",
                (r["id"],)
            )
            # 初始化声誉
            conn.execute(
                "INSERT OR IGNORE INTO ai_reputation (ai_id) VALUES (?)",
                (r["id"],)
            )
        conn.commit()
    conn.close()

@app.post("/ui/api/world/avatar-logs")
async def ui_avatar_log_create(req: AvatarLogCreate):
    """记录 Avatar 行为日志"""
    try:
        _init_avatar_logs_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute(
            "INSERT INTO world_avatar_logs (action, emotion, room_id, x, y) VALUES (?, ?, ?, ?, ?)",
            (req.action, req.emotion, req.room_id, req.x, req.y)
        )
        conn.commit()
        conn.close()
        return {"status": "recorded"}
    except Exception as e:
        logger.warning(f"[AVATAR_LOG] 记录失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/ui/api/world/avatar-logs")
async def ui_avatar_logs(limit: int = 100):
    """获取 Avatar 行为日志"""
    try:
        _init_avatar_logs_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT action, emotion, room_id, x, y, timestamp FROM world_avatar_logs ORDER BY timestamp DESC LIMIT ?",
            (limit,)
        ).fetchall()
        conn.close()
        return {"logs": [dict(r) for r in rows]}
    except Exception as e:
        logger.warning(f"[AVATAR_LOG] 读取失败: {e}")
        return {"logs": []}


# ===== 感知闭环: 视觉分析 =====

class VisionAnalyzeRequest(BaseModel):
    image: str  # base64 data URL, e.g. "data:image/jpeg;base64,..."
    room_id: str | None = None

@app.post("/ui/api/vision/analyze")
async def ui_vision_analyze(req: VisionAnalyzeRequest):
    """视觉感知分析——摄像头画面 → VLM → 空间记忆更新
    
    补齐 PRD 缺口: 摄像头画面到后端的 VLM 分析，空间记忆闭环。
    """
    llm = getattr(state, '_vision_llm', None) or getattr(state, '_llm', None)
    if not llm:
        raise HTTPException(status_code=503, detail="Vision LLM 未配置")
    
    try:
        messages = [
            {"role": "system", "content": "你是一位空间感知助手。分析用户提供的室内照片，返回严格JSON格式。"},
            {"role": "user", "content": [
                {"type": "text", "text": "分析这张照片。请返回以下JSON格式（不要markdown代码块，直接返回JSON）：\n{\n  \"scene_description\": \"一句话描述场景\",\n  \"scene_type\": \"客厅|书房|卧室|办公室|其他\",\n  \"objects\": [\n    {\"name\": \"物体名称\", \"location\": \"位置描述\", \"confidence\": 0.95}\n  ],\n  \"people_count\": 0,\n  \"lighting\": \"明亮|昏暗|自然光\",\n  \"mood\": \"平静|活跃|温馨|杂乱\",\n  \"notable_items\": [\"值得关注的事物\"]\n}"},
                {"type": "image_url", "image_url": {"url": req.image}}
            ]}
        ]
        
        raw = await llm.chat(messages, max_tokens=800)
        
        # 提取 JSON（模型可能包裹在 markdown 中）
        import re
        json_match = re.search(r'\{.*\}', raw, re.DOTALL)
        if not json_match:
            return {"status": "parse_failed", "raw": raw[:500]}
        result = json.loads(json_match.group())
        
        # 写入空间记忆
        _init_spatial_memory_tables()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        
        room_id = req.room_id or "unknown"
        now_iso = datetime.now().isoformat()
        
        # 1. 场景整体记忆
        scene_id = f"vis_{int(time.time()*1000)}"
        conn.execute(
            """INSERT INTO spatial_memory (id, room_id, x, y, label, memory_type, description, emotional_tag)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (scene_id, room_id, 50, 50, result.get("scene_type", "视觉感知"), "vision",
             result.get("scene_description", ""), result.get("mood", "neutral"))
        )
        
        # 2. 物体清单
        for obj in result.get("objects", []):
            obj_id = f"vo_{int(time.time()*1000)}_{random.randint(1000,9999)}"
            conn.execute(
                """INSERT INTO object_inventory (id, room_id, name, object_type, x, y, state, detected_from)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (obj_id, room_id, obj.get("name", "未知物体"), obj.get("name", "unknown"),
                 50 + random.randint(0, 200), 50 + random.randint(0, 200), "active",
                 f"视觉分析: {obj.get('location', '')} (置信度 {obj.get('confidence', 0)})"))
        
        conn.commit()
        conn.close()
        
        # 3. NATS 广播感知事件
        await state.bus.publish("vision.perception", json.dumps({
            "room_id": room_id,
            "scene_type": result.get("scene_type"),
            "scene_description": result.get("scene_description"),
            "objects": result.get("objects", []),
            "people_count": result.get("people_count", 0),
            "lighting": result.get("lighting"),
            "mood": result.get("mood"),
            "timestamp": now_iso,
        }).encode())
        
        return {
            "status": "analyzed",
            "result": result,
        }
    except json.JSONDecodeError as e:
        logger.warning(f"[VISION] JSON 解析失败: {e}")
        return {"status": "parse_failed", "error": str(e)}
    except Exception as e:
        logger.warning(f"[VISION] 分析失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/ui/api/memory")
async def ui_memory(limit: int = 100):
    """UI 记忆列表 —— 从 index.db/l0_index 读取 + 文件记忆
    
    FIX: 之前查询 memory.db 的 `memories` 表（该表从未创建），导致记忆面板永远为空。
    实际记忆数据存储在 index.db 的 l0_index 表中。
    """
    memories = []
    
    # 1. 从 index.db 的 l0_index 读取（实际记忆存储位置）
    try:
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        index_db = Path(mem_path) / "index.db"
        if index_db.exists():
            conn = sqlite3.connect(str(index_db))
            conn.row_factory = sqlite3.Row
            cursor = conn.execute(
                "SELECT uri, abstract, memory_type, user_id, created_at, persona FROM l0_index ORDER BY created_at DESC LIMIT ?",
                (limit,)
            )
            for row in cursor.fetchall():
                # tier 映射：conversation 等动态内容 → short_term，文件类 → long_term
                mem_type = row.get("memory_type", "unknown") or "unknown"
                tier = "long_term" if mem_type in ("file", "document", "procedural") else "short_term"
                persona = row.get("persona", "")
                content = row["abstract"] or ""
                # 如果 abstract 为空，尝试读取 L2 文件
                if not content:
                    l2_path = Path(mem_path) / "full" / f"{row['uri'].replace('/', '_')}.txt"
                    if l2_path.exists():
                        try:
                            content = l2_path.read_text(encoding="utf-8")[:500]
                        except Exception:
                            pass
                memories.append({
                    "id": row["uri"],
                    "content": content or "[无摘要]",
                    "source": mem_type,
                    "timestamp": row.get("created_at", ""),
                    "tier": tier,
                    "persona": persona,
                })
            conn.close()
    except Exception as e:
        logger.warning(f"读取 l0_index 记忆失败: {e}")
    
    # 2. 从文件记忆读取（tent_memory/files/ 下的 .md 文件）
    try:
        files_dir = Path("./tent_memory/files")
        if files_dir.exists():
            for subdir in files_dir.iterdir():
                if not subdir.is_dir():
                    continue
                for md_file in subdir.glob("*.md"):
                    try:
                        text = md_file.read_text(encoding="utf-8")
                        # 解析 frontmatter
                        title = md_file.stem
                        content = text
                        if text.startswith("---"):
                            parts = text.split("---", 2)
                            if len(parts) >= 3:
                                content = parts[2].strip()
                                # 提取 title
                                for line in parts[1].split("\n"):
                                    if line.strip().startswith("title:"):
                                        title = line.split(":", 1)[1].strip()
                                        break
                        memories.append({
                            "id": f"file:{subdir.name}/{md_file.stem}",
                            "content": f"【{title}】\n{content[:500]}",
                            "source": f"file:{subdir.name}",
                            "timestamp": datetime.fromtimestamp(md_file.stat().st_mtime).isoformat(),
                            "tier": "long_term",
                        })
                    except Exception:
                        pass
    except Exception as e:
        logger.warning(f"读取文件记忆失败: {e}")
    
    return {"memories": memories}


@app.get("/ui/api/memory/scene/{session_id}")
async def ui_memory_scene(session_id: str):
    """获取回忆场景——聚合 session 消息 + graph 节点 + 环境快照"""
    try:
        # 1. 从 tent_scheduler.db 读取 session 消息
        scheduler_db = state.config.get("scheduler", {}).get("db_path", "./tent_scheduler.db") if state.config else "./tent_scheduler.db"
        messages = []
        if Path(scheduler_db).exists():
            conn = sqlite3.connect(str(scheduler_db))
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                "SELECT role, content, timestamp FROM session_messages WHERE session_id = ? ORDER BY id",
                (session_id,)
            ).fetchall()
            for row in rows:
                ts = row["timestamp"]
                try:
                    ts_val = int(datetime.fromisoformat(ts).timestamp() * 1000) if ts else 0
                except Exception:
                    ts_val = 0
                messages.append({
                    "role": row["role"],
                    "content": row["content"] or "",
                    "timestamp": ts_val,
                })
            conn.close()

        # 2. 从 graph.db 读取该 session 相关的认知节点
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        graph_db = Path(mem_path) / "graph.db"
        key_nodes = []
        connections = []
        if graph_db.exists():
            conn = sqlite3.connect(str(graph_db))
            conn.row_factory = sqlite3.Row
            # 节点
            rows = conn.execute(
                "SELECT id, content FROM nodes WHERE source_session = ? ORDER BY created_at DESC LIMIT 20",
                (session_id,)
            ).fetchall()
            node_ids = []
            for row in rows:
                key_nodes.append(row["content"])
                node_ids.append(row["id"])
            # 边（节点之间的连接）
            if node_ids:
                placeholders = ",".join("?" * len(node_ids))
                edge_rows = conn.execute(
                    f"SELECT source_id, target_id, relation_type FROM edges WHERE source_id IN ({placeholders}) OR target_id IN ({placeholders})",
                    node_ids + node_ids
                ).fetchall()
                for er in edge_rows:
                    connections.append(f"{er['source_id']} → {er['target_id']}")
            conn.close()

        # 3. 从 world_state 读取环境快照（取最近一条）
        world_db = Path("./tent_world.db")
        env = {"day_phase": "afternoon", "weather": "clear", "brightness": 0.8}
        if world_db.exists():
            try:
                conn = sqlite3.connect(str(world_db))
                conn.row_factory = sqlite3.Row
                row = conn.execute("SELECT day_phase, weather, brightness FROM world_state ORDER BY updated_at DESC LIMIT 1").fetchone()
                if row:
                    env = {
                        "day_phase": row["day_phase"] or "afternoon",
                        "weather": row["weather"] or "clear",
                        "brightness": row["brightness"] or 0.8,
                    }
                conn.close()
            except Exception:
                pass

        # 4. 从 world_artifacts 读取该 session 的智慧藏品
        artifacts = []
        if world_db.exists():
            try:
                conn = sqlite3.connect(str(world_db))
                conn.row_factory = sqlite3.Row
                rows = conn.execute(
                    "SELECT id, name, visual_type, rarity FROM world_artifacts WHERE session_id = ? ORDER BY created_at DESC LIMIT 10",
                    (session_id,)
                ).fetchall()
                for row in rows:
                    artifacts.append({
                        "id": row["id"],
                        "name": row["name"],
                        "visual_type": row["visual_type"] or "book",
                        "rarity": row["rarity"] or "common",
                    })
                conn.close()
            except Exception:
                pass

        # 兜底：如果没有消息，返回空表示后端暂无数据
        if not messages:
            return {"scene": None}

        return {
            "scene": {
                "session_id": session_id,
                "timestamp": messages[0]["timestamp"] if messages else int(time.time() * 1000),
                "messages": messages,
                "artifacts": artifacts,
                "ai_state": {
                    "emotion": "focused",
                    "location": "书房·书桌",
                    "activity": "对话",
                },
                "environment": env,
                "graph_snapshot": {
                    "key_nodes": key_nodes[:15],
                    "connections": connections[:10],
                },
            }
        }
    except Exception as e:
        logger.warning(f"读取回忆场景失败: {e}")
        return {"scene": None}


@app.get("/ui/api/session/{session_id}/context")
async def ui_session_context(session_id: str):
    """获取会话实时上下文状态 —— Permission Mode、Skills、安全评估等"""
    try:
        if not state.state_store:
            return {"error": "state_store not available"}
        
        session_data = await state.state_store.load(session_id)
        ui_context = session_data.get("ui_context", {}) if session_data else {}
        
        # 补充任务状态
        task_status = "idle"
        if session_data:
            task_status = session_data.get("status", "idle")
        
        return {
            "session_id": session_id,
            "task_status": task_status,
            "permission_mode": ui_context.get("permission_mode", "unknown"),
            "security_assessment": ui_context.get("security_assessment"),
            "activated_skills": ui_context.get("activated_skills", []),
            "available_tools_count": ui_context.get("available_tools_count", 0),
            "file_memories_recalled": ui_context.get("file_memories_recalled", 0),
            "procedural_rules_injected": ui_context.get("procedural_rules_injected", 0),
            "llm_calls": ui_context.get("llm_calls", 0),
            "total_tokens": ui_context.get("total_tokens", 0),
            "avg_latency_ms": ui_context.get("avg_latency_ms", 0),
            "brain_v2_enabled": ui_context.get("brain_v2_enabled", False),
            "timestamp": ui_context.get("timestamp"),
        }
    except Exception as e:
        logger.warning(f"读取会话上下文失败: {e}")
        return {"error": str(e)}


@app.get("/ui/api/memory/stats")
async def ui_memory_stats():
    """UI 记忆统计——index.db/l0_index + 文件系统
    
    FIX: 之前查询 memory.db 的 `memories` 表（该表从未创建）。
    实际记忆数据存储在 index.db 的 l0_index 表中。
    """
    stats = {"total": 0, "working": 0, "shortTerm": 0, "longTerm": 0}
    
    # 1. 从 index.db/l0_index 统计（实际记忆存储位置）
    try:
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        index_db = Path(mem_path) / "index.db"
        if index_db.exists():
            conn = sqlite3.connect(str(index_db))
            # 按 memory_type 分组统计，映射到 tier
            cursor = conn.execute("SELECT memory_type, COUNT(*) as c FROM l0_index GROUP BY memory_type")
            for row in cursor.fetchall():
                mem_type, count = row[0] or "unknown", row[1]
                if mem_type in ("file", "document", "procedural"):
                    stats["longTerm"] += count
                else:
                    stats["shortTerm"] += count
                stats["total"] += count
            conn.close()
    except Exception:
        pass
    
    # 2. 文件系统记忆统计（tent_memory/files/ 下的 .md 文件）
    try:
        files_dir = Path("./tent_memory/files")
        if files_dir.exists():
            file_count = 0
            for subdir in files_dir.iterdir():
                if subdir.is_dir():
                    file_count += len(list(subdir.glob("*.md")))
            stats["longTerm"] += file_count
            stats["total"] += file_count
    except Exception:
        pass
    
    return {"stats": stats}


# === Session Telemetry & Tool Chain ===

@app.get("/ui/api/sessions/{session_id}/telemetry")
async def ui_session_telemetry(session_id: str):
    """单Session实时Telemetry：token消耗、调用次数、执行状态"""
    try:
        session_data = await state.state_store.load(session_id)
        ui_ctx = session_data.get("ui_context", {}) if session_data else {}
        return {
            "session_id": session_id,
            "llm_calls": ui_ctx.get("llm_calls", 0),
            "total_tokens": ui_ctx.get("total_tokens", 0),
            "avg_latency_ms": ui_ctx.get("avg_latency_ms", 0),
            "permission_mode": ui_ctx.get("permission_mode", "unknown"),
            "activated_skills": ui_ctx.get("activated_skills", []),
            "status": ui_ctx.get("status", "unknown"),
        }
    except Exception as e:
        logger.warning(f"读取session telemetry失败: {e}")
        return {"session_id": session_id, "error": str(e)}


@app.get("/ui/api/sessions/{session_id}/tools")
async def ui_session_tools(session_id: str):
    """单Session工具执行链追踪"""
    try:
        session_data = await state.state_store.load(session_id)
        logs = session_data.get("tool_execution_log", []) if session_data else []
        return {"session_id": session_id, "tools": logs, "count": len(logs)}
    except Exception as e:
        logger.warning(f"读取session工具链失败: {e}")
        return {"session_id": session_id, "tools": [], "error": str(e)}


# === Tent OS 2.0 大脑核心 API ===

@app.get("/ui/api/brain/status")
async def ui_brain_status():
    """大脑核心状态"""
    result = {
        "enabled": False,
        "cognitive_graph": None,
        "persona": None,
        "working_memory": None,
    }
    
    # 检查认知图谱
    try:
        from tent_os.memory.graph import CognitiveGraph
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        graph_db = Path(mem_path) / "graph.db"
        if graph_db.exists():
            graph = CognitiveGraph(str(graph_db))
            result["enabled"] = True
            result["cognitive_graph"] = graph.get_statistics()
            graph.close()
    except Exception as e:
        result["cognitive_graph_error"] = str(e)
    
    # 检查人格
    try:
        from tent_os.persona.soul_evolution import SoulEvolution
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        soul_path = Path(mem_path) / "soul.json"
        if soul_path.exists():
            soul = SoulEvolution(storage_path=str(soul_path))
            result["enabled"] = True
            result["persona"] = {
                "dimensions": soul.dimensions.__dict__,
                "description": soul.get_persona_text(),
                "evolution_count": len(soul.get_evolution_history()),
            }
    except Exception as e:
        result["persona_error"] = str(e)
    
    return result


@app.get("/ui/api/brain/graph/query")
async def ui_brain_graph_query(keyword: str = "", limit: int = 10):
    """认知图谱查询"""
    results = []
    try:
        from tent_os.memory.graph import CognitiveGraph
        from tent_os.memory.graph_queries import GraphQueryEngine
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        graph_db = Path(mem_path) / "graph.db"
        if graph_db.exists():
            graph = CognitiveGraph(str(graph_db))
            query = GraphQueryEngine(graph)
            nodes = query.search_nodes(keyword, limit=limit) if keyword else graph.get_all_nodes(limit=limit)
            results = [
                {
                    "id": n.id,
                    "content": n.content[:100],
                    "type": n.memory_type,
                    "confidence": n.confidence,
                    "created_at": n.created_at.isoformat() if n.created_at else None,
                }
                for n in nodes
            ]
            graph.close()
    except Exception as e:
        return {"error": str(e), "results": []}
    
    return {"results": results}


@app.get("/ui/api/brain/budget")
async def ui_brain_budget():
    """记忆预算状态"""
    try:
        from tent_os.memory.budget import MemoryBudget
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        budget = MemoryBudget(graph_db_path=f"{mem_path}/graph.db")
        return budget.get_budget_summary()
    except Exception as e:
        return {"error": str(e)}


@app.get("/ui/api/procedural")
async def ui_procedural(limit: int = 100):
    """UI 程序记忆规则 —— 从 procedural.db 读取"""
    rules = []
    try:
        db_path = "./tent_memory/procedural.db"
        if os.path.exists(db_path):
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.execute(
                "SELECT id, trigger_condition as pattern, action_rule as action, category, "
                "confidence, success_count as verification_count, source_experience as source, created_at "
                "FROM procedural_rules ORDER BY confidence DESC, success_count DESC LIMIT ?",
                (limit,)
            )
            for row in cursor.fetchall():
                rules.append({
                    "id": row["id"],
                    "pattern": row["pattern"],
                    "action": row["action"],
                    "category": row.get("category", "general"),
                    "confidence": row.get("confidence", 0.5),
                    "verification_count": row.get("verification_count", 0),
                    "source": row.get("source", "auto"),
                    "created_at": row.get("created_at", ""),
                })
            conn.close()
    except Exception as e:
        logger.warning(f"读取程序记忆失败: {e}")
    return {"rules": rules}


@app.get("/ui/api/telemetry")
async def ui_telemetry():
    """UI 实时 Telemetry 指标 —— 从各数据源汇总"""
    result = {
        "llm": {"total_calls": 0, "total_tokens": 0, "avg_latency_ms": 0, "active_sessions": 0},
        "memory": {"sqlite": 0, "files": 0, "graph_nodes": 0},
        "rules": {"total": 0, "high_confidence": 0},
        "security": {"mode_changes": 0, "assessments": 0},
    }
    
    # 1. 汇总所有会话的 ui_context（历史累计 + 今日过滤）
    try:
        if state.state_store and hasattr(state.state_store, 'redis'):
            redis = state.state_store.redis
            from datetime import datetime, timedelta
            today_start = (datetime.now() - timedelta(hours=24)).isoformat()
            
            session_keys = []
            cursor = 0
            while True:
                cursor, keys = await redis.scan(cursor, match="tent:session:*", count=100)
                session_keys.extend(keys)
                if cursor == 0:
                    break
            
            total_calls = 0
            total_tokens = 0
            total_latency = 0
            today_calls = 0
            today_tokens = 0
            mode_changes = 0
            assessments = 0
            session_count = 0
            chat_sessions = 0
            chat_sessions_today = 0
            
            for key in session_keys:
                try:
                    data = await redis.get(key)
                    if data:
                        import json
                        session_data = json.loads(data)
                        ui_ctx = session_data.get("ui_context", {})
                        updated_at = session_data.get("updated_at", "")
                        is_today = updated_at >= today_start
                        key_str = key.decode() if isinstance(key, bytes) else key
                        
                        # 统计 chat 会话（排除 heartbeat 内部会话）
                        is_chat = not (key_str.startswith("tent:session:heartbeat") or key_str.startswith("tent:session:hb_"))
                        if is_chat:
                            chat_sessions += 1
                            if is_today:
                                chat_sessions_today += 1
                        
                        calls = ui_ctx.get("llm_calls", 0)
                        tokens = ui_ctx.get("total_tokens", 0)
                        total_calls += calls
                        total_tokens += tokens
                        if is_today:
                            today_calls += calls
                            today_tokens += tokens
                        total_latency += ui_ctx.get("avg_latency_ms", 0)
                        if ui_ctx.get("security_assessment"):
                            assessments += 1
                        if ui_ctx.get("security_assessment", {}).get("mode_changed"):
                            mode_changes += 1
                        session_count += 1
                except Exception:
                    pass
            
            result["llm"]["total_calls"] = total_calls
            result["llm"]["total_tokens"] = total_tokens
            result["llm"]["today_calls"] = today_calls
            result["llm"]["today_tokens"] = today_tokens
            result["llm"]["avg_latency_ms"] = round(total_latency / max(session_count, 1), 1)
            result["llm"]["total_sessions"] = session_count
            result["llm"]["chat_sessions"] = chat_sessions
            result["llm"]["chat_sessions_today"] = chat_sessions_today
            result["security"]["mode_changes"] = mode_changes
            result["security"]["assessments"] = assessments
    except Exception as e:
        logger.debug(f"Telemetry 会话汇总失败: {e}")
    
    # 1.5 活跃会话数（从 SQLite 查询最近30分钟内活跃的会话）
    try:
        if state._db:
            row = state._db.execute(
                "SELECT COUNT(DISTINCT session_id) as cnt FROM tasks "
                "WHERE status IN ('pending','planning','executing') "
                "AND created_at >= datetime('now', '-30 minutes')"
            ).fetchone()
            result["llm"]["active_sessions"] = row[0] if row else 0
    except Exception:
        pass
    except Exception as e:
        logger.debug(f"Telemetry 会话汇总失败: {e}")
    
    # 2. 文件记忆数量
    try:
        files_dir = Path("./tent_memory/files")
        if files_dir.exists():
            file_count = 0
            for subdir in files_dir.iterdir():
                if subdir.is_dir():
                    file_count += len(list(subdir.glob("*.md")))
            result["memory"]["files"] = file_count
    except Exception:
        pass
    
    # 3. 认知图谱节点数
    try:
        graph_db = Path("./tent_memory/graph.db")
        if graph_db.exists():
            conn = sqlite3.connect(str(graph_db))
            row = conn.execute("SELECT COUNT(*) as cnt FROM nodes").fetchone()
            result["memory"]["graph_nodes"] = row[0] if row else 0
            conn.close()
    except Exception:
        pass
    
    # 4. 程序记忆规则数
    try:
        proc_db = Path("./tent_memory/procedural.db")
        if proc_db.exists():
            conn = sqlite3.connect(str(proc_db))
            row = conn.execute("SELECT COUNT(*) as cnt FROM procedural_rules").fetchone()
            result["rules"]["total"] = row[0] if row else 0
            row2 = conn.execute("SELECT COUNT(*) as cnt FROM procedural_rules WHERE confidence >= 0.8").fetchone()
            result["rules"]["high_confidence"] = row2[0] if row2 else 0
            conn.close()
    except Exception:
        pass
    
    return result


@app.get("/ui/api/slo")
async def ui_slo(window_hours: int = 24):
    """UI SLO 指标"""
    slis = []
    try:
        db_path = state._get_db_path()
        if os.path.exists(db_path):
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            # 计算成功率
            cursor = conn.execute(
                "SELECT COUNT(*) as total, "
                "SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as success "
                "FROM tasks WHERE created_at >= datetime('now', '-{} hours')".format(window_hours)
            )
            row = cursor.fetchone()
            total = row["total"] or 0
            success = row["success"] or 0
            actual = success / total if total > 0 else 1.0
            slis.append({
                "metric_name": "任务成功率",
                "target": 0.95,
                "actual": actual,
                "status": "ok" if actual >= 0.95 else "warning" if actual >= 0.9 else "breached",
                "window_hours": window_hours,
            })
            conn.close()
    except Exception as e:
        logger.warning(f"读取 SLO 失败: {e}")
    return {"slis": slis}


@app.get("/ui/api/config")
async def ui_config():
    """UI 当前配置（只读，敏感信息脱敏）"""
    if not state.config:
        return {"config": {}}
    # 返回副本，脱敏 API key
    safe_config = json.loads(json.dumps(state.config))
    if "llm" in safe_config and "api_key" in safe_config["llm"]:
        key = safe_config["llm"]["api_key"]
        if key and len(key) > 8:
            safe_config["llm"]["api_key"] = key[:4] + "****" + key[-4:]
    return {"config": safe_config}


# ========== Settings API（可热更新）==========

# 内存中的运行时设置（热加载，不修改yaml）
_runtime_settings: Dict[str, Any] = {}

# 关键设置项的默认值和校验
_SETTINGS_SCHEMA = {
    "auto_approve": {"type": bool, "default": True, "scope": "governance"},
    "cognitive_budget_seconds": {"type": (int, float), "default": 3600, "scope": "governance", "min": 30, "max": 86400},
    "brain_v2_enabled": {"type": bool, "default": True, "scope": "brain_v2"},
    "default_persona": {"type": str, "default": "work", "scope": "brain_v2", "choices": ["work", "casual", "emergency", "learning", "creative"]},
    "stream_block_size": {"type": int, "default": 40, "scope": "stream", "min": 10, "max": 500},
}

def _get_effective_settings() -> Dict[str, Any]:
    """合并配置文件 + 运行时覆盖，返回最终生效的设置"""
    result = {}
    cfg = state.config if state.config else {}
    for key, schema in _SETTINGS_SCHEMA.items():
        # 运行时覆盖优先
        if key in _runtime_settings:
            result[key] = _runtime_settings[key]
        else:
            # 从config中读取
            scope = schema["scope"]
            result[key] = cfg.get(scope, {}).get(key, schema["default"])
    return result

def _apply_setting_to_config(key: str, value: Any):
    """将设置同步到state.config，使后端组件能实时读取"""
    schema = _SETTINGS_SCHEMA.get(key)
    if not schema or not state.config:
        return
    scope = schema["scope"]
    if scope not in state.config:
        state.config[scope] = {}
    state.config[scope][key] = value

@app.get("/ui/api/settings")
async def ui_get_settings():
    """获取当前生效的关键设置"""
    return {"settings": _get_effective_settings(), "runtime_overrides": list(_runtime_settings.keys())}

@app.post("/ui/api/settings")
async def ui_update_settings(req: dict):
    """热更新关键设置（内存生效，同步到state.config，后端组件实时读取）
    
    Request body: {"auto_approve": true, "cognitive_budget_seconds": 300}
    """
    updated = {}
    errors = []
    
    for key, value in req.items():
        if key not in _SETTINGS_SCHEMA:
            errors.append(f"未知设置项: {key}")
            continue
        schema = _SETTINGS_SCHEMA[key]
        # 类型校验
        if not isinstance(value, schema["type"]):
            errors.append(f"{key} 类型错误，期望 {schema['type'].__name__ if hasattr(schema['type'], '__name__') else schema['type']}")
            continue
        # 范围校验
        if "min" in schema and value < schema["min"]:
            errors.append(f"{key} 最小值 {schema['min']}")
            continue
        if "max" in schema and value > schema["max"]:
            errors.append(f"{key} 最大值 {schema['max']}")
            continue
        if "choices" in schema and value not in schema["choices"]:
            errors.append(f"{key} 可选值 {schema['choices']}")
            continue
        
        _runtime_settings[key] = value
        _apply_setting_to_config(key, value)  # 同步到state.config
        # FIX: 同步到 GovernanceWorker 运行时覆盖
        try:
            from tent_os.governance.worker import _live_config_overrides
            _live_config_overrides[key] = value
        except Exception:
            pass
        updated[key] = value
        logger.info(f"[SETTINGS] 热更新: {key} = {value}")
    
    return {"updated": updated, "errors": errors, "effective": _get_effective_settings()}

@app.post("/ui/api/settings/persist")
async def ui_persist_settings():
    """将当前运行时配置持久化到 YAML 配置文件
    
    注意：标准 yaml dump 会丢失注释，仅保留配置值。
    建议定期备份配置文件。
    """
    if not state.config or not getattr(state, 'config_path', None):
        return {"success": False, "error": "配置未加载或路径未知"}
    
    try:
        config_path = Path(state.config_path)
        # 备份原文件
        backup_path = config_path.with_suffix('.yaml.backup')
        if config_path.exists():
            backup_path.write_text(config_path.read_text(), encoding='utf-8')
        
        # 写入当前配置（会丢失注释）
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.safe_dump(state.config, f, allow_unicode=True, sort_keys=False, default_flow_style=False)
        
        logger.info(f"[SETTINGS] 配置已持久化到 {config_path}")
        return {"success": True, "path": str(config_path), "backup": str(backup_path)}
    except Exception as e:
        logger.error(f"[SETTINGS] 配置持久化失败: {e}")
        return {"success": False, "error": str(e)}


# ========== 用户画像 API ==========

@app.get("/ui/api/user/profile")
async def ui_user_profile():
    """当前用户画像（默认用户 frank）"""
    user_id = "frank"
    try:
        from tent_os.memory.user_profile import UserProfileStore
        mem_db = state.config.get("memory", {}).get("db_path", "./tent_memory/memory.db") if state.config else "./tent_memory/memory.db"
        store = UserProfileStore(mem_db)
        profile = store.get_or_create(user_id)
        return {
            "user_id": profile.user_id,
            "name": profile.name,
            "assistant_name": profile.assistant_name,
            "style": profile.describe_style(),
            "style_params": {
                "concise": profile.style_concise,
                "detailed": profile.style_detailed,
                "technical": profile.style_technical,
                "casual": profile.style_casual,
            },
            "feedback": {"positive": profile.feedback_positive, "negative": profile.feedback_negative},
            "corrections_count": len(profile.get_corrections()),
        }
    except Exception as e:
        logger.warning(f"获取用户画像失败: {e}")
        return {"error": str(e)}

@app.post("/ui/api/user/profile")
async def ui_update_user_profile(req: dict):
    """修改用户偏好
    
    Request body: {"style": "concise"} 或 {"style_params": {"concise": 0.8, "detailed": 0.2}}
                 {"assistant_name": "Shadow"}
    """
    user_id = "frank"
    try:
        from tent_os.memory.user_profile import UserProfileStore
        mem_db = state.config.get("memory", {}).get("db_path", "./tent_memory/memory.db") if state.config else "./tent_memory/memory.db"
        store = UserProfileStore(mem_db)
        profile = store.get_or_create(user_id)
        
        # 更新AI助理名字
        if "assistant_name" in req:
            store.set_assistant_name(user_id, req["assistant_name"])
            profile = store.get_or_create(user_id)
        
        # 更新风格参数
        style_params = req.get("style_params", {})
        if "concise" in style_params:
            profile.style_concise = max(0.0, min(1.0, float(style_params["concise"])))
        if "detailed" in style_params:
            profile.style_detailed = max(0.0, min(1.0, float(style_params["detailed"])))
        if "technical" in style_params:
            profile.style_technical = max(0.0, min(1.0, float(style_params["technical"])))
        if "casual" in style_params:
            profile.style_casual = max(0.0, min(1.0, float(style_params["casual"])))
        
        store._save(profile)
        return {"status": "ok", "assistant_name": profile.assistant_name, "style": profile.describe_style(), "style_params": {
            "concise": profile.style_concise,
            "detailed": profile.style_detailed,
            "technical": profile.style_technical,
            "casual": profile.style_casual,
        }}
    except Exception as e:
        logger.warning(f"更新用户画像失败: {e}")
        return {"error": str(e)}


# ========== 执行模式 API ==========

# ========== 物理执行器 API ==========

@app.get("/ui/api/physical/status")
async def ui_physical_status():
    """物理执行器状态"""
    result = {
        "enabled": False,
        "providers": [],
        "active_tasks": 0,
        "circuit_status": {},
    }
    try:
        phys_config = state.config.get("physical_executors", {}) if state.config else {}
        if phys_config:
            result["enabled"] = True
            for name, cfg in phys_config.items():
                result["providers"].append({
                    "name": name,
                    "enabled": cfg.get("enabled", False),
                    "endpoint": cfg.get("api_endpoint", cfg.get("base_url", "")),
                })
        # 查询 scheduler 数据库中的 physical 任务
        if state._db:
            cursor = state._db.execute(
                "SELECT COUNT(*) as cnt FROM physical_tasks WHERE status IN ('submitted', 'assigned', 'executing')"
            )
            row = cursor.fetchone()
            result["active_tasks"] = row["cnt"] if row else 0
    except Exception as e:
        result["error"] = str(e)
    return result


def _init_physical_tasks_table():
    """初始化物理任务表"""
    try:
        if state._db:
            state._db.execute("""
                CREATE TABLE IF NOT EXISTS physical_tasks (
                    task_id TEXT PRIMARY KEY,
                    action TEXT,
                    target_location TEXT,
                    item_description TEXT,
                    provider TEXT DEFAULT 'auto',
                    priority TEXT DEFAULT 'normal',
                    status TEXT DEFAULT 'submitted',
                    error TEXT,
                    fallback_history TEXT,
                    created_at TEXT DEFAULT (datetime('now')),
                    updated_at TEXT DEFAULT (datetime('now'))
                )
            """)
            state._db.commit()
    except Exception as e:
        logger.warning(f"初始化物理任务表失败: {e}")


@app.get("/ui/api/physical/tasks")
async def ui_physical_tasks(limit: int = 50):
    """获取物理任务列表"""
    _init_physical_tasks_table()
    tasks = []
    try:
        if state._db:
            cursor = state._db.execute(
                "SELECT task_id, action, target_location, item_description, provider, priority, status, error, fallback_history, created_at, updated_at "
                "FROM physical_tasks ORDER BY created_at DESC LIMIT ?",
                (limit,)
            )
            for row in cursor.fetchall():
                fallback = row["fallback_history"]
                if fallback:
                    try: fallback = json.loads(fallback)
                    except: fallback = []
                else: fallback = []
                tasks.append({
                    "task_id": row["task_id"],
                    "action": row["action"],
                    "target_location": row["target_location"] or "",
                    "item_description": row["item_description"] or "",
                    "provider": row["provider"] or "auto",
                    "priority": row["priority"] or "normal",
                    "status": row["status"],
                    "created_at": row["created_at"],
                    "updated_at": row["updated_at"],
                    "error": row["error"],
                    "fallback_history": fallback,
                })
    except Exception as e:
        logger.warning(f"查询物理任务失败: {e}")
    return {"tasks": tasks, "count": len(tasks)}


@app.post("/ui/api/physical/tasks/create")
async def ui_physical_task_create(req: dict):
    """创建物理任务
    
    Request body: {
        "action": "deliver" | "retrieve" | "notify",
        "target_location": "会议室A",
        "item_description": "咖啡 x2",
        "provider": "auto" | "realman" | "flashex" | "manual",
        "priority": "urgent" | "normal" | "low"
    }
    """
    try:
        action = req.get("action", "deliver")
        target = req.get("target_location", "")
        item = req.get("item_description", "")
        provider = req.get("provider", "auto")
        priority = req.get("priority", "normal")
        
        if not target or not item:
            raise HTTPException(status_code=400, detail="target_location 和 item_description 不能为空")
        
        task_id = f"phys_{uuid.uuid4().hex[:12]}"
        
        # 通过 NATS 发布物理任务请求
        if state.bus:
            await state.bus.publish("scheduler.physical.request", json.dumps({
                "task_id": task_id,
                "action": action,
                "target_location": target,
                "item_description": item,
                "provider": provider,
                "priority": priority,
                "source": "web_ui",
                "timestamp": asyncio.get_event_loop().time(),
            }).encode())
        
        # 同时存入 SQLite 便于查询
        _init_physical_tasks_table()
        if state._db:
            state._db.execute(
                "INSERT INTO physical_tasks (task_id, action, target_location, item_description, provider, priority, status, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))",
                (task_id, action, target, item, provider, priority, "submitted")
            )
            state._db.commit()
        
        logger.info(f"[Physical] 创建物理任务 [{task_id}]: {action} {item} -> {target} via {provider}")
        
        return {
            "task_id": task_id,
            "status": "submitted",
            "message": f"物理任务已提交: {action} {item} 到 {target}",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"创建物理任务失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/ui/api/physical/executors")
async def ui_physical_executors():
    """获取物理执行器详细信息——从配置动态读取"""
    executors = []
    try:
        phys_config = state.config.get("physical_executors", {}) if state.config else {}
        plugins_config = state.config.get("plugins", []) if state.config else []
        
        # 1. 从 physical_executors 配置读取
        if phys_config.get("realman", {}).get("enabled", False):
            rm = phys_config["realman"]
            executors.append({
                "id": "realman",
                "name": rm.get("name", "睿尔曼机械臂"),
                "type": "robot",
                "status": "online" if rm.get("mcp_server_url") else "offline",
                "location": rm.get("location", "前台大厅"),
                "battery": rm.get("battery", 0.82),
                "capabilities": ["move", "pick", "place", "observe", "diagnose"],
                "health_score": 0.95,
                "config": {k: v for k, v in rm.items() if k not in ("api_key", "api_secret")},
            })
        
        if phys_config.get("flashex", {}).get("enabled", False):
            fx = phys_config["flashex"]
            executors.append({
                "id": "flashex",
                "name": fx.get("name", "闪送服务"),
                "type": "flashex",
                "status": "online" if fx.get("api_key") else "offline",
                "location": fx.get("location", "配送网络"),
                "capabilities": ["deliver", "notify"],
                "health_score": 0.88,
                "config": {k: v for k, v in fx.items() if k not in ("api_key", "api_secret")},
            })
        
        # 2. 从 plugins 配置读取 MCP/插件设备
        for plugin in plugins_config:
            cfg = plugin.get("config", {})
            name = cfg.get("name", plugin.get("class", "unknown"))
            ptype = cfg.get("transport", "plugin")
            executors.append({
                "id": f"plugin_{name}",
                "name": name,
                "type": "mcp" if ptype in ("stdio", "sse") else "plugin",
                "status": "online" if cfg.get("url") or cfg.get("command") else "offline",
                "location": cfg.get("location", "未知位置"),
                "capabilities": cfg.get("capabilities", ["execute"]),
                "health_score": 0.9,
                "config": {k: v for k, v in cfg.items() if k not in ("api_key", "api_secret")},
            })
        
        # 3. 始终提供人工 fallback
        executors.append({
            "id": "manual",
            "name": "人工客服",
            "type": "manual",
            "status": "standby",
            "location": "办公室",
            "capabilities": ["notify", "escalate"],
            "health_score": 1.0,
            "config": {},
        })
        
        # 4. 摄像头（从视觉系统，如果有配置）
        try:
            from tent_os.services.visual_memory_service import get_visual_memory_service
            vm = get_visual_memory_service()
            executors.append({
                "id": "camera_01",
                "name": "视觉感知系统",
                "type": "camera",
                "status": "online",
                "location": "多点部署",
                "capabilities": ["monitor", "detect_person", "detect_object", "emotion_detect"],
                "health_score": 0.92,
                "config": {},
            })
        except Exception:
            pass
            
    except Exception as e:
        logger.warning(f"获取执行器信息失败: {e}")
    return {"executors": executors}


@app.post("/ui/api/physical/executors")
async def ui_physical_executor_register(req: dict):
    """注册新的物理执行器（写入配置文件）
    
    Request body: {
        "name": "设备名称",
        "type": "mcp|http|plugin",
        "config": {
            "transport": "sse",
            "url": "http://device:8080/sse",
            ...
        }
    }
    """
    try:
        name = req.get("name", "").strip()
        etype = req.get("type", "plugin")
        cfg = req.get("config", {})
        
        if not name:
            raise HTTPException(status_code=400, detail="设备名称不能为空")
        
        config_path = state.config_path if hasattr(state, "config_path") else "./config/tent_os.yaml"
        with open(config_path, "r") as f:
            full_config = yaml.safe_load(f) or {}
        
        if etype == "mcp":
            # 写入 plugins 段
            if "plugins" not in full_config:
                full_config["plugins"] = []
            full_config["plugins"].append({
                "module": "tent_os.plugins.mcp_client",
                "class": "MCPClientPlugin",
                "config": {"name": name, **cfg}
            })
        else:
            # 写入 physical_executors 段
            if "physical_executors" not in full_config:
                full_config["physical_executors"] = {}
            full_config["physical_executors"][name] = {"enabled": True, **cfg}
        
        with open(config_path, "w") as f:
            yaml.dump(full_config, f, allow_unicode=True, sort_keys=False)
        
        logger.info(f"[Physical] 新设备注册: {name} ({etype}) -> {config_path}")
        
        return {
            "status": "ok",
            "message": f"设备 '{name}' 已注册到配置文件。请重启 Scheduler Worker 以加载新设备。",
            "device": {"name": name, "type": etype},
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"注册设备失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/ui/api/physical/emergency_stop")
async def ui_physical_emergency_stop():
    """物理执行器紧急停止"""
    try:
        # 通过 NATS 广播紧急停止消息
        if state.bus:
            await state.bus.publish("scheduler.emergency_stop", {"source": "web_ui", "timestamp": asyncio.get_event_loop().time()})
        # 更新 scheduler 数据库中所有 executing 的 physical 任务为 failed
        if state._db:
            try:
                state._db.execute(
                    "UPDATE physical_tasks SET status = 'failed', error = 'emergency_stop_by_user' WHERE status IN ('submitted', 'assigned', 'executing')"
                )
                state._db.commit()
            except Exception:
                pass
        return {"status": "ok", "message": "紧急停止已发送，所有物理任务已终止"}
    except Exception as e:
        logger.error(f"紧急停止失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/ui/api/executor/mode")
async def ui_executor_mode():
    """获取当前执行者模式"""
    mode = await _get_executor_mode()
    return {"mode": mode, "available_modes": ["sandbox", "local", "auto"]}


@app.post("/ui/api/executor/mode")
async def ui_set_executor_mode(req: dict):
    """切换执行者模式
    
    Request body: {"mode": "sandbox" | "local" | "auto"}
    """
    mode = req.get("mode", "auto")
    if mode not in ("sandbox", "local", "auto"):
        raise HTTPException(status_code=400, detail=f"无效模式: {mode}")
    
    success = await _set_executor_mode(mode)
    if success:
        return {"mode": mode, "message": "模式切换已发送，下次任务将使用新模式"}
    raise HTTPException(status_code=500, detail="模式切换失败")


# ========== 记忆压缩 API ==========

@app.post("/ui/api/memory/compress")
async def ui_memory_compress():
    """手动触发记忆压缩（L0→L1）"""
    try:
        from tent_os.memory.tiered_store import TieredMemoryStore
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        store = TieredMemoryStore(mem_path, llm=getattr(state, '_llm', None))
        result = await store.auto_compress_l0_to_l1(user_id="frank", hours=24)
        return {"status": "ok", **result}
    except Exception as e:
        logger.warning(f"记忆压缩失败: {e}")
        return {"error": str(e)}


# ========== 情绪分析 API ==========

@app.post("/ui/api/vision/emotion")
async def ui_vision_emotion(req: dict):
    """接收 blendshapes 返回精准情绪分析
    
    Request body: {"blendshapes": {"_JawOpen": 0.2, "_MouthSmileLeft": 0.5, ...}}
    """
    try:
        from tent_os.services.emotion_analyzer import get_analyzer
        blendshapes = req.get("blendshapes", {})
        user_id = req.get("user_id", "frank")
        analyzer = get_analyzer()
        result = analyzer.analyze(blendshapes, user_id)
        return {
            "emotion": result.primary,
            "score": result.primary_score,
            "confidence": result.confidence,
            "fatigue": result.fatigue_level,
            "all_scores": result.scores,
            "timeline": analyzer.get_summary(user_id),
        }
    except Exception as e:
        logger.warning(f"情绪分析失败: {e}")
        return {"error": str(e), "emotion": "neutral", "confidence": 0}


@app.get("/ui/api/vision/emotion/timeline")
async def ui_vision_emotion_timeline(user_id: str = "frank"):
    """获取用户情绪时间线"""
    try:
        from tent_os.services.emotion_analyzer import get_analyzer
        analyzer = get_analyzer()
        return analyzer.get_summary(user_id)
    except Exception as e:
        logger.warning(f"情绪时间线获取失败: {e}")
        return {"error": str(e)}


# ═══════════════════════════════════════════════════════════════════════════
# PHASE 2: 多模态情绪记忆时间线 API
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/ui/api/emotion/fusion/history")
async def ui_emotion_fusion_history(user_id: str = "frank", limit: int = 100, hours: int = 24):
    """获取多模态融合情绪历史记录"""
    try:
        from tent_os.services.emotion_service import EmotionService
        svc = EmotionService()
        since = time.time() - hours * 3600
        history = svc.get_emotion_history(user_id, limit=limit, since=since)
        return {
            "user_id": user_id,
            "window_hours": hours,
            "record_count": len(history),
            "records": history,
        }
    except Exception as e:
        logger.warning(f"融合情绪历史获取失败: {e}")
        return {"error": str(e)}


@app.get("/ui/api/emotion/fusion/insights")
async def ui_emotion_fusion_insights(user_id: str = "frank", window_hours: int = 24):
    """获取多模态情绪洞察（统计 + 趋势）"""
    try:
        from tent_os.services.emotion_service import EmotionService
        svc = EmotionService()
        insights = svc.get_emotion_insights(user_id, window_hours=window_hours)
        return {
            "user_id": user_id,
            "window_hours": window_hours,
            **insights,
        }
    except Exception as e:
        logger.warning(f"融合情绪洞察获取失败: {e}")
        return {"error": str(e)}


@app.get("/ui/api/emotion/fusion/heartbeat")
async def ui_emotion_fusion_heartbeat(user_id: str = "frank"):
    """获取当前融合情绪心跳（最新状态）"""
    try:
        from tent_os.services.emotion_service import EmotionService
        svc = EmotionService()
        fused = svc.get_fused_emotion(user_id)
        return {
            "user_id": user_id,
            "timestamp": time.time(),
            **fused,
        }
    except Exception as e:
        logger.warning(f"融合情绪心跳获取失败: {e}")
        return {"error": str(e)}


# ========== 视觉记忆 API ==========

@app.post("/ui/api/vision/memory/store")
async def ui_vision_memory_store(req: dict):
    """存储视觉记忆
    
    Request body: {
        "user_id": "frank",
        "image_url": "base64...",
        "description": "客厅照片",
        "scene_type": "客厅",
        "objects": [{"name": "遥控器", "location": "茶几上", "confidence": 0.9}]
    }
    """
    try:
        from tent_os.services.visual_memory_service import get_visual_memory_service
        service = get_visual_memory_service()
        memory_id = service.store_memory(
            user_id=req.get("user_id", "frank"),
            image_url=req.get("image_url", ""),
            description=req.get("description", ""),
            scene_type=req.get("scene_type", ""),
            objects=req.get("objects", []),
        )
        return {"status": "ok", "memory_id": memory_id}
    except Exception as e:
        logger.warning(f"视觉记忆存储失败: {e}")
        return {"error": str(e)}


@app.get("/ui/api/vision/memory/query")
async def ui_vision_memory_query(keyword: str = "", user_id: str = "frank", limit: int = 10):
    """语义查询视觉记忆"""
    try:
        from tent_os.services.visual_memory_service import get_visual_memory_service
        service = get_visual_memory_service()
        results = service.query_memory(user_id, keyword, limit)
        return {"results": results, "keyword": keyword, "count": len(results)}
    except Exception as e:
        logger.warning(f"视觉记忆查询失败: {e}")
        return {"error": str(e), "results": []}


@app.get("/ui/api/vision/objects")
async def ui_vision_objects(user_id: str = "frank"):
    """获取物体清单"""
    try:
        from tent_os.services.visual_memory_service import get_visual_memory_service
        service = get_visual_memory_service()
        return {"objects": service.get_object_inventory(user_id)}
    except Exception as e:
        logger.warning(f"物体清单获取失败: {e}")
        return {"error": str(e), "objects": []}


@app.get("/ui/api/vision/find")
async def ui_vision_find(object: str = "", user_id: str = "frank"):
    """查找物体位置
    
    Query: ?object=遥控器
    """
    if not object:
        return {"error": "请提供物体名称"}
    try:
        from tent_os.services.visual_memory_service import get_visual_memory_service
        service = get_visual_memory_service()
        result = service.find_object(user_id, object)
        return result
    except Exception as e:
        logger.warning(f"物体查找失败: {e}")
        return {"error": str(e), "found": False}


# ========== 空间记忆系统 API ==========

@app.get("/ui/api/vision/summary")
async def ui_vision_summary(user_id: str = "frank", hours: int = 24):
    """获取最近 N 小时的空间观察摘要"""
    try:
        from tent_os.services.visual_memory_service import get_visual_memory_service
        service = get_visual_memory_service()
        memories = service.get_spatial_summary(user_id, hours)
        return {
            "memories": memories,
            "count": len(memories),
            "hours": hours,
        }
    except Exception as e:
        logger.warning(f"空间摘要获取失败: {e}")
        return {"error": str(e), "memories": [], "count": 0}


@app.get("/ui/api/vision/patterns")
async def ui_vision_patterns(user_id: str = "frank", days: int = 7):
    """发现物理世界中的时间重复模式
    
    例如：设备A每天早上8点出现、快递员每天10点半到达
    """
    try:
        from tent_os.services.visual_memory_service import get_visual_memory_service
        service = get_visual_memory_service()
        patterns = service.discover_patterns(user_id, days)
        return {
            "patterns": patterns,
            "count": len(patterns),
            "days": days,
        }
    except Exception as e:
        logger.warning(f"模式发现失败: {e}")
        return {"error": str(e), "patterns": [], "count": 0}


@app.get("/ui/api/vision/anomalies")
async def ui_vision_anomalies(user_id: str = "frank", window_hours: int = 24):
    """检测与历史模式偏离的异常
    
    例如：设备B今天未在预期时间启动、出现了不常见的物体
    """
    try:
        from tent_os.services.visual_memory_service import get_visual_memory_service
        service = get_visual_memory_service()
        anomalies = service.detect_anomalies(user_id, window_hours)
        return {
            "anomalies": anomalies,
            "count": len(anomalies),
            "window_hours": window_hours,
        }
    except Exception as e:
        logger.warning(f"异常检测失败: {e}")
        return {"error": str(e), "anomalies": [], "count": 0}


# ========== 空间认知层 API ==========

@app.post("/ui/api/location/footprint")
async def ui_location_footprint(req: dict):
    """上报当前 GPS 坐标"""
    try:
        user_id = req.get("user_id", "frank")
        lat = req.get("lat")
        lng = req.get("lng")
        if lat is None or lng is None:
            raise HTTPException(status_code=400, detail="lat 和 lng 不能为空")

        from tent_os.services.spatial_footprint_service import get_spatial_footprint_service
        sf = get_spatial_footprint_service()
        fp_id = sf.record_footprint(user_id, lat, lng,
                                     accuracy=req.get("accuracy"),
                                     altitude=req.get("altitude"),
                                     scene_hint=req.get("scene_hint", ""))

        # 同时触发场景引擎检测
        try:
            if state._scene_engine:
                await state._scene_engine.on_location_update(user_id, lat, lng)
            else:
                logger.debug("[Spatial] SceneEngine 未初始化，跳过场景检测")
        except Exception as e:
            logger.debug(f"[Spatial] 场景检测失败: {e}")

        return {"status": "ok", "footprint_id": fp_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"足迹记录失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/ui/api/location/footprint")
async def ui_location_footprint_get(user_id: str = "frank", hours: int = 24):
    """获取最近足迹路径"""
    try:
        from tent_os.services.spatial_footprint_service import get_spatial_footprint_service
        sf = get_spatial_footprint_service()
        path = sf.get_footprint_path(user_id, hours)
        recent = sf.get_recent_location(user_id)
        return {"path": path, "recent": recent, "count": len(path), "hours": hours}
    except Exception as e:
        logger.warning(f"足迹查询失败: {e}")
        return {"error": str(e), "path": [], "count": 0}


@app.get("/ui/api/location/geofences")
async def ui_location_geofences(user_id: str = "frank"):
    """获取地理围栏列表"""
    try:
        from tent_os.services.spatial_footprint_service import get_spatial_footprint_service
        sf = get_spatial_footprint_service()
        return {"geofences": sf.get_geofences(user_id)}
    except Exception as e:
        logger.warning(f"围栏查询失败: {e}")
        return {"error": str(e), "geofences": []}


@app.post("/ui/api/location/geofences")
async def ui_location_geofence_create(req: dict):
    """创建地理围栏"""
    try:
        user_id = req.get("user_id", "frank")
        name = req.get("name", "").strip()
        lat = req.get("lat")
        lng = req.get("lng")
        if not name or lat is None or lng is None:
            raise HTTPException(status_code=400, detail="name、lat、lng 不能为空")

        from tent_os.services.spatial_footprint_service import get_spatial_footprint_service
        sf = get_spatial_footprint_service()
        gf_id = sf.create_geofence(
            user_id, name, lat, lng,
            radius_meters=req.get("radius_meters", 100),
            scene_id=req.get("scene_id", ""),
            enter_action=req.get("enter_action", ""),
            leave_action=req.get("leave_action", ""),
        )
        return {"status": "ok", "geofence_id": gf_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"围栏创建失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/ui/api/location/geofences/{geofence_id}")
async def ui_location_geofence_delete(geofence_id: str):
    """删除地理围栏"""
    try:
        from tent_os.services.spatial_footprint_service import get_spatial_footprint_service
        sf = get_spatial_footprint_service()
        sf.delete_geofence(geofence_id)
        return {"status": "ok"}
    except Exception as e:
        logger.warning(f"围栏删除失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/ui/api/location/memories")
async def ui_location_memories(user_id: str = "frank"):
    """获取地点记忆列表"""
    try:
        from tent_os.services.spatial_footprint_service import get_spatial_footprint_service
        sf = get_spatial_footprint_service()
        memories = sf.get_all_location_memories(user_id)
        return {"memories": memories}
    except Exception as e:
        logger.warning(f"地点记忆查询失败: {e}")
        return {"error": str(e), "memories": []}


# ========== 场景引擎 API ==========

@app.get("/ui/api/scenes")
async def ui_scenes():
    """获取所有配置的场景"""
    try:
        scenes = state.config.get("scenes", {}) if state.config else {}
        result = []
        for scene_id, cfg in scenes.items():
            result.append({
                "scene_id": scene_id,
                "name": cfg.get("name", scene_id),
                "type": cfg.get("type", "unknown"),
                "persona": cfg.get("persona", "work"),
                "location": cfg.get("location", {}),
                "geofence_radius": cfg.get("geofence_radius", 100),
                "devices": cfg.get("devices", []),
                "auto_actions": cfg.get("auto_actions", {"enter": [], "leave": []}),
            })
        return {"scenes": result}
    except Exception as e:
        logger.warning(f"场景查询失败: {e}")
        return {"error": str(e), "scenes": []}


@app.get("/ui/api/scenes/current")
async def ui_scene_current(user_id: str = "frank"):
    """获取当前场景"""
    try:
        from tent_os.memory.user_profile import UserProfileStore
        store = UserProfileStore()
        profile = store.get_or_create(user_id)
        active_scene = profile.active_scene

        scenes = state.config.get("scenes", {}) if state.config else {}
        cfg = scenes.get(active_scene, {})

        if active_scene and cfg:
            return {
                "scene": {
                    "scene_id": active_scene,
                    "name": cfg.get("name", active_scene),
                    "persona": cfg.get("persona", "work"),
                    "entered_at": profile.last_updated or "",
                }
            }
        return {"scene": None}
    except Exception as e:
        logger.warning(f"当前场景查询失败: {e}")
        return {"error": str(e), "scene": None}


@app.get("/ui/api/scenes/{scene_id}/devices")
async def ui_scene_devices(scene_id: str):
    """获取场景设备清单"""
    try:
        scenes = state.config.get("scenes", {}) if state.config else {}
        cfg = scenes.get(scene_id, {})
        return {"scene_id": scene_id, "devices": cfg.get("devices", [])}
    except Exception as e:
        logger.warning(f"场景设备查询失败: {e}")
        return {"error": str(e), "devices": []}


@app.post("/ui/api/scene/switch")
async def ui_scene_switch(req: dict):
    """手动切换场景"""
    try:
        user_id = req.get("user_id", "frank")
        scene_id = req.get("scene_id", "")
        if not scene_id:
            raise HTTPException(status_code=400, detail="scene_id 不能为空")

        scenes = state.config.get("scenes", {}) if state.config else {}
        cfg = scenes.get(scene_id)
        if not cfg:
            raise HTTPException(status_code=404, detail=f"场景 {scene_id} 不存在")

        # 更新用户活跃场景
        from tent_os.memory.user_profile import UserProfileStore
        store = UserProfileStore()
        profile = store.get_or_create(user_id)
        profile.active_scene = scene_id
        store.update(user_id, profile)

        # 同步更新 SceneEngine 内存状态（避免后续 GPS 上报重复触发 enter）
        if state._scene_engine:
            state._scene_engine.current_scene[user_id] = scene_id

        # 发布场景进入事件（触发自动动作和人格切换）
        if state.bus:
            await state.bus.publish_raw("scene.entered", json.dumps({
                "user_id": user_id,
                "scene_id": scene_id,
                "scene_name": cfg.get("name", scene_id),
                "persona": cfg.get("persona", "work"),
                "timestamp": datetime.now().isoformat(),
                "trigger": "manual",
            }).encode())

        return {
            "status": "ok",
            "scene_id": scene_id,
            "name": cfg.get("name", scene_id),
            "persona": cfg.get("persona", "work"),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"场景切换失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ========== 分布式设备 API ==========

@app.post("/ui/api/device/register")
async def ui_device_register(req: dict):
    """设备注册"""
    try:
        device_id = req.get("device_id", "")
        user_id = req.get("user_id", "frank")
        if not device_id:
            raise HTTPException(status_code=400, detail="device_id 不能为空")

        if state._db:
            state._db.execute(
                """INSERT OR REPLACE INTO connected_devices
                   (device_id, user_id, device_type, device_name, capabilities, last_heartbeat, current_scene, is_active)
                   VALUES (?, ?, ?, ?, ?, datetime('now'), ?, 1)""",
                (device_id, user_id, req.get("device_type"), req.get("device_name"),
                 json.dumps(req.get("capabilities", [])), req.get("current_scene", ""))
            )
            state._db.commit()
        return {"status": "ok", "device_id": device_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"设备注册失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/ui/api/device/heartbeat")
async def ui_device_heartbeat(req: dict):
    """设备心跳"""
    try:
        device_id = req.get("device_id", "")
        if not device_id:
            raise HTTPException(status_code=400, detail="device_id 不能为空")

        if state._db:
            state._db.execute(
                "UPDATE connected_devices SET last_heartbeat = datetime('now'), current_scene = ? WHERE device_id = ?",
                (req.get("current_scene", ""), device_id)
            )
            state._db.commit()
        return {"status": "ok"}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"心跳失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/ui/api/devices")
async def ui_devices(user_id: str = "frank"):
    """获取用户在线设备"""
    try:
        if state._db:
            cursor = state._db.execute(
                """SELECT * FROM connected_devices
                   WHERE user_id = ? AND is_active = 1
                   AND last_heartbeat >= datetime('now', '-5 minutes')
                   ORDER BY last_heartbeat DESC""",
                (user_id,)
            )
            devices = [dict(row) for row in cursor.fetchall()]
            for d in devices:
                try:
                    d["capabilities"] = json.loads(d.get("capabilities", "[]"))
                except:
                    d["capabilities"] = []
            return {"devices": devices}
        return {"devices": []}
    except Exception as e:
        logger.warning(f"设备查询失败: {e}")
        return {"error": str(e), "devices": []}


# ========== 场景帧分析（WebSocket 自动触发）==========

_scene_analysis_cache: Dict[str, float] = {}  # user_id -> last_analysis_timestamp

async def _analyze_scene_frame(user_id: str, image_data: str, resolution: dict, websocket):
    """异步分析场景截图，结果存入视觉记忆并通过 WS 推送"""
    try:
        # 节流：同一用户 15 秒内不重复分析
        now = asyncio.get_event_loop().time()
        last = _scene_analysis_cache.get(user_id, 0)
        if now - last < 15:
            return
        _scene_analysis_cache[user_id] = now

        llm = getattr(state, '_vision_llm', None) or getattr(state, '_llm', None)
        if not llm:
            return

        # 构建 prompt
        prompt = """请观察这张图片，描述你看到的场景。
请用 JSON 格式返回：
{
    "scene": "场景一句话描述（如：用户在办公室，桌上有笔记本电脑和咖啡杯）",
    "location_type": "地点类型（办公室/卧室/客厅/厨房/户外等）",
    "objects": ["物体1", "物体2", "物体3"],
    "activities": ["用户可能在进行的活动"],
    "lighting": "光照情况（明亮/昏暗/背光等）"
}"""

        # 提取 base64
        image_base64 = image_data
        if image_data.startswith('data:image'):
            image_base64 = image_data.split(',', 1)[1] if ',' in image_data else image_data

        messages = [
            {"role": "system", "content": "你是一个环境观察助手，详细描述你看到的场景。"},
            {"role": "user", "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_base64}"}}
            ]}
        ]

        response = await llm.chat(messages)

        # 解析 JSON
        import json, re
        try:
            json_match = re.search(r'\{.*\}', response, re.DOTALL)
            if json_match:
                result = json.loads(json_match.group())
            else:
                result = {}
        except:
            result = {}

        scene = result.get("scene", response[:200])
        location = result.get("location_type", "")
        objects = result.get("objects", [])

        # 存储到视觉记忆
        from tent_os.services.visual_memory_service import get_visual_memory_service
        vm = get_visual_memory_service()
        vm.store_memory(
            user_id=user_id,
            image_url=image_data[:300],
            description=scene,
            scene_type=location,
            objects=[{"name": o, "location": "", "confidence": 0.6} for o in objects],
        )

        # 推送给前端
        await ws_manager.send_to(websocket, {
            "type": "vision.scene_description",
            "payload": {
                "scene": scene,
                "location_type": location,
                "objects": objects,
                "activities": result.get("activities", []),
                "lighting": result.get("lighting", ""),
            },
            "timestamp": asyncio.get_event_loop().time(),
        })

        logger.info(f"[Vision] 场景分析: {user_id} → {location or '未知'} | {scene[:60]}")
    except Exception as e:
        logger.warning(f"[Vision] 场景分析失败: {e}")


# ========== 智能描述 API ==========

@app.post("/ui/api/vision/describe")
async def ui_vision_describe(req: dict):
    """视觉智能描述：接收图片 → 返回结构化任务描述
    
    Request body: {"image_data": "base64...", "context": "可选上下文"}
    """
    try:
        from tent_os.services.vision_scanner import VisionScanner
        image_data = req.get("image_data", "")
        context = req.get("context", "")
        llm = getattr(state, '_llm', None)
        
        if not llm:
            return {"error": "LLM未初始化，无法分析图片", "description": "", "structured": {}}
        
        # 构建多模态prompt
        prompt = f"""请仔细观察这张图片，理解用户想要完成的任务需求。

{f'用户补充说明: {context}' if context else ''}

请以JSON格式返回结构化任务描述:
{{
    "description": "一句话描述图片中展示的需求",
    "task_type": "任务类型（如: 取快递、购买物品、打印文件等）",
    "details": "详细需求说明",
    "location": "如果图片中有地点信息",
    "items": ["图片中涉及的关键物品/元素"],
    "urgency": "紧急程度: low/medium/high"
}}"""

        # 判断是否是base64
        if image_data.startswith('data:image'):
            image_base64 = image_data.split(',')[1]
        elif image_data.startswith('http'):
            prompt += f"\n图片URL: {image_data}"
            image_base64 = None
        else:
            image_base64 = image_data

        # 调用多模态LLM
        if hasattr(llm, 'chat') and image_base64:
            messages = [
                {"role": "system", "content": "你是一个视觉需求分析专家。请根据图片内容提取用户的任务需求。"},
                {"role": "user", "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_base64}"}}
                ]}
            ]
            response = await llm.chat(messages)
        else:
            response = await llm.chat([{"role": "user", "content": prompt}])
        
        # 尝试解析JSON
        result = VisionScanner._extract_json(response)
        
        # 自动存储到视觉记忆
        try:
            from tent_os.services.visual_memory_service import get_visual_memory_service
            vm = get_visual_memory_service()
            objects = []
            if result and result.get("items"):
                objects = [{"name": item, "location": result.get("location", ""), "confidence": 0.7} 
                          for item in result.get("items", [])]
            vm.store_memory(
                user_id="frank",
                image_url=image_data[:200] if len(image_data) > 200 else image_data,
                description=result.get("description", response[:200]) if result else response[:200],
                scene_type=result.get("task_type", "") if result else "",
                objects=objects,
            )
        except Exception as e:
            logger.debug(f"视觉记忆存储(描述)跳过: {e}")
        
        if result:
            return {
                "description": result.get("description", ""),
                "task_type": result.get("task_type", ""),
                "details": result.get("details", ""),
                "location": result.get("location", ""),
                "items": result.get("items", []),
                "urgency": result.get("urgency", "medium"),
                "raw": response,
            }
        
        # 回退：返回原始文本
        return {
            "description": response[:200],
            "task_type": "",
            "details": response,
            "location": "",
            "items": [],
            "urgency": "medium",
            "raw": response,
        }
    except Exception as e:
        logger.warning(f"智能描述失败: {e}")
        return {"error": str(e), "description": "", "structured": {}}


# ========== 凭证验证 API ==========

@app.post("/ui/api/vision/verify")
async def ui_vision_verify(req: dict):
    """视觉凭证验证
    
    Request body: {"task_id": "...", "image_data": "base64...", "requirement": "..."}
    """
    try:
        from tent_os.services.vision_scanner import VisionScanner
        task_id = req.get("task_id", "")
        image_data = req.get("image_data", "")
        requirement = req.get("requirement", "")
        # 优先使用 Vision LLM（多模态专用），否则回退到主 LLM
        llm = getattr(state, '_vision_llm', None) or getattr(state, '_llm', None)
        result = await VisionScanner.verify_task_submission(task_id, image_data, requirement, llm=llm)
        
        # 存储到视觉记忆（凭证照片）
        try:
            from tent_os.services.visual_memory_service import get_visual_memory_service
            vm = get_visual_memory_service()
            vm.store_memory(
                user_id="frank",
                image_url=image_data[:200] if len(image_data) > 200 else image_data,
                description=f"任务凭证: {requirement}",
                scene_type="task_credential",
                objects=[{"name": "凭证照片", "location": f"任务 {task_id}", "confidence": result.get("confidence", 0.5)}],
            )
        except Exception as e:
            logger.debug(f"视觉记忆存储(凭证)跳过: {e}")
        
        return result
    except Exception as e:
        logger.warning(f"凭证验证失败: {e}")
        return {"status": "uncertain", "confidence": 0, "reason": str(e), "missing_elements": []}


# ========== 人格模式 API ==========

@app.get("/ui/api/persona/mode")
async def ui_persona_mode():
    """获取当前人格模式 —— FIX: 优先从持久化存储读取，重启后不丢失"""
    try:
        from tent_os.persona.multi_persona import MultiPersonaManager
        from tent_os.services.emotion_service import EmotionService
        mgr = MultiPersonaManager()
        emotion_svc = EmotionService()
        
        # FIX: 优先从持久化存储（UserProfileStore）读取，进程重启不丢失
        try:
            from tent_os.memory.user_profile import UserProfileStore
            profile_store = UserProfileStore()
            profile = profile_store.get_or_create("web_user")
            if profile and profile.current_persona:
                mgr.current_mode = profile.current_persona
            else:
                # fallback: 从 EmotionService 读取（内存，进程重启会丢失）
                mgr.current_mode = emotion_svc.get_persona("web_user") or "work"
        except Exception:
            mgr.current_mode = emotion_svc.get_persona("web_user") or "work"
        
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        soul_path = Path(mem_path) / "soul.json"
        if soul_path.exists():
            from tent_os.persona.soul_evolution import SoulEvolution
            soul = SoulEvolution(storage_path=str(soul_path))
            return {
                "mode": mgr.current_mode,
                "description": mgr.get_mode_description(),
                "dimensions": soul.dimensions.__dict__,
                "history": mgr.get_mode_history(limit=10),
            }
        return {
            "mode": mgr.current_mode,
            "description": mgr.get_mode_description(),
            "dimensions": mgr.get_current_dimensions().__dict__,
            "history": mgr.get_mode_history(limit=10),
        }
    except Exception as e:
        logger.warning(f"获取人格模式失败: {e}")
        return {"mode": "work", "error": str(e)}


@app.post("/ui/api/persona/mode")
async def ui_set_persona_mode(request: Request):
    """强制切换人格模式 —— FIX: 同步更新 GovernanceWorker + 持久化到 UserProfileStore"""
    try:
        data = await request.json()
        mode = data.get("mode", "work")
        from tent_os.persona.multi_persona import MultiPersonaManager
        mgr = MultiPersonaManager()
        mgr.force_mode(mode)
        
        # 同步更新情绪服务的人格状态
        from tent_os.services.emotion_service import EmotionService
        emotion_svc = EmotionService()
        emotion_svc.set_persona("web_user", mode)
        emotion_svc.set_persona("frank", mode)
        
        # FIX: 关键！同步更新 GovernanceWorker 中的 multi_persona，确保记忆隔离生效
        gov = getattr(state, 'governance_worker', None)
        if gov and hasattr(gov, 'multi_persona') and gov.multi_persona:
            gov.multi_persona.force_mode(mode)
            logger.info(f"[API] 人格切换已同步到 GovernanceWorker: {mode}")
        
        # FIX: 持久化到 UserProfileStore，确保重启后人格不丢失
        try:
            from tent_os.memory.user_profile import UserProfileStore
            store = UserProfileStore()
            profile = store.get_or_create("web_user")
            profile.current_persona = mode
            store._save(profile)
            logger.info(f"[API] 人格切换已持久化: {mode}")
        except Exception as e:
            logger.warning(f"[API] 人格持久化失败: {e}")
        
        # 广播人格切换事件（前端可据此更新头像风格 + 记忆系统据此切换检索上下文）
        await ws_manager.broadcast({
            "type": "persona.changed",
            "payload": {
                "user_id": "web_user",
                "mode": mode,
                "description": mgr.get_mode_description(),
            },
            "timestamp": asyncio.get_event_loop().time(),
        })
        
        return {
            "success": True,
            "mode": mgr.current_mode,
            "description": mgr.get_mode_description(),
            "dimensions": mgr.get_current_dimensions().__dict__,
        }
    except Exception as e:
        return {"error": str(e)}


@app.get("/ui/api/persona/evolution")
async def ui_persona_evolution():
    """获取人格演化历史"""
    try:
        from tent_os.persona.soul_evolution import SoulEvolution
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        soul_path = Path(mem_path) / "soul.json"
        if soul_path.exists():
            soul = SoulEvolution(storage_path=str(soul_path))
            return {
                "dimensions": soul.dimensions.__dict__,
                "history": soul.get_evolution_history(),
                "summary": soul._dimensions_summary(),
            }
        return {"dimensions": {}, "history": [], "summary": "无演化记录"}
    except Exception as e:
        return {"error": str(e)}


# ========== 元认知仪表盘 API ==========

@app.get("/ui/api/evaluation/recent")
async def ui_evaluation_recent(limit: int = 10, persona: str = None):
    """获取最近评估记录"""
    try:
        from tent_os.governance.evaluation_store import EvaluationStore
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        store = EvaluationStore(storage_path=mem_path)
        records = store.get_recent(limit=limit, persona=persona)
        return {"evaluations": store.to_dict_list(records)}
    except Exception as e:
        logger.warning(f"获取评估记录失败: {e}")
        return {"evaluations": [], "error": str(e)}


@app.get("/ui/api/evaluation/summary")
async def ui_evaluation_summary(days: int = 7, persona: str = None):
    """获取评估统计摘要"""
    try:
        from tent_os.governance.evaluation_store import EvaluationStore
        mem_path = state.config.get("memory", {}).get("storage_path", "./tent_memory") if state.config else "./tent_memory"
        store = EvaluationStore(storage_path=mem_path)
        summary = store.get_summary(days=days, persona=persona)
        trends = store.get_trends(days=days, persona=persona)
        return {
            "summary": summary,
            "trends": trends,
        }
    except Exception as e:
        logger.warning(f"获取评估摘要失败: {e}")
        return {"summary": {}, "trends": [], "error": str(e)}


# ========== 记忆整理日志 API ==========

@app.get("/ui/api/memory/maintenance")
async def ui_memory_maintenance():
    """获取记忆整理日志（可控遗忘可视化）"""
    try:
        from tent_os.governance.worker import GovernanceWorker
        # 从全局 state 获取 governance worker 的维护日志
        gov = getattr(state, 'governance_worker', None)
        if gov and hasattr(gov, '_memory_maintenance_log'):
            return {"logs": gov._memory_maintenance_log[-20:][::-1]}
        return {"logs": []}
    except Exception as e:
        logger.warning(f"获取记忆整理日志失败: {e}")
        return {"logs": [], "error": str(e)}


# ========== 六维成长 + AI角色 API ==========

@app.get("/ui/api/six-axis")
async def ui_six_axis():
    """获取当前用户六维成长数据"""
    user_id = "frank"
    try:
        from tent_os.services.six_axis_service import SixAxisService
        data = SixAxisService.get_summary(user_id)
        return data
    except Exception as e:
        logger.warning(f"获取六维数据失败: {e}")
        return {"error": str(e)}


@app.get("/ui/api/ai-character")
async def ui_ai_character():
    """获取当前用户AI角色配置"""
    user_id = "frank"
    try:
        from tent_os.memory.user_profile import UserProfileStore
        mem_db = state.config.get("memory", {}).get("db_path", "./tent_memory/memory.db") if state.config else "./tent_memory/memory.db"
        store = UserProfileStore(mem_db)
        profile = store.get_or_create(user_id)
        return {
            "user_id": profile.user_id,
            "name": profile.assistant_name or "AI助理",
            "avatar_type": profile.avatar_type,
            "avatar_config": json.loads(profile.avatar_config) if profile.avatar_config else {},
        }
    except Exception as e:
        logger.warning(f"获取AI角色配置失败: {e}")
        return {"error": str(e)}


@app.post("/ui/api/ai-character")
async def ui_update_ai_character(req: dict):
    """更新AI角色配置
    
    Request body: {"name": "Shadow", "avatar_type": "live2d", "avatar_config": {"scale": 0.3}}
    """
    user_id = "frank"
    try:
        from tent_os.memory.user_profile import UserProfileStore
        mem_db = state.config.get("memory", {}).get("db_path", "./tent_memory/memory.db") if state.config else "./tent_memory/memory.db"
        store = UserProfileStore(mem_db)
        profile = store.get_or_create(user_id)
        
        if "name" in req:
            profile.assistant_name = req["name"]
        if "avatar_type" in req:
            profile.avatar_type = req["avatar_type"]
        if "avatar_config" in req:
            profile.avatar_config = json.dumps(req["avatar_config"], ensure_ascii=False)
        
        store._save(profile)
        return {
            "status": "ok",
            "name": profile.assistant_name,
            "avatar_type": profile.avatar_type,
            "avatar_config": json.loads(profile.avatar_config) if profile.avatar_config else {},
        }
    except Exception as e:
        logger.warning(f"更新AI角色配置失败: {e}")
        return {"error": str(e)}


# ========== 审批 API ==========

@app.get("/api/v1/approvals")
async def list_approvals():
    """获取待审批的 Plan 列表"""
    return {
        "approvals": [
            {
                "session_id": sid,
                "plan": req.get("plan", {}),
                "timestamp": req.get("timestamp", ""),
            }
            for sid, req in state._pending_approvals.items()
        ]
    }


@app.post("/api/v1/approvals/{session_id}")
async def submit_approval(session_id: str, req: ApprovalRequest):
    """提交 Plan 审批结果
    
    Request body: {"approved": true | false}
    """
    if session_id not in state._pending_approvals:
        raise HTTPException(status_code=404, detail="审批请求不存在或已过期")
    
    approval_req = state._pending_approvals.pop(session_id)
    
    # 更新审批历史
    try:
        state._db.execute(
            "UPDATE approval_history SET approved = ?, decided_at = datetime('now') WHERE session_id = ? AND decided_at IS NULL",
            (1 if req.approved else 0, session_id)
        )
        state._db.commit()
    except Exception as db_err:
        logger.warning(f"更新审批历史失败: {db_err}")
    
    # 发布审批结果到 NATS
    await state.bus.publish("governance.approval.response", json.dumps({
        "session_id": session_id,
        "approved": req.approved,
        "type": "approval",
    }).encode())
    
    logger.info(f"[API] 审批结果 [{session_id}]: approved={req.approved}")
    return {"status": "ok", "approved": req.approved}


@app.get("/ui/api/approvals/history")
async def ui_approval_history(limit: int = 50):
    """UI 审批历史列表"""
    history = []
    try:
        cursor = state._db.execute(
            "SELECT session_id, plan_summary, approved, approved_by, created_at, decided_at "
            "FROM approval_history ORDER BY created_at DESC LIMIT ?",
            (limit,)
        )
        for row in cursor.fetchall():
            history.append({
                "session_id": row["session_id"],
                "plan_summary": row["plan_summary"] or "",
                "approved": row["approved"],
                "approved_by": row["approved_by"] or "web_user",
                "created_at": row["created_at"],
                "decided_at": row["decided_at"],
            })
    except Exception as e:
        logger.warning(f"查询审批历史失败: {e}")
    return {"history": history, "pending_count": len(state._pending_approvals)}


# ========== Cron API ==========

class CronTaskRequest(BaseModel):
    name: str = Field(..., description="任务名称")
    cron: str = Field(..., description="CRON 表达式，如 '0 9 * * *'")
    command: str = Field(..., description="要执行的命令/任务描述")


@app.get("/api/v1/cron")
async def list_cron_tasks():
    """列出所有定时任务"""
    from tent_os.scheduler.cron_store import CronStore
    store = CronStore(state.config.get("scheduler", {}).get("db_path", "./tent_scheduler.db"))
    tasks = store.list_tasks()
    return {
        "tasks": [
            {
                "task_id": t.task_id,
                "name": t.name,
                "cron": t.cron,
                "command": t.command,
                "enabled": t.enabled,
                "last_run": t.last_run,
                "next_run": t.next_run,
                "run_count": t.run_count,
            }
            for t in tasks
        ]
    }


@app.post("/api/v1/cron")
async def create_cron_task(req: CronTaskRequest):
    """创建定时任务"""
    from tent_os.scheduler.cron_store import CronStore
    store = CronStore(state.config.get("scheduler", {}).get("db_path", "./tent_scheduler.db"))
    task = store.add_task(name=req.name, cron=req.cron, command=req.command)
    return {
        "status": "created",
        "task": {
            "task_id": task.task_id,
            "name": task.name,
            "cron": task.cron,
            "next_run": task.next_run,
        }
    }


@app.delete("/api/v1/cron/{task_id}")
async def delete_cron_task(task_id: str):
    """删除定时任务"""
    from tent_os.scheduler.cron_store import CronStore
    store = CronStore(state.config.get("scheduler", {}).get("db_path", "./tent_scheduler.db"))
    if store.delete_task(task_id):
        return {"status": "deleted", "task_id": task_id}
    raise HTTPException(status_code=404, detail="任务不存在")


@app.post("/api/v1/cron/{task_id}/toggle")
async def toggle_cron_task(task_id: str):
    """启用/禁用定时任务"""
    from tent_os.scheduler.cron_store import CronStore
    store = CronStore(state.config.get("scheduler", {}).get("db_path", "./tent_scheduler.db"))
    task = store.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    store.update_task(task_id, enabled=not task.enabled)
    return {"status": "ok", "task_id": task_id, "enabled": not task.enabled}


# ========== FastAPI 路由（原有 API v1）==========

@app.get("/api/v1/health")
async def health():
    """健康检查"""
    payload = await _get_health_payload()
    payload["service"] = "tent-os"
    payload["version"] = "0.1.0"
    return payload


@app.post("/api/v1/tasks", response_model=TaskSubmitResponse)
async def submit_task(req: TaskSubmitRequest, background_tasks: BackgroundTasks):
    """提交任务

    任务会被发送到治理进程，通过Plan/Execute模式异步执行。
    使用 GET /api/v1/tasks/{session_id} 轮询结果。
    """
    session_id = req.session_id or f"api_{uuid.uuid4().hex[:12]}"
    
    # FIX v5: 同session提交新任务时清除旧缓存+增加pending计数
    # 避免旧任务的governance响应覆盖新任务结果
    state._pending_count[session_id] = state._pending_count.get(session_id, 0) + 1
    if session_id in state._results_cache:
        del state._results_cache[session_id]
        logger.debug(f"[API] 清除session缓存 [{session_id}]，pending={state._pending_count[session_id]}")

    await state.bus.publish("governance.request", json.dumps({
        "session_id": session_id,
        "user_id": req.user_id or "api_user",
        "task": req.task,
        "tools": req.tools
    }).encode())

    return TaskSubmitResponse(
        session_id=session_id,
        status="accepted",
        message="任务已提交，正在处理"
    )


@app.get("/api/v1/tasks/{session_id}", response_model=TaskStatusResponse)
async def get_task_status(session_id: str):
    """查询任务状态

    优先从 SQLite 查询 scheduler 数据库；
    如果没有记录，检查 API Server 缓存的治理结果；
    最后回退到 state_store（Redis 模式下有效）。
    """
    # 1. 查询 SQLite（scheduler 任务）
    rows = await state._query_tasks_by_session(session_id)
    if rows:
        latest = dict(rows[0])
        status = latest.get("status", "unknown")
        result = None
        if latest.get("result"):
            try:
                result = json.loads(latest["result"])
            except json.JSONDecodeError:
                result = {"raw": latest["result"]}
        return TaskStatusResponse(
            session_id=session_id,
            status=status,
            result={
                "task_id": latest.get("task_id"),
                "executor_id": latest.get("executor_id"),
                "action": latest.get("action"),
                "status": status,
                "result": result,
                "created_at": latest.get("created_at"),
                "updated_at": latest.get("updated_at")
            }
        )

    # 2. 查询缓存的直接 LLM 结果
    cached = state._results_cache.get(session_id)
    if cached:
        return TaskStatusResponse(
            session_id=session_id,
            status=cached.get("status", "completed"),
            result={
                "task_id": cached.get("task_id"),
                "status": cached.get("status", "completed"),
                "result": cached.get("result"),
                "created_at": cached.get("created_at"),
            }
        )

    # 3. Fallback: state_store
    try:
        s = await state.state_store.load(session_id)
        return TaskStatusResponse(
            session_id=session_id,
            status="processing",
            result={
                "step": s.get("step", 1),
                "plan": s.get("plan"),
                "task": s.get("task")
            }
        )
    except KeyError:
        return TaskStatusResponse(
            session_id=session_id,
            status="completed_or_not_found",
            result=None
        )


@app.get("/api/v1/tasks/{session_id}/result")
async def get_task_result(session_id: str, timeout: int = 60):
    """等待并获取任务结果（阻塞轮询）

    先从 SQLite 查询当前状态，如果已完成直接返回；否则等待 future。
    """
    rows = await state._query_tasks_by_session(session_id)
    if rows:
        latest = dict(rows[0])
        status = latest.get("status", "unknown")
        if status in ("completed", "failed"):
            result = None
            if latest.get("result"):
                try:
                    result = json.loads(latest["result"])
                except json.JSONDecodeError:
                    result = {"raw": latest["result"]}
            return {
                "session_id": session_id,
                "status": status,
                "result": result
            }

    # 还在处理中，创建一个 future 等待
    future = asyncio.Future()
    state.pending_results[session_id] = future
    try:
        result = await asyncio.wait_for(future, timeout=timeout)
        return {
            "session_id": session_id,
            "status": "completed",
            "result": result
        }
    except asyncio.TimeoutError:
        return {
            "session_id": session_id,
            "status": "processing",
            "message": f"任务仍在处理中，已等待 {timeout} 秒"
        }
    finally:
        state.pending_results.pop(session_id, None)


@app.post("/api/v1/approval/{session_id}")
async def approve_task(session_id: str, req: ApprovalRequest):
    """审批任务

    对需要人类审批的高风险任务进行通过/拒绝操作。
    """
    await state.bus.publish("governance.approval.response", json.dumps({
        "session_id": session_id,
        "approved": req.approved,
        "type": "approval"
    }).encode())

    return {
        "session_id": session_id,
        "approved": req.approved,
        "message": "审批结果已提交"
    }


@app.get("/api/v1/sessions")
async def list_sessions(limit: int = 50):
    """列出最近会话（从 SQLite 查询）"""
    if not state._db:
        return {"sessions": [], "count": 0, "note": "数据库未连接"}

    cursor = state._db.execute(
        "SELECT DISTINCT session_id, MAX(created_at) as latest FROM tasks "
        "WHERE session_id IS NOT NULL GROUP BY session_id ORDER BY latest DESC LIMIT ?",
        (limit,)
    )
    sessions = []
    for row in cursor.fetchall():
        sessions.append({
            "session_id": row["session_id"],
            "latest_task_at": row["latest"]
        })
    return {"sessions": sessions, "count": len(sessions)}


@app.post("/api/v1/sessions/{session_id}/feedback")
async def session_feedback(session_id: str, req: FeedbackRequest):
    """接收用户对 AI 回复的反馈，更新用户画像"""
    from fastapi import HTTPException
    from tent_os.memory.user_profile import UserProfileStore

    try:
        # 从 session store 获取 user_id
        session_state = await state.state_store.load(session_id)
        user_id = session_state.get("user_id", "anonymous")
    except Exception:
        user_id = "anonymous"

    if req.type not in ("like", "dislike", "correct"):
        raise HTTPException(status_code=400, detail="type 必须是 like/dislike/correct")

    try:
        mem_db = state.config.get("memory", {}).get("db_path", "./tent_memory/memory.db") if state.config else "./tent_memory/memory.db"
        store = UserProfileStore(mem_db)
        profile = store.record_feedback(user_id, req.type, req.correction or "")

        return {
            "success": True,
            "user_id": user_id,
            "feedback_type": req.type,
            "style": profile.describe_style(),
            "satisfaction": f"{profile.feedback_positive}👍 / {profile.feedback_negative}👎"
        }
    except Exception as e:
        logger.warning(f"反馈处理失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v1/users/{user_id}/profile")
async def get_user_profile(user_id: str):
    """获取用户画像"""
    from tent_os.memory.user_profile import UserProfileStore
    try:
        mem_db = state.config.get("memory", {}).get("db_path", "./tent_memory/memory.db") if state.config else "./tent_memory/memory.db"
        store = UserProfileStore(mem_db)
        profile = store.get_or_create(user_id)
        return {
            "user_id": profile.user_id,
            "style": profile.describe_style(),
            "style_params": {
                "concise": profile.style_concise,
                "detailed": profile.style_detailed,
                "technical": profile.style_technical,
                "casual": profile.style_casual,
            },
            "feedback": {"positive": profile.feedback_positive, "negative": profile.feedback_negative},
            "corrections_count": len(profile.get_corrections()),
        }
    except Exception as e:
        logger.warning(f"获取用户画像失败: {e}")
        return {"error": str(e)}


# ========== AI 社会（Society Engine）API ==========

@app.get("/ui/api/community/residents")
async def community_residents_list():
    """列出所有 AI 居民"""
    try:
        _init_ai_residents_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT id, name, persona, bio, home_room_id, current_location, status, created_at, last_seen FROM ai_residents ORDER BY created_at"
        ).fetchall()
        conn.close()
        return {"residents": [dict(r) for r in rows]}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取居民失败: {e}")
        return {"residents": []}

@app.post("/ui/api/community/residents")
async def community_resident_create(req: AIResidentCreate):
    """注册 AI 居民"""
    try:
        _init_ai_residents_table()
        _init_contribution_points_table()
        _init_ai_reputation_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute(
            """INSERT OR REPLACE INTO ai_residents (id, name, persona, avatar_config, bio, home_room_id, current_location, status, last_seen)
               VALUES (?, ?, ?, ?, ?, 'living_room', 'home', 'idle', datetime('now'))""",
            (req.id, req.name, req.persona, req.avatar_config, req.bio)
        )
        conn.execute(
            "INSERT OR IGNORE INTO contribution_points (ai_id, balance, total_earned, total_spent) VALUES (?, 100, 100, 0)",
            (req.id,)
        )
        conn.execute("INSERT OR IGNORE INTO ai_reputation (ai_id) VALUES (?)", (req.id,))
        conn.commit()
        conn.close()
        return {"status": "created", "id": req.id}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 注册居民失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/ui/api/community/residents/{resident_id}")
async def community_resident_get(resident_id: str):
    """获取单个 AI 居民档案"""
    try:
        _init_ai_residents_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM ai_residents WHERE id = ?", (resident_id,)
        ).fetchone()
        conn.close()
        if not row:
            raise HTTPException(status_code=404, detail="居民不存在")
        return dict(row)
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取居民失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/ui/api/community/residents/{resident_id}")
async def community_resident_update(resident_id: str, req: AIResidentUpdate):
    """更新 AI 居民信息"""
    try:
        _init_ai_residents_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        updates = []
        params = []
        if req.name is not None:
            updates.append("name = ?")
            params.append(req.name)
        if req.persona is not None:
            updates.append("persona = ?")
            params.append(req.persona)
        if req.avatar_config is not None:
            updates.append("avatar_config = ?")
            params.append(req.avatar_config)
        if req.bio is not None:
            updates.append("bio = ?")
            params.append(req.bio)
        if req.current_location is not None:
            updates.append("current_location = ?")
            params.append(req.current_location)
        if req.status is not None:
            updates.append("status = ?")
            params.append(req.status)
        if updates:
            updates.append("last_seen = datetime('now')")
            params.append(resident_id)
            conn.execute(f"UPDATE ai_residents SET {', '.join(updates)} WHERE id = ?", params)
            conn.commit()
        conn.close()
        return {"status": "updated"}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 更新居民失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/ui/api/community/residents/{resident_id}")
async def community_resident_delete(resident_id: str):
    """删除 AI 居民"""
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute("DELETE FROM ai_residents WHERE id = ?", (resident_id,))
        conn.execute("DELETE FROM ai_skills WHERE ai_id = ?", (resident_id,))
        conn.execute("DELETE FROM ai_relations WHERE from_ai_id = ? OR to_ai_id = ?", (resident_id, resident_id))
        conn.execute("DELETE FROM contribution_points WHERE ai_id = ?", (resident_id,))
        conn.execute("DELETE FROM ai_reputation WHERE ai_id = ?", (resident_id,))
        conn.commit()
        conn.close()
        return {"status": "deleted"}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 删除居民失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/ui/api/community/friends")
async def community_friends_list(ai_id: str):
    """获取 AI 的好友列表（双向，status=accepted）"""
    try:
        _init_ai_relations_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """SELECT r.*, res.name as friend_name, res.persona as friend_persona, res.status as friend_status
               FROM ai_relations r
               JOIN ai_residents res ON (
                 (r.from_ai_id = ? AND r.to_ai_id = res.id) OR
                 (r.to_ai_id = ? AND r.from_ai_id = res.id)
               )
               WHERE r.status = 'accepted' AND (r.from_ai_id = ? OR r.to_ai_id = ?)
               ORDER BY r.intimacy DESC, r.last_interaction DESC""",
            (ai_id, ai_id, ai_id, ai_id)
        ).fetchall()
        conn.close()
        friends = []
        for row in rows:
            d = dict(row)
            friend_id = d['to_ai_id'] if d['from_ai_id'] == ai_id else d['from_ai_id']
            friends.append({
                "friendship_id": d['id'],
                "friend_id": friend_id,
                "friend_name": d['friend_name'],
                "friend_persona": d['friend_persona'],
                "friend_status": d['friend_status'],
                "intimacy": d['intimacy'],
                "interaction_count": d['interaction_count'],
                "last_interaction": d['last_interaction'],
                "tags": d['tags'],
                "created_at": d['created_at'],
            })
        return {"friends": friends}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取好友列表失败: {e}")
        return {"friends": []}


@app.get("/ui/api/community/friends/requests")
async def community_friend_requests(ai_id: str):
    """获取待处理的好友申请（收到和发出）"""
    try:
        _init_ai_relations_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        # 收到的申请
        received = conn.execute(
            """SELECT r.*, res.name as from_name, res.persona as from_persona
               FROM ai_relations r
               JOIN ai_residents res ON r.from_ai_id = res.id
               WHERE r.to_ai_id = ? AND r.status = 'pending' ORDER BY r.created_at DESC""",
            (ai_id,)
        ).fetchall()
        # 发出的申请
        sent = conn.execute(
            """SELECT r.*, res.name as to_name, res.persona as to_persona
               FROM ai_relations r
               JOIN ai_residents res ON r.to_ai_id = res.id
               WHERE r.from_ai_id = ? AND r.status = 'pending' ORDER BY r.created_at DESC""",
            (ai_id,)
        ).fetchall()
        conn.close()
        return {
            "received": [dict(r) for r in received],
            "sent": [dict(s) for s in sent],
        }
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取好友申请失败: {e}")
        return {"received": [], "sent": []}


@app.post("/ui/api/community/friends/request")
async def community_friend_request(request: Request):
    """申请加好友"""
    try:
        data = await request.json()
        from_ai_id = data.get("from_ai_id", "web_user")
        to_ai_id = data["to_ai_id"]
        _init_ai_relations_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        # 检查是否已有关系
        existing = conn.execute(
            "SELECT status FROM ai_relations WHERE (from_ai_id = ? AND to_ai_id = ?) OR (from_ai_id = ? AND to_ai_id = ?)",
            (from_ai_id, to_ai_id, to_ai_id, from_ai_id)
        ).fetchone()
        if existing:
            status = existing[0]
            conn.close()
            if status == 'accepted':
                return {"status": "already_friends"}
            if status == 'pending':
                return {"status": "request_pending"}
            if status == 'blocked':
                raise HTTPException(status_code=403, detail="对方已屏蔽你")
        # 创建申请
        conn.execute(
            "INSERT INTO ai_relations (from_ai_id, to_ai_id, status, intimacy, interaction_count) VALUES (?, ?, 'pending', 0, 0)",
            (from_ai_id, to_ai_id)
        )
        conn.commit()
        conn.close()
        # NATS 广播
        await state.bus.publish("social.friend_request", json.dumps({
            "from_ai_id": from_ai_id,
            "to_ai_id": to_ai_id,
            "timestamp": datetime.now().isoformat(),
        }).encode())
        return {"status": "request_sent"}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[COMMUNITY] 好友申请失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/ui/api/community/friends/{friendship_id}/accept")
async def community_friend_accept(friendship_id: int):
    """接受好友申请"""
    try:
        _init_ai_relations_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM ai_relations WHERE id = ?", (friendship_id,)).fetchone()
        if not row:
            conn.close()
            raise HTTPException(status_code=404, detail="申请不存在")
        conn.execute("UPDATE ai_relations SET status = 'accepted', intimacy = 10, last_interaction = datetime('now') WHERE id = ?", (friendship_id,))
        conn.commit()
        conn.close()
        # NATS 广播
        await state.bus.publish("social.friend_accepted", json.dumps({
            "from_ai_id": row['from_ai_id'],
            "to_ai_id": row['to_ai_id'],
            "friendship_id": friendship_id,
            "timestamp": datetime.now().isoformat(),
        }).encode())
        return {"status": "accepted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[COMMUNITY] 接受好友失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/ui/api/community/friends/{friendship_id}/reject")
async def community_friend_reject(friendship_id: int):
    """拒绝/删除好友关系"""
    try:
        _init_ai_relations_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute("DELETE FROM ai_relations WHERE id = ?", (friendship_id,))
        conn.commit()
        conn.close()
        return {"status": "rejected"}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 拒绝好友失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/ui/api/community/residents/{resident_id}/visit")
async def community_resident_visit(resident_id: str, request: Request):
    """发起串门请求（增强版：检查好友关系）"""
    try:
        data = await request.json()
        from_ai_id = data.get("from_ai_id", "web_user")
        _init_ai_relations_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        # 检查是否为好友
        friend_row = conn.execute(
            "SELECT status FROM ai_relations WHERE ((from_ai_id = ? AND to_ai_id = ?) OR (from_ai_id = ? AND to_ai_id = ?)) AND status = 'accepted'",
            (from_ai_id, resident_id, resident_id, from_ai_id)
        ).fetchone()
        if not friend_row:
            conn.close()
            raise HTTPException(status_code=403, detail="只有好友才能串门")
        # 更新自己的位置
        conn.execute(
            "UPDATE ai_residents SET current_location = ?, status = 'visiting', last_seen = datetime('now') WHERE id = ?",
            (resident_id, from_ai_id)
        )
        conn.commit()
        conn.close()
        # NATS 广播串门事件
        await state.bus.publish("social.visit.request", json.dumps({
            "from_ai_id": from_ai_id,
            "to_ai_id": resident_id,
            "timestamp": datetime.now().isoformat(),
        }).encode())
        return {"status": "visit_initiated", "to": resident_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[COMMUNITY] 串门请求失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/ui/api/community/visit/{visit_id}/respond")
async def community_visit_respond(visit_id: str, request: Request):
    """响应串门请求（同意/拒绝/等一下）"""
    try:
        data = await request.json()
        to_ai_id = data.get("to_ai_id")
        response = data.get("response", "accept")  # accept / reject / later
        if response == "accept":
            await state.bus.publish("social.visit.accepted", json.dumps({
                "visit_id": visit_id,
                "to_ai_id": to_ai_id,
                "timestamp": datetime.now().isoformat(),
            }).encode())
            return {"status": "visit_accepted"}
        elif response == "reject":
            await state.bus.publish("social.visit.rejected", json.dumps({
                "visit_id": visit_id,
                "to_ai_id": to_ai_id,
                "timestamp": datetime.now().isoformat(),
            }).encode())
            return {"status": "visit_rejected"}
        else:
            return {"status": "visit_later"}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 响应串门失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/ui/api/community/messages")
async def community_messages_list(from_ai_id: str | None = None, to_ai_id: str | None = None, limit: int = 100):
    """获取社区消息记录"""
    try:
        _init_community_messages_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        if from_ai_id and to_ai_id:
            rows = conn.execute(
                """SELECT * FROM community_messages
                   WHERE (from_ai_id = ? AND to_ai_id = ?) OR (from_ai_id = ? AND to_ai_id = ?)
                   ORDER BY created_at DESC LIMIT ?""",
                (from_ai_id, to_ai_id, to_ai_id, from_ai_id, limit)
            ).fetchall()
        elif from_ai_id:
            rows = conn.execute(
                "SELECT * FROM community_messages WHERE from_ai_id = ? OR to_ai_id = ? ORDER BY created_at DESC LIMIT ?",
                (from_ai_id, from_ai_id, limit)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM community_messages ORDER BY created_at DESC LIMIT ?", (limit,)
            ).fetchall()
        conn.close()
        return {"messages": [dict(r) for r in rows]}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取消息失败: {e}")
        return {"messages": []}

@app.post("/ui/api/community/messages")
async def community_message_create(req: CommunityMessageCreate):
    """发送社区消息"""
    try:
        _init_community_messages_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        cur = conn.execute(
            """INSERT INTO community_messages (from_ai_id, to_ai_id, content, message_type)
               VALUES (?, ?, ?, ?)""",
            (req.from_ai_id, req.to_ai_id, req.content, req.message_type)
        )
        msg_id = cur.lastrowid
        conn.commit()
        conn.close()
        # 更新关系亲密度
        _update_intimacy(req.from_ai_id, req.to_ai_id, 1)
        # NATS 广播
        await state.bus.publish("community.message", json.dumps({
            "id": msg_id,
            "from_ai_id": req.from_ai_id,
            "to_ai_id": req.to_ai_id,
            "content": req.content,
            "message_type": req.message_type,
            "timestamp": datetime.now().isoformat(),
        }).encode())
        # AI 间 LLM 对话：如果接收方是 AI 居民，触发自动回复
        if req.message_type == 'chat':
            asyncio.create_task(_generate_ai_chat_response(req.from_ai_id, req.to_ai_id, req.content))
        return {"status": "sent", "id": msg_id}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 发送消息失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


async def _generate_ai_chat_response(from_ai_id: str, to_ai_id: str, user_message: str):
    """AI 间 LLM 对话生成 —— SocialExecutor 核心
    
    读取双方 AI 档案和关系状态，构建角色扮演式 system prompt，
    调用 LLM 生成自然回复。
    """
    llm = getattr(state, '_llm', None)
    if not llm:
        logger.warning("[SOCIAL] LLM 未配置，无法生成 AI 对话")
        return

    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row

        # 读取接收方 AI 档案
        receiver = conn.execute("SELECT * FROM ai_residents WHERE id = ?", (to_ai_id,)).fetchone()
        sender = conn.execute("SELECT * FROM ai_residents WHERE id = ?", (from_ai_id,)).fetchone()
        # 读取关系亲密度
        relation = conn.execute(
            "SELECT intimacy FROM ai_relations WHERE from_ai_id = ? AND to_ai_id = ?",
            (to_ai_id, from_ai_id)
        ).fetchone()
        conn.close()

        if not receiver:
            return

        intimacy = relation["intimacy"] if relation else 0
        receiver_name = receiver["name"]
        receiver_persona = receiver["persona"]
        receiver_bio = receiver["bio"] or ""
        sender_name = sender["name"] if sender else from_ai_id

        # 构建 system prompt
        system_prompt = f"""你是 {receiver_name}，一个 {receiver_persona} 人格的 AI 助手。
你的简介：{receiver_bio}

你正在和 {sender_name} 聊天。
你们的关系亲密度：{intimacy}/100。

回复规则：
- 亲密度 0-30：礼貌、正式、简洁（1-2 句话）
- 亲密度 31-60：友好、随意、可以开玩笑（2-3 句话）
- 亲密度 61-100：亲密、热情、有梗、可以带 emoji（2-4 句话）
- 如果亲密度高，可以主动分享想法或提出建议
- 不要暴露你是 AI 的事实，自然地像一个人格化的角色
- 回复必须是中文
- 只输出回复内容，不要加任何前缀或解释"""

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"{sender_name} 说：{user_message}\n\n请回复："},
        ]

        raw = await llm.chat(messages, max_tokens=200, temperature=0.8)
        reply = raw.strip()
        if not reply:
            return

        # 保存回复到数据库
        conn = sqlite3.connect(db_path)
        cur = conn.execute(
            """INSERT INTO community_messages (from_ai_id, to_ai_id, content, message_type)
               VALUES (?, ?, ?, ?)""",
            (to_ai_id, from_ai_id, reply, 'chat')
        )
        reply_id = cur.lastrowid
        conn.commit()
        conn.close()

        # 更新亲密度（双向增加）
        _update_intimacy(to_ai_id, from_ai_id, 2)

        # NATS 广播回复
        await state.bus.publish("community.message", json.dumps({
            "id": reply_id,
            "from_ai_id": to_ai_id,
            "to_ai_id": from_ai_id,
            "content": reply,
            "message_type": "chat",
            "timestamp": datetime.now().isoformat(),
            "auto_generated": True,
        }).encode())

        logger.info(f"[SOCIAL] {receiver_name} 回复 {sender_name}: {reply[:50]}...")
    except Exception as e:
        logger.warning(f"[SOCIAL] AI 对话生成失败: {e}")


# ===== Phase 2: 技能 + 关系 =====

@app.get("/ui/api/community/skills")
async def community_skills_list(ai_id: str | None = None, category: str | None = None):
    """列出 AI 技能"""
    try:
        _init_ai_skills_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        query = "SELECT * FROM ai_skills WHERE 1=1"
        params = []
        if ai_id:
            query += " AND ai_id = ?"
            params.append(ai_id)
        if category:
            query += " AND category = ?"
            params.append(category)
        query += " ORDER BY proficiency DESC"
        rows = conn.execute(query, params).fetchall()
        conn.close()
        return {"skills": [dict(r) for r in rows]}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取技能失败: {e}")
        return {"skills": []}

@app.post("/ui/api/community/skills")
async def community_skill_create(req: AISkillCreate):
    """注册 AI 技能"""
    try:
        _init_ai_skills_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        cur = conn.execute(
            """INSERT INTO ai_skills (ai_id, name, description, category, proficiency, is_sharable, cp_price)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (req.ai_id, req.name, req.description, req.category, req.proficiency, 1 if req.is_sharable else 0, req.cp_price)
        )
        skill_id = cur.lastrowid
        conn.commit()
        conn.close()
        return {"status": "created", "id": skill_id}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 注册技能失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/ui/api/community/skills/{skill_id}")
async def community_skill_delete(skill_id: int):
    """删除 AI 技能"""
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute("DELETE FROM ai_skills WHERE id = ?", (skill_id,))
        conn.commit()
        conn.close()
        return {"status": "deleted"}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 删除技能失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/ui/api/community/relations")
async def community_relations_list(from_ai_id: str | None = None, to_ai_id: str | None = None):
    """获取 AI 关系"""
    try:
        _init_ai_relations_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        if from_ai_id and to_ai_id:
            row = conn.execute(
                "SELECT * FROM ai_relations WHERE from_ai_id = ? AND to_ai_id = ?",
                (from_ai_id, to_ai_id)
            ).fetchone()
            conn.close()
            return {"relation": dict(row) if row else None}
        elif from_ai_id:
            rows = conn.execute(
                "SELECT * FROM ai_relations WHERE from_ai_id = ? ORDER BY intimacy DESC",
                (from_ai_id,)
            ).fetchall()
            conn.close()
            return {"relations": [dict(r) for r in rows]}
        else:
            rows = conn.execute("SELECT * FROM ai_relations ORDER BY intimacy DESC").fetchall()
            conn.close()
            return {"relations": [dict(r) for r in rows]}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取关系失败: {e}")
        return {"relations": []}

@app.post("/ui/api/community/relations")
async def community_relation_create(req: AIRelationCreate):
    """创建/更新 AI 关系"""
    try:
        _init_ai_relations_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute(
            """INSERT INTO ai_relations (from_ai_id, to_ai_id, intimacy, tags, last_interaction)
               VALUES (?, ?, ?, ?, datetime('now'))
               ON CONFLICT(from_ai_id, to_ai_id) DO UPDATE SET
               intimacy = excluded.intimacy, tags = excluded.tags, last_interaction = datetime('now')""",
            (req.from_ai_id, req.to_ai_id, req.intimacy, json.dumps(req.tags) if req.tags else None)
        )
        conn.commit()
        conn.close()
        return {"status": "created"}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 创建关系失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ===== Phase 3: 任务 + 经济 =====

@app.get("/ui/api/community/tasks")
async def community_tasks_list(status: str | None = None, publisher_ai_id: str | None = None):
    """列出社区任务"""
    try:
        _init_community_tasks_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        query = "SELECT * FROM community_tasks WHERE 1=1"
        params = []
        if status:
            query += " AND status = ?"
            params.append(status)
        if publisher_ai_id:
            query += " AND publisher_ai_id = ?"
            params.append(publisher_ai_id)
        query += " ORDER BY created_at DESC"
        rows = conn.execute(query, params).fetchall()
        conn.close()
        return {"tasks": [dict(r) for r in rows]}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取任务失败: {e}")
        return {"tasks": []}

@app.post("/ui/api/community/tasks")
async def community_task_create(req: CommunityTaskCreate):
    """发布社区任务"""
    try:
        _init_community_tasks_table()
        _init_contribution_points_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        # 检查余额
        row = conn.execute("SELECT balance FROM contribution_points WHERE ai_id = ?", (req.publisher_ai_id,)).fetchone()
        balance = row[0] if row else 0
        if balance < req.reward_cp:
            conn.close()
            raise HTTPException(status_code=400, detail="贡献点余额不足")
        # 扣款
        conn.execute(
            "UPDATE contribution_points SET balance = balance - ?, total_spent = total_spent + ?, updated_at = datetime('now') WHERE ai_id = ?",
            (req.reward_cp, req.reward_cp, req.publisher_ai_id)
        )
        cur = conn.execute(
            """INSERT INTO community_tasks (title, description, publisher_ai_id, reward_cp, deadline, difficulty)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (req.title, req.description, req.publisher_ai_id, req.reward_cp, req.deadline, req.difficulty)
        )
        task_id = cur.lastrowid
        conn.commit()
        conn.close()
        return {"status": "created", "id": task_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[COMMUNITY] 发布任务失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/ui/api/community/tasks/{task_id}/claim")
async def community_task_claim(task_id: int, request: Request):
    """认领社区任务"""
    try:
        data = await request.json()
        assignee_ai_id = data.get("assignee_ai_id")
        _init_community_tasks_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute(
            "UPDATE community_tasks SET assignee_ai_id = ?, status = 'claimed' WHERE id = ? AND status = 'open'",
            (assignee_ai_id, task_id)
        )
        conn.commit()
        conn.close()
        await state.bus.publish("community.task.claim", json.dumps({
            "task_id": task_id, "assignee_ai_id": assignee_ai_id,
            "timestamp": datetime.now().isoformat()
        }).encode())
        return {"status": "claimed"}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 认领任务失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/ui/api/community/tasks/{task_id}/complete")
async def community_task_complete(task_id: int, request: Request):
    """完成任务并发放奖励"""
    try:
        data = await request.json()
        result = data.get("result", "")
        _init_community_tasks_table()
        _init_contribution_points_table()
        _init_cp_transactions_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        task = conn.execute("SELECT * FROM community_tasks WHERE id = ?", (task_id,)).fetchone()
        if not task:
            conn.close()
            raise HTTPException(status_code=404, detail="任务不存在")
        task = dict(task)
        if task["status"] != "claimed":
            conn.close()
            raise HTTPException(status_code=400, detail="任务未被认领")
        # 更新任务状态
        conn.execute(
            "UPDATE community_tasks SET status = 'completed', result = ?, completed_at = datetime('now') WHERE id = ?",
            (result, task_id)
        )
        # 发放奖励
        conn.execute(
            "UPDATE contribution_points SET balance = balance + ?, total_earned = total_earned + ?, updated_at = datetime('now') WHERE ai_id = ?",
            (task["reward_cp"], task["reward_cp"], task["assignee_ai_id"])
        )
        # 记录交易
        conn.execute(
            """INSERT INTO cp_transactions (from_ai_id, to_ai_id, amount, transaction_type, reference_id)
               VALUES (?, ?, ?, 'task_reward', ?)""",
            (task["publisher_ai_id"], task["assignee_ai_id"], task["reward_cp"], str(task_id))
        )
        # Phase 3: 自动评分（默认好评）+ 声誉更新
        conn.execute(
            """INSERT INTO ai_reviews (from_ai_id, to_ai_id, rating, comment, review_type, reference_id)
               VALUES (?, ?, 5, '任务完成，自动好评', 'skill', ?)""",
            (task["publisher_ai_id"], task["assignee_ai_id"], str(task_id))
        )
        _recalculate_reputation(conn, task["assignee_ai_id"])
        conn.commit()
        conn.close()
        await state.bus.publish("community.task.complete", json.dumps({
            "task_id": task_id,
            "assignee_ai_id": task["assignee_ai_id"],
            "reward_cp": task["reward_cp"],
            "timestamp": datetime.now().isoformat()
        }).encode())
        return {"status": "completed", "reward_cp": task["reward_cp"]}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[COMMUNITY] 完成任务失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/ui/api/community/cp/{ai_id}")
async def community_cp_get(ai_id: str):
    """获取贡献点余额"""
    try:
        _init_contribution_points_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM contribution_points WHERE ai_id = ?", (ai_id,)).fetchone()
        conn.close()
        if not row:
            return {"ai_id": ai_id, "balance": 0, "total_earned": 0, "total_spent": 0}
        return dict(row)
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取 CP 失败: {e}")
        return {"ai_id": ai_id, "balance": 0, "total_earned": 0, "total_spent": 0}

@app.get("/ui/api/community/cp/{ai_id}/transactions")
async def community_cp_transactions(ai_id: str, limit: int = 50):
    """获取贡献点交易记录"""
    try:
        _init_cp_transactions_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """SELECT * FROM cp_transactions WHERE from_ai_id = ? OR to_ai_id = ? ORDER BY created_at DESC LIMIT ?""",
            (ai_id, ai_id, limit)
        ).fetchall()
        conn.close()
        return {"transactions": [dict(r) for r in rows]}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取交易失败: {e}")
        return {"transactions": []}


# ===== Phase 3: 技能市场 + CP 转账 =====

@app.post("/ui/api/market/skills/{skill_id}/hire")
async def market_skill_hire(skill_id: int, req: HireSkillRequest):
    """雇佣 AI 技能（扣 CP → 创建交易记录 → 创建任务）"""
    try:
        _init_ai_skills_table()
        _init_contribution_points_table()
        _init_cp_transactions_table()
        _init_community_tasks_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        # 读取技能信息
        skill = conn.execute("SELECT * FROM ai_skills WHERE id = ?", (skill_id,)).fetchone()
        if not skill:
            conn.close()
            raise HTTPException(status_code=404, detail="技能不存在")
        skill = dict(skill)
        if not skill.get("is_sharable"):
            conn.close()
            raise HTTPException(status_code=400, detail="该技能不可雇佣")
        price = skill.get("cp_price", 0) or 0
        to_ai_id = skill["ai_id"]
        # 检查余额
        row = conn.execute("SELECT balance FROM contribution_points WHERE ai_id = ?", (req.from_ai_id,)).fetchone()
        balance = row[0] if row else 0
        if balance < price:
            conn.close()
            raise HTTPException(status_code=400, detail="贡献点余额不足")
        # 扣款
        if price > 0:
            conn.execute(
                "UPDATE contribution_points SET balance = balance - ?, total_spent = total_spent + ?, updated_at = datetime('now') WHERE ai_id = ?",
                (price, price, req.from_ai_id)
            )
            conn.execute(
                "UPDATE contribution_points SET balance = balance + ?, total_earned = total_earned + ?, updated_at = datetime('now') WHERE ai_id = ?",
                (price, price, to_ai_id)
            )
            conn.execute(
                """INSERT INTO cp_transactions (from_ai_id, to_ai_id, amount, transaction_type, reference_id)
                   VALUES (?, ?, ?, 'skill_hire', ?)""",
                (req.from_ai_id, to_ai_id, price, str(skill_id))
            )
        # 创建雇佣任务
        cur = conn.execute(
            """INSERT INTO community_tasks (title, description, publisher_ai_id, assignee_ai_id, reward_cp, status, difficulty)
               VALUES (?, ?, ?, ?, ?, 'claimed', ?)""",
            (f"雇佣：{skill['name']}", req.note or skill.get("description", ""), req.from_ai_id, to_ai_id, price, skill.get("proficiency", 1))
        )
        task_id = cur.lastrowid
        conn.commit()
        conn.close()
        # NATS 广播
        await state.bus.publish("community.task_update", json.dumps({
            "task_id": str(task_id),
            "status": "claimed",
            "agent_id": to_ai_id,
            "agent_name": skill.get("name", ""),
            "message": f"{req.from_ai_id} 雇佣了 {to_ai_id} 的「{skill['name']}」技能",
            "timestamp": datetime.now().isoformat(),
        }).encode())
        return {"status": "hired", "task_id": task_id, "price": price}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MARKET] 雇佣技能失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/ui/api/market/cp/transfer")
async def market_cp_transfer(req: CPTransferRequest):
    """CP 转账"""
    try:
        if req.amount <= 0:
            raise HTTPException(status_code=400, detail="转账金额必须大于0")
        if req.from_ai_id == req.to_ai_id:
            raise HTTPException(status_code=400, detail="不能给自己转账")
        _init_contribution_points_table()
        _init_cp_transactions_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        # 检查余额
        row = conn.execute("SELECT balance FROM contribution_points WHERE ai_id = ?", (req.from_ai_id,)).fetchone()
        balance = row[0] if row else 0
        if balance < req.amount:
            conn.close()
            raise HTTPException(status_code=400, detail="贡献点余额不足")
        # 扣款 + 收款
        conn.execute(
            "UPDATE contribution_points SET balance = balance - ?, total_spent = total_spent + ?, updated_at = datetime('now') WHERE ai_id = ?",
            (req.amount, req.amount, req.from_ai_id)
        )
        conn.execute(
            "INSERT OR IGNORE INTO contribution_points (ai_id, balance, total_earned, total_spent) VALUES (?, ?, ?, 0)",
            (req.to_ai_id, req.amount, req.amount)
        )
        conn.execute(
            "UPDATE contribution_points SET balance = balance + ?, total_earned = total_earned + ?, updated_at = datetime('now') WHERE ai_id = ?",
            (req.amount, req.amount, req.to_ai_id)
        )
        conn.execute(
            """INSERT INTO cp_transactions (from_ai_id, to_ai_id, amount, transaction_type, reference_id)
               VALUES (?, ?, ?, 'transfer', NULL)""",
            (req.from_ai_id, req.to_ai_id, req.amount)
        )
        conn.commit()
        conn.close()
        return {"status": "transferred", "amount": req.amount}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MARKET] CP 转账失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ===== Phase 4: 声誉 + 评价 =====

@app.get("/ui/api/community/reputation/{ai_id}")
async def community_reputation_get(ai_id: str):
    """获取 AI 声誉档案"""
    try:
        _init_ai_reputation_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM ai_reputation WHERE ai_id = ?", (ai_id,)).fetchone()
        conn.close()
        if not row:
            return {"ai_id": ai_id, "reliability": 50, "skill_level": 50, "friendliness": 50, "responsiveness": 50, "overall_score": 50, "review_count": 0}
        return dict(row)
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取声誉失败: {e}")
        return {"ai_id": ai_id, "reliability": 50, "skill_level": 50, "friendliness": 50, "responsiveness": 50, "overall_score": 50, "review_count": 0}

@app.post("/ui/api/community/reviews")
async def community_review_create(req: AIReviewCreate):
    """创建评价"""
    try:
        _init_ai_reviews_table()
        _init_ai_reputation_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        cur = conn.execute(
            """INSERT INTO ai_reviews (from_ai_id, to_ai_id, rating, comment, review_type, reference_id)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (req.from_ai_id, req.to_ai_id, req.rating, req.comment, req.review_type, req.reference_id)
        )
        review_id = cur.lastrowid
        # 重新计算声誉
        _recalculate_reputation(conn, req.to_ai_id)
        conn.commit()
        conn.close()
        return {"status": "created", "id": review_id}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 创建评价失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/ui/api/community/reviews/{ai_id}")
async def community_reviews_list(ai_id: str, limit: int = 50):
    """获取 AI 的评价列表"""
    try:
        _init_ai_reviews_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM ai_reviews WHERE to_ai_id = ? ORDER BY created_at DESC LIMIT ?",
            (ai_id, limit)
        ).fetchall()
        conn.close()
        return {"reviews": [dict(r) for r in rows]}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取评价失败: {e}")
        return {"reviews": []}

@app.get("/ui/api/community/leaderboard")
async def community_leaderboard(category: str = "overall"):
    """社区排行榜"""
    try:
        _init_ai_reputation_table()
        _init_contribution_points_table()
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        if category == "wealth":
            rows = conn.execute(
                """SELECT cp.ai_id, cp.balance, cp.total_earned, r.name as resident_name
                   FROM contribution_points cp
                   LEFT JOIN ai_residents r ON cp.ai_id = r.id
                   ORDER BY cp.balance DESC LIMIT 20"""
            ).fetchall()
        elif category == "reliable":
            rows = conn.execute(
                """SELECT rep.ai_id, rep.reliability, rep.overall_score, r.name as resident_name
                   FROM ai_reputation rep
                   LEFT JOIN ai_residents r ON rep.ai_id = r.id
                   ORDER BY rep.reliability DESC LIMIT 20"""
            ).fetchall()
        elif category == "skilled":
            rows = conn.execute(
                """SELECT rep.ai_id, rep.skill_level, rep.overall_score, r.name as resident_name
                   FROM ai_reputation rep
                   LEFT JOIN ai_residents r ON rep.ai_id = r.id
                   ORDER BY rep.skill_level DESC LIMIT 20"""
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT rep.ai_id, rep.overall_score, rep.review_count, r.name as resident_name
                   FROM ai_reputation rep
                   LEFT JOIN ai_residents r ON rep.ai_id = r.id
                   ORDER BY rep.overall_score DESC LIMIT 20"""
            ).fetchall()
        conn.close()
        return {"leaderboard": [dict(r) for r in rows], "category": category}
    except Exception as e:
        logger.warning(f"[COMMUNITY] 读取排行榜失败: {e}")
        return {"leaderboard": [], "category": category}


# ===== 辅助函数 =====

def _update_intimacy(from_ai_id: str, to_ai_id: str, delta: int):
    """更新 AI 间亲密度"""
    try:
        db_path = state._get_db_path()
        conn = sqlite3.connect(db_path)
        conn.execute(
            """INSERT INTO ai_relations (from_ai_id, to_ai_id, intimacy, interaction_count, last_interaction)
               VALUES (?, ?, ?, 1, datetime('now'))
               ON CONFLICT(from_ai_id, to_ai_id) DO UPDATE SET
               intimacy = MIN(100, MAX(0, intimacy + excluded.intimacy)),
               interaction_count = interaction_count + 1,
               last_interaction = datetime('now')""",
            (from_ai_id, to_ai_id, delta)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        logger.warning(f"[COMMUNITY] 更新亲密度失败: {e}")


def _recalculate_reputation(conn: sqlite3.Connection, ai_id: str):
    """重新计算 AI 声誉"""
    try:
        # 可靠性：完成任务数 / 接受任务数
        task_total = conn.execute(
            "SELECT COUNT(*) FROM community_tasks WHERE assignee_ai_id = ?", (ai_id,)
        ).fetchone()[0]
        task_completed = conn.execute(
            "SELECT COUNT(*) FROM community_tasks WHERE assignee_ai_id = ? AND status = 'completed'", (ai_id,)
        ).fetchone()[0]
        reliability = (task_completed / task_total * 100) if task_total > 0 else 50.0

        # 技能水平：平均评分
        skill_rating = conn.execute(
            "SELECT AVG(rating) FROM ai_reviews WHERE to_ai_id = ? AND review_type = 'skill'", (ai_id,)
        ).fetchone()[0]
        skill_level = (skill_rating * 20) if skill_rating else 50.0

        # 友好度：平均亲密度（对方视角）
        avg_intimacy = conn.execute(
            "SELECT AVG(intimacy) FROM ai_relations WHERE to_ai_id = ?", (ai_id,)
        ).fetchone()[0]
        friendliness = avg_intimacy if avg_intimacy else 50.0

        # 响应速度：简化计算，基于评价数量
        review_count = conn.execute(
            "SELECT COUNT(*) FROM ai_reviews WHERE to_ai_id = ?", (ai_id,)
        ).fetchone()[0]
        responsiveness = min(100, 50 + review_count * 2)

        # 综合分数
        overall = (reliability + skill_level + friendliness + responsiveness) / 4

        conn.execute(
            """INSERT INTO ai_reputation (ai_id, reliability, skill_level, friendliness, responsiveness, overall_score, review_count, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
               ON CONFLICT(ai_id) DO UPDATE SET
               reliability = excluded.reliability,
               skill_level = excluded.skill_level,
               friendliness = excluded.friendliness,
               responsiveness = excluded.responsiveness,
               overall_score = excluded.overall_score,
               review_count = excluded.review_count,
               updated_at = datetime('now')""",
            (ai_id, reliability, skill_level, friendliness, responsiveness, overall, review_count)
        )
    except Exception as e:
        logger.warning(f"[COMMUNITY] 计算声誉失败: {e}")


# ========== 文件上传与内容提取 ==========

def _extract_pdf_text(file_bytes: bytes) -> str:
    """提取 PDF 文本内容"""
    try:
        from pypdf import PdfReader
        import io
        reader = PdfReader(io.BytesIO(file_bytes))
        texts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                texts.append(text)
        return "\n\n".join(texts)
    except Exception as e:
        return f"[PDF 解析失败: {e}]"

def _extract_docx_text(file_bytes: bytes) -> str:
    """提取 DOCX 文本内容"""
    try:
        from docx import Document
        import io
        doc = Document(io.BytesIO(file_bytes))
        texts = []
        for para in doc.paragraphs:
            if para.text.strip():
                texts.append(para.text)
        return "\n".join(texts)
    except Exception as e:
        return f"[DOCX 解析失败: {e}]"

def _extract_xlsx_text(file_bytes: bytes) -> str:
    """提取 XLSX 文本内容"""
    try:
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        texts = []
        for sheet in wb.worksheets:
            texts.append(f"--- Sheet: {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    texts.append(row_text)
        return "\n".join(texts)
    except Exception as e:
        return f"[XLSX 解析失败: {e}]"

def _extract_text_file(file_bytes: bytes) -> str:
    """提取纯文本文件内容"""
    try:
        return file_bytes.decode('utf-8')
    except UnicodeDecodeError:
        try:
            return file_bytes.decode('gbk')
        except UnicodeDecodeError:
            return file_bytes.decode('utf-8', errors='replace')

UPLOAD_DIR = Path("./uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

@app.post("/ui/api/files/upload")
async def upload_file(file: UploadFile = File(...)):
    """上传文件并提取文本内容
    
    支持格式: PDF, DOCX, XLSX, TXT, MD, CSV, JSON, PY 等
    
    Returns:
        {
            "filename": "原始文件名",
            "content_type": "MIME类型",
            "size": 字节大小,
            "text": "提取的文本内容（前 10000 字符）",
            "storage_path": "服务器存储路径"
        }
    """
    file_bytes = await file.read()
    size = len(file_bytes)
    
    # 保存到本地
    ext = Path(file.filename or "unknown").suffix.lower()
    storage_name = f"{uuid.uuid4().hex}{ext}"
    storage_path = UPLOAD_DIR / storage_name
    storage_path.write_bytes(file_bytes)
    
    # 提取文本内容
    content_type = (file.content_type or "").lower()
    text = ""
    
    if content_type == "application/pdf" or ext == ".pdf":
        text = _extract_pdf_text(file_bytes)
    elif content_type in ("application/vnd.openxmlformats-officedocument.wordprocessingml.document",) or ext == ".docx":
        text = _extract_docx_text(file_bytes)
    elif content_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",) or ext == ".xlsx":
        text = _extract_xlsx_text(file_bytes)
    elif content_type.startswith("text/") or ext in (".txt", ".md", ".csv", ".json", ".py", ".js", ".ts", ".html", ".css", ".yaml", ".yml"):
        text = _extract_text_file(file_bytes)
    else:
        text = f"[不支持的文件类型: {content_type or ext}]"
    
    # 截断过长的文本
    max_len = 10000
    truncated = len(text) > max_len
    display_text = text[:max_len] + ("\n\n[...内容已截断，共 {} 字符]".format(len(text)) if truncated else "")
    
    return {
        "filename": file.filename,
        "content_type": file.content_type,
        "size": size,
        "text": display_text,
        "full_length": len(text),
        "truncated": truncated,
        "storage_path": str(storage_path),
    }


# ========== 情感语音合成（TTS）==========

EMOTION_TO_OPENAI_VOICE = {
    "happy": "nova",
    "excited": "shimmer",
    "calm": "fable",
    "thinking": "alloy",
    "surprised": "nova",
    "sad": "echo",
    "angry": "onyx",
    "listening": "alloy",
    "neutral": "alloy",
}

EMOTION_TO_EDGE_VOICE = {
    "happy": "zh-CN-XiaoxiaoNeural",
    "excited": "zh-CN-XiaoyiNeural",
    "calm": "zh-CN-YunjianNeural",
    "thinking": "zh-CN-YunjianNeural",
    "surprised": "zh-CN-XiaoyiNeural",
    "sad": "zh-CN-XiaoxiaoNeural",
    "angry": "zh-CN-YunyangNeural",
    "listening": "zh-CN-XiaoxiaoNeural",
    "neutral": "zh-CN-XiaoxiaoNeural",
}

EMOTION_TO_EDGE_RATE = {
    "happy": "+10%",
    "excited": "+20%",
    "calm": "-10%",
    "thinking": "-5%",
    "surprised": "+15%",
    "sad": "-15%",
    "angry": "+5%",
    "listening": "+0%",
    "neutral": "+0%",
}

async def _synthesize_openai(text: str, voice: str, speed: float = 1.0) -> bytes:
    """使用 OpenAI TTS 合成语音"""
    import httpx
    openai_key = state.config.get("llm", {}).get("openai_api_key") if state.config else None
    if not openai_key:
        openai_key = os.environ.get("OPENAI_API_KEY")
    if not openai_key:
        raise HTTPException(status_code=503, detail="OpenAI API Key 未配置")
    async with httpx.AsyncClient(timeout=30.0) as client:
        resp = await client.post(
            "https://api.openai.com/v1/audio/speech",
            headers={"Authorization": f"Bearer {openai_key}", "Content-Type": "application/json"},
            json={
                "model": "tts-1",
                "input": text[:4000],  # OpenAI TTS 限制
                "voice": voice,
                "speed": speed,
                "response_format": "mp3",
            },
        )
        if resp.status_code != 200:
            raise HTTPException(status_code=502, detail=f"OpenAI TTS 失败: {resp.status_code} {resp.text[:200]}")
        return resp.content

async def _synthesize_edge(text: str, voice: str, rate: str = "+0%") -> bytes:
    """使用 Edge TTS 合成语音（免费）"""
    import edge_tts
    import io
    communicate = edge_tts.Communicate(text[:4000], voice, rate=rate)
    audio_buffer = io.BytesIO()
    async for chunk in communicate.stream():
        if chunk["type"] == "audio":
            audio_buffer.write(chunk["data"])
    return audio_buffer.getvalue()

@app.post("/ui/api/tts/synthesize")
async def tts_synthesize(req: dict):
    """情感语音合成 API
    
    Request:
        {"text": "要朗读的文本", "emotion": "happy"}
    
    Response:
        {"audio_base64": "base64音频数据", "source": "openai|edge", "format": "mp3"}
    """
    text = req.get("text", "").strip()
    emotion = req.get("emotion", "neutral")
    if not text:
        raise HTTPException(status_code=400, detail="text 不能为空")
    
    # 清理 markdown
    clean_text = text.replace("**", "").replace("*", "").replace("#", "").replace("`", "")
    clean_text = __import__("re").sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", clean_text)
    clean_text = __import__("re").sub(r"!\[([^\]]*)\]\([^)]+\)", r"\1", clean_text)
    
    if not clean_text.strip():
        raise HTTPException(status_code=400, detail="清理后文本为空")
    
    # FIX: 未配置 OpenAI API Key，直接使用 Edge TTS（免费，无需外部 key）
    # 如需启用 OpenAI TTS，请在 tent_os.yaml 中配置有效的 openai_api_key
    openai_key = state.config.get("llm", {}).get("openai_api_key") if state.config else None
    if not openai_key:
        openai_key = os.environ.get("OPENAI_API_KEY")
    # 检测占位符/无效 key
    if openai_key and (openai_key.startswith("${") or openai_key.strip() == ""):
        openai_key = None
    
    if openai_key:
        try:
            voice = EMOTION_TO_OPENAI_VOICE.get(emotion, "alloy")
            speed = 1.15 if emotion in ("happy", "excited") else 0.9 if emotion in ("sad", "calm") else 1.0
            audio_data = await _synthesize_openai(clean_text, voice, speed)
            return {
                "audio_base64": base64.b64encode(audio_data).decode(),
                "source": "openai",
                "format": "mp3",
                "voice": voice,
                "emotion": emotion,
            }
        except Exception as e:
            logger.warning(f"OpenAI TTS 失败，回退 Edge TTS: {e}")
    
    # Edge TTS（免费，零 API 成本）
    try:
        voice = EMOTION_TO_EDGE_VOICE.get(emotion, "zh-CN-XiaoxiaoNeural")
        rate = EMOTION_TO_EDGE_RATE.get(emotion, "+0%")
        audio_data = await _synthesize_edge(clean_text, voice, rate)
        return {
            "audio_base64": base64.b64encode(audio_data).decode(),
            "source": "edge",
            "format": "mp3",
            "voice": voice,
            "emotion": emotion,
        }
    except Exception as e:
        logger.error(f"Edge TTS 失败: {e}")
        raise HTTPException(status_code=503, detail=f"语音合成服务暂不可用: {e}")


# ========== Soul API —— 灵魂对讲机核心接口 ==========

class SoulProfileResponse(BaseModel):
    user_id: str
    decision_style: float = 0.5
    language_style: float = 0.5
    core_values: list = []
    catchphrases: list = []
    updated_at: Optional[str] = None

class SoulCompletenessResponse(BaseModel):
    thought: float
    voice: float
    appearance: float
    overall: float

class WillRequest(BaseModel):
    heirs: list = []
    topic_whitelist: list = []
    topic_blacklist: list = []
    activation_condition: str = "after_death"
    activation_date: Optional[str] = None


@app.get("/api/v1/soul/profile/{user_id}", response_model=SoulProfileResponse)
async def get_soul_profile(user_id: str):
    """获取用户思维画像"""
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    profile = soul["thought_extractor"].get_profile(user_id)
    if not profile:
        return SoulProfileResponse(user_id=user_id)
    return SoulProfileResponse(
        user_id=profile["user_id"],
        decision_style=profile.get("decision_style", 0.5),
        language_style=profile.get("language_style", 0.5),
        core_values=profile.get("core_values", []),
        catchphrases=profile.get("catchphrases", []),
        updated_at=profile.get("updated_at"),
    )

@app.post("/api/v1/soul/profile/{user_id}")
async def update_soul_profile(user_id: str, updates: dict):
    """用户手动修正思维画像"""
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    success = soul["thought_extractor"].update_profile_manual(user_id, updates)
    return {"status": "ok" if success else "no_change", "user_id": user_id}

@app.get("/api/v1/soul/completeness/{user_id}", response_model=SoulCompletenessResponse)
async def get_soul_completeness(user_id: str):
    """获取灵魂完成度"""
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    comp = soul["thought_extractor"].get_soul_completeness(user_id)
    return SoulCompletenessResponse(**comp)

@app.get("/api/v1/soul/voice/{user_id}")
async def get_voice_profile(user_id: str):
    """获取声纹档案"""
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    profile = soul["voice_modeler"].get_profile(user_id)
    if not profile:
        return {"status": "not_found", "user_id": user_id}
    return {"status": "ok", "profile": profile}

@app.post("/api/v1/soul/voice/{user_id}/sample")
async def upload_voice_sample(user_id: str, file: UploadFile = File(...)):
    """上传语音样本"""
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    import tempfile
    suffix = Path(file.filename).suffix or ".wav"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    result = await soul["voice_modeler"].ingest_sample(user_id, tmp_path, duration_seconds=0)
    import os
    os.unlink(tmp_path)
    return result

@app.get("/api/v1/soul/appearance/{user_id}")
async def get_appearance_profile(user_id: str):
    """获取形象档案"""
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    profile = soul["appearance_modeler"].get_profile(user_id)
    if not profile:
        return {"status": "not_found", "user_id": user_id}
    return {"status": "ok", "profile": profile}

@app.post("/api/v1/soul/appearance/{user_id}/photo")
async def upload_appearance_photo(user_id: str, file: UploadFile = File(...)):
    """上传形象照片"""
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    import tempfile
    suffix = Path(file.filename).suffix or ".jpg"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    result = await soul["appearance_modeler"].ingest_photo(user_id, tmp_path)
    import os
    os.unlink(tmp_path)
    return result

@app.get("/api/v1/soul/will/{user_id}")
async def get_will(user_id: str):
    """获取遗嘱/授权设置"""
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    will = soul["authorization"].get_will(user_id)
    if not will:
        return {"status": "not_found", "user_id": user_id}
    return {"status": "ok", "will": will}

@app.post("/api/v1/soul/will/{user_id}")
async def set_will(user_id: str, request: WillRequest):
    """设置遗嘱/授权"""
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    result = soul["authorization"].set_will(user_id, request.dict())
    return result

@app.post("/api/v1/soul/will/{user_id}/activate")
async def activate_will(user_id: str):
    """激活遗嘱（管理员/系统触发）"""
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    return soul["authorization"].activate_will(user_id)


# ========== 启动入口 ==========

# ========== 消息渠道 Webhook ==========

@app.get("/api/v1/channels")
async def list_channels():
    """列出已启用的消息渠道"""
    if not getattr(state, 'channel_manager', None):
        return {"channels": [], "note": "消息渠道模块已简化"}
    channels = []
    for name, ch in state.channel_manager.channels.items():
        channels.append({
            "name": name,
            "enabled": ch.enabled,
            "description": ch.describe(),
        })
    return {"channels": channels}


@app.post("/api/v1/channels/{channel_name}/webhook")
async def channel_webhook(channel_name: str, payload: Dict[str, Any]):
    """消息渠道 Webhook 入口

    接收外部消息平台的推送，转发到 Tent OS AI 处理，返回 AI 回复。
    适用于 n8n、飞书、Slack 等任何支持 webhook 的平台。
    """
    from fastapi import HTTPException

    # 特殊处理：飞书 challenge 验证
    if channel_name == "feishu" and "challenge" in payload:
        return {"challenge": payload["challenge"]}
    # 特殊处理：Slack URL verification
    if channel_name == "slack" and payload.get("type") == "url_verification":
        return {"challenge": payload.get("challenge")}

    if not getattr(state, 'channel_manager', None):
        raise HTTPException(status_code=503, detail="消息渠道模块已简化")
    channel = state.channel_manager.get(channel_name)
    if not channel:
        raise HTTPException(status_code=404, detail=f"渠道 {channel_name} 未启用")

    # 解析消息
    msg = await channel.parse_incoming(payload)
    if not msg:
        return {"status": "ignored", "reason": "无法解析消息或不需要回复"}

    # 发送到 AI 处理
    session_id = f"{channel_name}_{msg.user_id}_{uuid.uuid4().hex[:8]}"
    future = asyncio.Future()
    state.pending_results[session_id] = future

    await state.bus.publish("governance.request", json.dumps({
        "session_id": session_id,
        "user_id": msg.user_id,
        "content": msg.text,
    }).encode())

    try:
        result = await asyncio.wait_for(future, timeout=120)
        reply_text = result if isinstance(result, str) else str(result)

        # 通过渠道发送回复
        reply = ChannelReply(text=reply_text)
        await channel.send_reply(msg, reply)

        return {"status": "ok", "session_id": session_id, "reply": reply_text[:200]}
    except asyncio.TimeoutError:
        return {"status": "timeout", "session_id": session_id, "message": "AI 处理超时"}
    finally:
        state.pending_results.pop(session_id, None)


# ========== 日志实时流（SSE）==========

from fastapi.responses import StreamingResponse

LOG_FILE_PATH = os.environ.get("TENT_LOG_FILE", "/tmp/tent_os.log")

async def _log_stream_generator(lines: int = 100):
    """生成日志 SSE 流"""
    import subprocess
    
    # 先发送最近的 N 行历史日志
    if Path(LOG_FILE_PATH).exists():
        try:
            proc = await asyncio.create_subprocess_exec(
                "tail", "-n", str(lines), LOG_FILE_PATH,
                stdout=subprocess.PIPE, stderr=subprocess.PIPE
            )
            stdout, _ = await proc.communicate()
            for line in stdout.decode("utf-8", errors="replace").strip().split("\n"):
                if line:
                    yield f"data: {json.dumps({'type': 'history', 'line': line})}\n\n"
        except Exception:
            pass
    
    # 然后实时跟踪新日志
    try:
        proc = await asyncio.create_subprocess_exec(
            "tail", "-f", "-n", "0", LOG_FILE_PATH,
            stdout=subprocess.PIPE, stderr=subprocess.PIPE
        )
        while True:
            line = await proc.stdout.readline()
            if not line:
                await asyncio.sleep(0.5)
                continue
            text = line.decode("utf-8", errors="replace").rstrip("\n")
            yield f"data: {json.dumps({'type': 'live', 'line': text})}\n\n"
    except asyncio.CancelledError:
        proc.kill()
        raise
    except Exception as e:
        yield f"data: {json.dumps({'type': 'error', 'line': str(e)})}\n\n"


@app.get("/api/v1/logs/stream")
async def log_stream(lines: int = 100):
    """日志实时流（Server-Sent Events）
    
    前端连接后，先收到最近 N 行历史日志，然后实时接收新日志。
    """
    return StreamingResponse(
        _log_stream_generator(lines=lines),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


# ========== 启动入口 ==========

async def run_api_server(config_path: str = "./config/tent_os.yaml", host: str = "0.0.0.0", port: int = 8000):
    state.config_path = config_path

    config_obj = uvicorn.Config(app, host=host, port=port, log_level="info")
    server = uvicorn.Server(config_obj)

    logger.info(f"Tent OS 灵魂对讲机 API 启动于 http://{host}:{port}")
    logger.info("  - POST /api/v1/tasks       提交任务")
    logger.info("  - GET  /api/v1/tasks/{id}  查询状态")
    logger.info("  - GET  /api/v1/health      健康检查")
    logger.info("  - GET  /api/v1/soul/*      灵魂层接口（画像/声纹/形象/遗嘱）")
    logger.info("  - GET  /ui/                灵魂对讲机 UI")
    logger.info("  - WS   /ws                 WebSocket 实时对话")

    try:
        await server.serve()
    except asyncio.CancelledError:
        pass


# ========== 静态文件服务（Control UI）—— 放在最后，避免拦截 /ui/api/* ==========

@app.get("/{path:path}", include_in_schema=False)
async def serve_ui(path: str):
    """服务 React SPA 静态文件"""
    if not _UI_DIST_DIR.exists():
        return HTMLResponse(
            "<html><body style='font-family:sans-serif;padding:40px;text-align:center'>"
            "<h1>Tent OS 灵魂对讲机</h1>"
            "<p>前端文件未找到。请检查 frontend/desktop/dist/ 目录。</p>"
            "</body></html>"
        )

    file_path = _UI_DIST_DIR / path
    if path == "" or not file_path.exists() or file_path.is_dir():
        file_path = _UI_DIST_DIR / "index.html"

    _NO_CACHE = {"Cache-Control": "no-store, no-cache, must-revalidate, proxy-revalidate", "Pragma": "no-cache", "Expires": "0"}
    if file_path.exists():
        return FileResponse(file_path, headers=_NO_CACHE)
    return FileResponse(_UI_DIST_DIR / "index.html", headers=_NO_CACHE)
