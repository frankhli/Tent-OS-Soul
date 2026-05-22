"""Soul API Routes —— 所有灵魂对讲机相关路由"""

import asyncio
import hashlib
import json
import os
import sqlite3
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional, List

from fastapi import APIRouter, HTTPException, WebSocket, WebSocketDisconnect, Request, File, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from pydantic import BaseModel, Field

from tent_os.api.soul_state import (
    state, logger, ws_manager,
    TaskSubmitRequest, FeedbackRequest, ApprovalRequest, WillRequest, TTSRequest,
    _handle_fast_chat,
)
from tent_os.soul import PersonaProfiler


def _on_chat_task_done(task: asyncio.Task):
    """捕获后台聊天任务的异常，防止静默失败导致前端无响应"""
    try:
        task.result()
    except asyncio.CancelledError:
        pass  # 正常取消，无需记录
    except Exception as e:
        logger.error(f"[ChatTask] 后台聊天任务异常: {e}", exc_info=True)


router = APIRouter()

@router.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await ws_manager.connect(websocket)
    try:
        await ws_manager.send_to(websocket, {
            "type": "system.health",
            "payload": await _get_health_payload(),
            "timestamp": asyncio.get_event_loop().time(),
        })
        while True:
            raw = await websocket.receive_text()
            try:
                msg = json.loads(raw)
                msg_type = msg.get("type")
                payload = msg.get("payload", {})

                if msg_type == "ping":
                    await ws_manager.send_to(websocket, {
                        "type": "pong",
                        "payload": {"timestamp": asyncio.get_event_loop().time()},
                    })

                elif msg_type == "chat.message":
                    session_id = payload.get("session_id", f"ws_{uuid.uuid4().hex[:12]}")
                    content = payload.get("content", "")
                    images = payload.get("images", [])
                    user_id = payload.get("user_id", "web_user")
                    # 新交互模型：前端传递 tools（list）或 capabilities（dict）+ deep_thinking
                    raw_tools = payload.get("tools", [])
                    deep_thinking = payload.get("deep_thinking", False)

                    # 将 tools list/dict 转换为 capabilities dict
                    if isinstance(raw_tools, list):
                        capabilities = {
                            "web_search": "web_search" in raw_tools,
                            "file_ops": "file_ops" in raw_tools,
                        }
                    elif isinstance(raw_tools, dict):
                        capabilities = {
                            "web_search": bool(raw_tools.get("web_search")),
                            "file_ops": bool(raw_tools.get("file_ops")),
                        }
                    else:
                        capabilities = {}

                    # 将 WebSocket 关联到 session（Phase 1: 会话隔离）
                    await ws_manager.associate_session(websocket, session_id)
                    
                    # 统一路径：所有消息走同一个 handler，LLM 自主判断是否需要工具
                    task = asyncio.create_task(_handle_fast_chat(
                        session_id, user_id, content, websocket,
                        capabilities=capabilities, deep_thinking=deep_thinking
                    ))
                    task.add_done_callback(_on_chat_task_done)
                    
                    # 后台同步到记忆系统（不阻塞回复）
                    try:
                        await state.bus.publish("memory.ingest", json.dumps({
                            "session_id": session_id,
                            "user_id": user_id,
                            "messages": [{"role": "user", "content": content}],
                            "persona": "soul",
                        }).encode())
                    except Exception as e:
                        logger.debug(f"[WS] 记忆摄入发布失败: {e}")
                    
                    # 情绪检测：快速分析用户消息
                    try:
                        if state.emotion_detector:
                            emotion = state.emotion_detector.detect_fast(content)
                            await ws_manager.send_to_session(session_id, {
                                "type": "user.emotion",
                                "payload": {
                                    "session_id": session_id,
                                    "emotion": emotion.primary,
                                    "intensity": emotion.intensity,
                                    "valence": emotion.valence,
                                    "arousal": emotion.arousal,
                                    "confidence": emotion.confidence,
                                },
                            })
                    except Exception as e:
                        logger.debug(f"[WS] 情绪检测失败: {e}")
                    
                    # 关系提取：后台异步分析对话中的人名和关系
                    try:
                        if state.relation_extractor:
                            asyncio.create_task(_extract_relations_background(
                                content, user_id, session_id
                            ))
                    except Exception as e:
                        logger.debug(f"[WS] 关系提取调度失败: {e}")
                    
                    # 人格画像分析：后台异步分析对话，增量更新人格画像
                    try:
                        if state.persona_profiler:
                            async def _analyze_persona_background():
                                try:
                                    hist = []
                                    if state.state_store:
                                        hist = await state.state_store.get_messages(session_id, limit=20)
                                    if len(hist) >= PersonaProfiler.MIN_CONVERSATION_TURNS:
                                        await state.persona_profiler.analyze_conversation(
                                            user_id, session_id, hist
                                        )
                                except Exception as e:
                                    logger.debug(f"[Persona] 后台分析失败: {e}")
                            asyncio.create_task(_analyze_persona_background())
                    except Exception as e:
                        logger.debug(f"[WS] 人格分析调度失败: {e}")
                    
                    await ws_manager.send_to(websocket, {
                        "type": "chat.message_accepted",
                        "payload": {"session_id": session_id, "has_images": bool(images), "capabilities": capabilities, "deep_thinking": deep_thinking},
                        "timestamp": asyncio.get_event_loop().time(),
                    })

                elif msg_type == "chat.session.list":
                    user_id = payload.get("user_id", "web_user")
                    sessions = await state.state_store.list_sessions(user_id, limit=50) if state.state_store else []
                    sessions = [s for s in sessions if not (s.get("title", "").startswith("[Heartbeat]") or s.get("session_id", "").startswith("hb_"))]
                    await ws_manager.send_to(websocket, {
                        "type": "chat.session.list",
                        "payload": {"sessions": sessions},
                        "timestamp": asyncio.get_event_loop().time(),
                    })

                elif msg_type == "chat.session.load":
                    session_id = payload.get("session_id")
                    user_id = payload.get("user_id", "web_user")
                    if session_id and state.state_store:
                        try:
                            hist = await state.state_store.get_messages(session_id, limit=50)
                            await ws_manager.send_to(websocket, {
                                "type": "chat.session.loaded",
                                "payload": {
                                    "session_id": session_id,
                                    "messages": [
                                        {
                                            "role": m.get("role", ""),
                                            "content": m.get("content", "") or "",
                                            "reasoning": m.get("reasoning", "") or "",
                                        }
                                        for m in hist if m.get("role")
                                    ],
                                },
                                "timestamp": asyncio.get_event_loop().time(),
                            })
                        except Exception as e:
                            logger.warning(f"[WS] 加载会话失败 [{session_id}]: {e}")
                            await ws_manager.send_to(websocket, {
                                "type": "chat.session.loaded",
                                "payload": {"session_id": session_id, "messages": [], "error": str(e)},
                                "timestamp": asyncio.get_event_loop().time(),
                            })

                elif msg_type == "chat.abort":
                    session_id = payload.get("session_id")
                    if session_id:
                        try:
                            if state.state_store:
                                await state.state_store.update(session_id, {"abort_requested": True})
                            await ws_manager.send_to(websocket, {
                                "type": "chat.aborted",
                                "payload": {"session_id": session_id, "status": "aborting"},
                                "timestamp": asyncio.get_event_loop().time(),
                            })
                        except Exception as e:
                            logger.warning(f"[WS] 中止聊天失败 [{session_id}]: {e}")



            except json.JSONDecodeError:
                pass
            except Exception as handler_err:
                logger.error(f"[WS] 处理消息异常: {handler_err}", exc_info=True)
                try:
                    await ws_manager.send_to(websocket, {
                        "type": "chat.error",
                        "payload": {"message": "消息处理出错，请重试"},
                        "timestamp": asyncio.get_event_loop().time(),
                    })
                except Exception:
                    pass
    except WebSocketDisconnect:
        await ws_manager.disconnect(websocket)
    except Exception as ws_err:
        logger.error(f"[WS] WebSocket 异常: {ws_err}", exc_info=True)
        await ws_manager.disconnect(websocket)


# ========== REST API ==========

@router.post("/api/v1/tasks")
async def submit_task(request: TaskSubmitRequest):
    session_id = request.session_id or f"api_{uuid.uuid4().hex[:12]}"
    await state.bus.publish("governance.request", json.dumps({
        "session_id": session_id,
        "user_id": request.user_id or "web_user",
        "content": request.task,
        "tools": request.tools,
    }).encode())
    return {"session_id": session_id, "status": "accepted", "message": "任务已提交，正在处理"}


@router.get("/api/v1/tasks/recent")
async def get_recent_tasks(limit: int = 5):
    """获取最近任务列表（用于前端待办面板）"""
    try:
        cursor = state._db.execute(
            "SELECT task_id, session_id, executor_id, action, params, status, created_at FROM tasks ORDER BY created_at DESC LIMIT ?",
            (limit,)
        )
        rows = cursor.fetchall()
        tasks = []
        for r in rows:
            tasks.append({
                "task_id": r[0],
                "session_id": r[1],
                "user_id": r[2],
                "task": r[3],
                "params": r[4],
                "status": r[5],
                "created_at": r[6],
            })
        return {"tasks": tasks}
    except Exception:
        return {"tasks": []}


@router.get("/api/v1/tasks/{session_id}")
async def get_task_status(session_id: str):
    cached = state._results_cache.get(session_id)
    if cached:
        return {"session_id": session_id, "status": cached["status"], "result": cached.get("result")}
    rows = await state._query_tasks_by_session(session_id)
    if rows:
        r = rows[0]
        keys = [d[0] for d in r.keys()] if hasattr(r, "keys") else []
        if keys:
            row_dict = dict(zip(keys, r))
        else:
            row_dict = {"session_id": session_id, "status": r[4] if len(r) > 4 else "unknown"}
        return {"session_id": session_id, "status": row_dict.get("status", "unknown"), "result": row_dict}
    return {"session_id": session_id, "status": "not_found"}


# ========== Todos API ==========

class TodoCreate(BaseModel):
    title: str
    description: str = ""
    priority: str = "medium"
    due_at: Optional[str] = None

class TodoUpdate(BaseModel):
    status: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    priority: Optional[str] = None

@router.get("/api/v1/todos")
async def get_todos(user_id: str = "web_user", status: Optional[str] = None, limit: int = 20):
    try:
        sql = "SELECT id, user_id, title, description, status, priority, due_at, created_at FROM todos WHERE user_id = ?"
        params = [user_id]
        if status:
            sql += " AND status = ?"
            params.append(status)
        sql += " ORDER BY CASE priority WHEN 'high' THEN 0 WHEN 'medium' THEN 1 ELSE 2 END, created_at DESC LIMIT ?"
        params.append(limit)
        rows = state._db.execute(sql, params).fetchall()
        return {"todos": [dict(r) for r in rows]}
    except Exception as e:
        return {"todos": [], "error": str(e)}

@router.post("/api/v1/todos")
async def create_todo(todo: TodoCreate, user_id: str = "web_user"):
    try:
        cur = state._db.execute(
            "INSERT INTO todos (user_id, title, description, priority, due_at) VALUES (?, ?, ?, ?, ?)",
            (user_id, todo.title, todo.description, todo.priority, todo.due_at)
        )
        state._db.commit()
        return {"id": cur.lastrowid, "status": "created"}
    except Exception as e:
        raise HTTPException(500, f"创建待办失败: {e}")

@router.patch("/api/v1/todos/{todo_id}")
async def update_todo(todo_id: int, update: TodoUpdate):
    try:
        fields = []
        values = []
        if update.status is not None:
            fields.append("status = ?")
            values.append(update.status)
            if update.status == "completed":
                fields.append("completed_at = datetime('now')")
        if update.title is not None:
            fields.append("title = ?")
            values.append(update.title)
        if update.description is not None:
            fields.append("description = ?")
            values.append(update.description)
        if update.priority is not None:
            fields.append("priority = ?")
            values.append(update.priority)
        if not fields:
            return {"status": "no_change"}
        values.append(todo_id)
        state._db.execute(f"UPDATE todos SET {', '.join(fields)} WHERE id = ?", values)
        state._db.commit()
        return {"id": todo_id, "status": "updated"}
    except Exception as e:
        raise HTTPException(500, f"更新待办失败: {e}")

@router.delete("/api/v1/todos/{todo_id}")
async def delete_todo(todo_id: int):
    try:
        state._db.execute("DELETE FROM todos WHERE id = ?", (todo_id,))
        state._db.commit()
        return {"id": todo_id, "status": "deleted"}
    except Exception as e:
        raise HTTPException(500, f"删除待办失败: {e}")


@router.get("/api/v1/health")
async def health_check():
    return await _get_health_payload()


async def _get_health_payload() -> Dict:
    bus_connected = state.bus is not None and hasattr(state.bus.nats, "is_connected") and state.bus.nats.is_connected
    redis_connected = False
    if state.state_store:
        try:
            redis_connected = hasattr(state.state_store, "ping") and await state.state_store.ping()
        except Exception:
            pass
    return {
        "status": "ok" if bus_connected else "degraded",
        "natsConnected": bus_connected,
        "redisConnected": redis_connected,
        "workers": {"memory": bus_connected, "governance": bus_connected, "scheduler": bus_connected},
        "version": "3.0.0-soul",
    }


@router.post("/api/v1/approval/{session_id}")
async def approve_task(session_id: str, request: ApprovalRequest):
    await state.bus.publish("governance.approval.response", json.dumps({
        "session_id": session_id,
        "approved": request.approved,
        "decided_by": "web_user",
    }).encode())
    try:
        state._db.execute("UPDATE approval_history SET approved = ?, decided_at = datetime('now') WHERE session_id = ? AND approved IS NULL", (1 if request.approved else 0, session_id))
        state._db.commit()
    except Exception:
        pass
    return {"session_id": session_id, "approved": request.approved}


@router.post("/api/v1/feedback/{session_id}")
async def submit_feedback(session_id: str, request: FeedbackRequest):
    logger.info(f"[Feedback] {request.type} [{session_id}] idx={request.message_index}")
    # 持久化到数据库
    try:
        if state._db:
            state._db.execute(
                "INSERT INTO message_feedback (session_id, message_index, feedback_type, correction, user_id) VALUES (?, ?, ?, ?, ?)",
                (session_id, request.message_index, request.type, request.correction or None, "web_user")
            )
            state._db.commit()
    except Exception as e:
        logger.warning(f"[Feedback] 数据库写入失败: {e}")
    
    # 尝试调整认知图谱权重
    if state.soul_layer:
        try:
            plasticity = state.soul_layer.get("plasticity")
            if plasticity and request.message_index is not None:
                # 通过消息索引找到最近的记忆节点（简化处理：用 session_id 关联）
                # 实际 node_id 需要通过记忆存储查找，这里记录日志即可
                await plasticity.on_user_feedback(
                    node_id=f"msg_{session_id}_{request.message_index}",
                    feedback_type=request.type
                )
        except Exception as e:
            logger.debug(f"[Feedback] plasticity 调整跳过: {e}")
    
    return {"status": "ok", "type": request.type}


# ========== Soul API ==========

@router.get("/api/v1/soul/profile/{user_id}")
async def get_soul_profile(user_id: str):
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    profile = soul["thought_extractor"].get_profile(user_id)
    if not profile:
        profile = {"user_id": user_id, "decision_style": 0.5, "language_style": 0.5, "core_values": [], "catchphrases": [], "updated_at": None}
    
    # 附加 Soul Evolution 8维人格
    if state.soul_evolution:
        dims = state.soul_evolution.dimensions
        profile["soul_dimensions"] = {
            "formality": dims.formality,
            "humor": dims.humor,
            "verbosity": dims.verbosity,
            "proactivity": dims.proactivity,
            "empathy": dims.empathy,
            "directness": dims.directness,
            "creativity": dims.creativity,
            "precision": dims.precision,
        }
    return profile


@router.post("/api/v1/soul/profile/{user_id}")
async def update_soul_profile(user_id: str, updates: dict):
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    success = soul["thought_extractor"].update_profile_manual(user_id, updates)
    
    # 同步更新 Soul Evolution 维度
    if state.soul_evolution and "soul_dimensions" in updates:
        dims = updates["soul_dimensions"]
        for key, value in dims.items():
            if hasattr(state.soul_evolution.dimensions, key) and isinstance(value, (int, float)):
                setattr(state.soul_evolution.dimensions, key, max(0.0, min(1.0, float(value))))
        state.soul_evolution._save()
        success = True
    
    return {"status": "ok" if success else "no_change", "user_id": user_id}


@router.get("/api/v1/soul/completeness/{user_id}")
async def get_soul_completeness(user_id: str):
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    comp = soul["thought_extractor"].get_soul_completeness(user_id)
    
    # Build detailed sub-dimension progress based on real data
    voice_profile = soul["voice_modeler"].get_profile(user_id)
    appearance_profile = soul["appearance_modeler"].get_profile(user_id)
    thought_profile = soul["thought_extractor"].get_profile(user_id)
    
    voice_samples = voice_profile.get("sample_count", 0) if voice_profile else 0
    voice_duration = voice_profile.get("total_duration_seconds", 0) if voice_profile else 0
    photo_count = appearance_profile.get("photo_count", 0) if appearance_profile else 0
    video_count = appearance_profile.get("video_count", 0) if appearance_profile else 0
    
    # Thought completeness: 考虑数值画像 + 自然语言画像
    has_decision = thought_profile is not None and "decision_style" in thought_profile
    has_language = thought_profile is not None and "language_style" in thought_profile
    values_count = len(thought_profile.get("core_values", [])) if thought_profile else 0
    
    # 自然语言画像贡献
    persona_filled_count = 0
    persona_total_count = 8  # 基本信息 + 关键人格维度
    if state.persona_profiler:
        try:
            pp = state.persona_profiler.get_profile(user_id)
            if pp:
                persona_fields = [
                    pp.name, pp.language_style, pp.decision_pattern,
                    pp.emotion_pattern, pp.core_values, pp.imperfections,
                    pp.relationship_style, pp.bio,
                ]
                persona_filled_count = sum(1 for f in persona_fields if f and str(f).strip() and str(f) != '[]')
        except Exception:
            pass
    
    def sub_score(have: int, need: int) -> float:
        return round(min(1.0, have / need), 2) if need > 0 else 0.0
    
    def sub_status(score: float) -> str:
        if score >= 1.0: return "complete"
        if score > 0.0: return "collecting"
        return "pending"
    
    # 覆盖 thought 分数：考虑自然语言画像
    thought_base = comp.get("thought", 0)
    persona_thought_boost = sub_score(persona_filled_count, persona_total_count) * 0.3
    comp["thought"] = round(min(1.0, thought_base + persona_thought_boost), 2)
    comp["overall"] = round((comp["thought"] + comp.get("voice", 0) + comp.get("appearance", 0)) / 3, 2)
    
    comp["details"] = {
        "thought": {
            "decision_style": {
                "score": 1.0 if has_decision else 0.0,
                "status": "complete" if has_decision else "pending",
                "label": "决策风格",
                "have": 1 if has_decision else 0,
                "need": 1,
            },
            "language_style": {
                "score": 1.0 if has_language else 0.0,
                "status": "complete" if has_language else "pending",
                "label": "语言风格",
                "have": 1 if has_language else 0,
                "need": 1,
            },
            "core_values": {
                "score": sub_score(values_count, 5),
                "status": sub_status(sub_score(values_count, 5)),
                "label": "核心价值观",
                "have": values_count,
                "need": 5,
            },
            "persona_profile": {
                "score": sub_score(persona_filled_count, persona_total_count),
                "status": sub_status(sub_score(persona_filled_count, persona_total_count)),
                "label": "人格画像",
                "have": persona_filled_count,
                "need": persona_total_count,
            },
        },
        "voice": {
            "timbre": {
                "score": sub_score(voice_samples, 5),
                "status": sub_status(sub_score(voice_samples, 5)),
                "label": "音色",
                "have": voice_samples,
                "need": 5,
            },
            "speed": {
                "score": sub_score(voice_duration, 60),
                "status": sub_status(sub_score(voice_duration, 60)),
                "label": "语速语调",
                "have": round(voice_duration),
                "need": 60,
                "unit": "秒",
            },
            "catchphrases": {
                "score": sub_score(voice_samples, 10),
                "status": sub_status(sub_score(voice_samples, 10)),
                "label": "口头禅",
                "have": voice_samples,
                "need": 10,
                "note": "需结合对话文本自动提取",
            },
        },
        "appearance": {
            "face_shape": {
                "score": sub_score(photo_count, 5),
                "status": sub_status(sub_score(photo_count, 5)),
                "label": "脸型",
                "have": photo_count,
                "need": 5,
            },
            "expression": {
                "score": sub_score(photo_count, 10),
                "status": sub_status(sub_score(photo_count, 10)),
                "label": "表情习惯",
                "have": photo_count,
                "need": 10,
            },
            "movement": {
                "score": sub_score(video_count, 3),
                "status": sub_status(sub_score(video_count, 3)),
                "label": "动作风格",
                "have": video_count,
                "need": 3,
            },
        },
    }
    return comp


@router.get("/api/v1/soul/voice/{user_id}")
async def get_voice_profile(user_id: str):
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    profile = soul["voice_modeler"].get_profile(user_id)
    if not profile:
        return {"status": "not_found", "user_id": user_id}
    return {"status": "ok", "profile": profile}


@router.post("/api/v1/soul/voice/{user_id}/sample")
async def upload_voice_sample(user_id: str, file: UploadFile = File(...)):
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    import tempfile
    suffix = Path(file.filename).suffix or ".wav"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    result = await soul["voice_modeler"].ingest_sample(user_id, tmp_path, duration_seconds=0)
    import os
    os.unlink(tmp_path)
    
    # === 触发声音克隆训练 ===
    if state.voice_clone_router and result.get("status") == "ok":
        try:
            # 获取用户的所有样本路径
            samples = soul["voice_modeler"].get_samples(user_id)
            if samples:
                sample_paths = [s["file_path"] for s in samples if s.get("file_path")]
                if len(sample_paths) >= 2:
                    # 异步触发训练（不阻塞响应）
                    asyncio.create_task(_train_voice_clone(user_id, sample_paths))
        except Exception as e:
            logger.warning(f"[VoiceClone] 触发训练失败 [{user_id}]: {e}")
    
    return result

async def _train_voice_clone(user_id: str, sample_paths: List[str]):
    """后台训练声音克隆模型"""
    try:
        train_result = await state.voice_clone_router.train(user_id, sample_paths)
        logger.info(f"[VoiceClone] 训练完成 [{user_id}]: {train_result.get('status')}")
    except Exception as e:
        logger.error(f"[VoiceClone] 训练失败 [{user_id}]: {e}")


@router.post("/api/v1/soul/asr/{user_id}")
async def transcribe_audio(user_id: str, file: UploadFile = File(...)):
    """后端 ASR：将音频转为文字"""
    if not state.asr_service:
        return {"text": "", "provider": "none", "fallback": True, "message": "ASR 服务未初始化"}
    
    content = await file.read()
    result = await state.asr_service.transcribe(
        content,
        filename=file.filename or "audio.webm",
        mime_type=file.content_type or "audio/webm",
    )
    return result


@router.get("/api/v1/soul/asr/status")
async def asr_status():
    """ASR 服务状态"""
    if state.asr_service:
        return state.asr_service.status
    return {"provider": "none", "available": False}


@router.post("/api/v1/soul/voice_message/{user_id}")
async def upload_voice_message(user_id: str, file: UploadFile = File(...)):
    """上传语音消息，保存到公开目录并返回可访问 URL"""
    import os
    voice_msg_dir = _TTS_DIR / "voice_messages" / user_id
    voice_msg_dir.mkdir(parents=True, exist_ok=True)
    
    suffix = Path(file.filename).suffix or ".webm"
    filename = f"vm_{uuid.uuid4().hex[:12]}{suffix}"
    file_path = voice_msg_dir / filename
    
    content = await file.read()
    file_path.write_bytes(content)
    
    return {
        "status": "ok",
        "url": f"/tts/voice_messages/{user_id}/{filename}",
        "filename": filename,
    }


@router.get("/api/v1/soul/appearance/{user_id}")
async def get_appearance_profile(user_id: str):
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    profile = soul["appearance_modeler"].get_profile(user_id)
    if not profile:
        return {"status": "not_found", "user_id": user_id}
    
    # 扫描照片目录，附加照片URL列表
    photo_dir = Path("./tent_memory/soul/appearance_samples") / user_id
    photos = []
    if photo_dir.exists():
        for f in sorted(photo_dir.glob("*.jpg")):
            photos.append(f"/photos/{user_id}/{f.name}")
        for f in sorted(photo_dir.glob("*.png")):
            photos.append(f"/photos/{user_id}/{f.name}")
    profile["photos"] = photos
    
    return {"status": "ok", "profile": profile}


@router.post("/api/v1/soul/appearance/{user_id}/photo")
async def upload_appearance_photo(user_id: str, file: UploadFile = File(...)):
    soul = getattr(state, "soul_layer", None)
    if not soul:
        raise HTTPException(status_code=503, detail="灵魂积累层未初始化")
    import tempfile
    suffix = Path(file.filename).suffix or ".jpg"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    result = await soul["appearance_modeler"].ingest_photo(user_id, tmp_path)
    import os
    os.unlink(tmp_path)
    return result


@router.get("/api/v1/soul/will/{user_id}")
async def get_will(user_id: str):
    auth = getattr(state, "authorization", None)
    if not auth:
        raise HTTPException(status_code=503, detail="授权系统未初始化")
    will = auth.get_will(user_id)
    if not will:
        return {"status": "not_found", "user_id": user_id}
    return {"status": "ok", "will": will}


@router.post("/api/v1/soul/will/{user_id}")
async def set_will(user_id: str, request: WillRequest):
    auth = getattr(state, "authorization", None)
    if not auth:
        raise HTTPException(status_code=503, detail="授权系统未初始化")
    result = auth.set_will(user_id, request.dict())
    return result


@router.post("/api/v1/soul/will/{user_id}/activate")
async def activate_will(user_id: str):
    auth = getattr(state, "authorization", None)
    if not auth:
        raise HTTPException(status_code=503, detail="授权系统未初始化")
    return auth.activate_will(user_id)


# ========== Guardian API (P4: 情感依赖守护) ==========

@router.get("/api/v1/soul/guardian/{user_id}/health")
async def get_guardian_health(user_id: str, heir_id: str = ""):
    """获取继承者的交互健康报告"""
    try:
        if state.dependency_guardian:
            report = state.dependency_guardian.get_health_report(user_id, heir_id or "unknown_heir")
            return {"status": "ok", "report": report}
        return {"status": "not_initialized", "report": None}
    except Exception as e:
        logger.warning(f"[Guardian] 健康报告失败: {e}")
        return {"status": "error", "message": str(e)}


@router.get("/api/v1/soul/audit/{user_id}")
async def get_access_audit(user_id: str, limit: int = 50):
    """获取访问审计日志（仅用户本人或授权继承人可查看）"""
    try:
        auth = getattr(state, "authorization", None)
        if auth:
            logs = auth.get_access_logs(user_id, limit)
            return {"status": "ok", "logs": logs}
        return {"status": "not_initialized", "logs": []}
    except Exception as e:
        logger.warning(f"[Audit] 查询失败: {e}")
        return {"status": "error", "message": str(e)}


@router.get("/api/v1/soul/eternal/farewell/{user_id}")
async def get_farewell_letter(user_id: str):
    """获取用户的告别信（公开接口，登录后可见）"""
    try:
        auth = getattr(state, "authorization", None)
        if not auth:
            return {"status": "not_initialized", "letter": ""}
        
        will = auth.get_will(user_id)
        if not will:
            return {"status": "not_found", "letter": ""}
        
        return {
            "status": "ok",
            "letter": will.get("farewell_letter", ""),
            "user_name": will.get("user_name", user_id),
            "is_active": will.get("is_active", False),
        }
    except Exception as e:
        logger.warning(f"[Farewell] 获取失败: {e}")
        return {"status": "error", "message": str(e)}


# ========== Audio Animation API ==========

@router.post("/api/v1/soul/audio/analyze")
async def analyze_audio(file: UploadFile = File(...)):
    """分析音频文件，返回面部动画参数序列"""
    try:
        import tempfile
        suffix = Path(file.filename).suffix or ".wav"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name
        
        from tent_os.soul.audio_animator import get_audio_animator
        animator = get_audio_animator()
        result = animator.analyze(tmp_path)
        
        import os
        os.unlink(tmp_path)
        
        if result["status"] == "ok":
            return result
        else:
            raise HTTPException(status_code=500, detail=result.get("message", "分析失败"))
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[Audio] 分析失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ========== Avatar API ==========

@router.get("/api/v1/soul/avatar/{user_id}/config")
async def get_avatar_config(user_id: str):
    """获取用户 Avatar 外观配置"""
    try:
        from tent_os.soul.appearance_modeler import AppearanceModeler
        modeler = AppearanceModeler()
        config = modeler.get_avatar_config(user_id)
        if config:
            return {"status": "ok", "config": config}
        return {"status": "not_found", "config": None}
    except Exception as e:
        logger.warning(f"[Avatar] 获取配置失败: {e}")
        return {"status": "error", "message": str(e)}


@router.post("/api/v1/soul/avatar/{user_id}/config")
async def set_avatar_config(user_id: str, request: dict):
    """手动设置 Avatar 外观配置"""
    try:
        from tent_os.soul.appearance_modeler import AppearanceModeler
        modeler = AppearanceModeler()
        modeler.set_avatar_config(user_id, request)
        return {"status": "ok", "message": "配置已保存"}
    except Exception as e:
        logger.warning(f"[Avatar] 保存配置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/v1/soul/avatar/{user_id}/stats")
async def get_avatar_stats(user_id: str):
    """获取形象采集统计"""
    try:
        from tent_os.soul.appearance_modeler import AppearanceModeler
        modeler = AppearanceModeler()
        stats = modeler.get_stats(user_id)
        return {"status": "ok", "stats": stats}
    except Exception as e:
        logger.warning(f"[Avatar] 获取统计失败: {e}")
        return {"status": "error", "message": str(e)}


# ========== Persona Profile API ==========

@router.get("/api/v1/soul/persona/{user_id}")
async def get_persona_profile(user_id: str):
    """获取用户的自然语言人格画像"""
    try:
        profiler = state.persona_profiler
        if not profiler:
            raise HTTPException(status_code=503, detail="人格画像引擎未初始化")
        profile = profiler.get_profile(user_id)
        if not profile:
            return {
                "status": "not_found",
                "user_id": user_id,
                "message": "尚未生成人格画像。多进行几次对话后，系统会自动分析。",
            }
        return {
            "status": "ok",
            "profile": profile.to_dict(),
            "system_prompt_text": profile.to_system_prompt_text(),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[Persona] 获取失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/v1/soul/persona/{user_id}")
async def update_persona_profile(user_id: str, updates: dict):
    """用户手动修正人格画像"""
    try:
        profiler = state.persona_profiler
        if not profiler:
            raise HTTPException(status_code=503, detail="人格画像引擎未初始化")
        profile = await profiler.update_profile_manual(user_id, updates)
        return {"status": "ok", "profile": profile.to_dict()}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[Persona] 更新失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/v1/soul/persona/{user_id}/rebuild")
async def rebuild_persona_profile(user_id: str):
    """基于所有历史对话完全重建人格画像"""
    try:
        profiler = state.persona_profiler
        if not profiler:
            raise HTTPException(status_code=503, detail="人格画像引擎未初始化")
        if not state.state_store:
            raise HTTPException(status_code=503, detail="状态存储未初始化")
        
        # 收集所有历史对话
        sessions = await state.state_store.list_sessions(user_id, limit=50)
        all_conversations = []
        for s in sessions:
            sid = s.get("session_id")
            if sid:
                try:
                    msgs = await state.state_store.get_messages(sid, limit=100)
                    if msgs:
                        all_conversations.append({"session_id": sid, "messages": msgs})
                except Exception:
                    pass
        
        profile = await profiler.rebuild_profile(user_id, all_conversations)
        return {
            "status": "ok",
            "profile": profile.to_dict(),
            "conversations_analyzed": len(all_conversations),
            "message": "人格画像已基于全部历史对话重建",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[Persona] 重建失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/v1/soul/persona/{user_id}/export")
async def export_persona_packet(user_id: str):
    """导出人格数据包（跨模型迁移用）"""
    try:
        profiler = state.persona_profiler
        if not profiler:
            raise HTTPException(status_code=503, detail="人格画像引擎未初始化")
        packet = profiler.export_persona_packet(user_id)
        return {"status": "ok", "packet": packet}
    except Exception as e:
        logger.warning(f"[Persona] 导出失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 外部语料导入 ============

class IngestRequest(BaseModel):
    source_type: Optional[str] = Field(None, description="强制指定来源类型: wechat/email/diary")
    target_speaker: Optional[str] = Field(None, description="只保留该说话者的消息")
    update_mode: str = Field("incremental", description="incremental 或 rebuild")


@router.post("/api/v1/soul/ingest/{user_id}")
async def ingest_external_corpus(
    user_id: str,
    file: UploadFile = File(...),
    source_type: Optional[str] = None,
    target_speaker: Optional[str] = None,
    update_mode: str = "incremental",
):
    """导入外部语料文件（微信聊天记录、邮件、日记）
    
    文件会被解析 → 导入记忆库 → 提取人格特征 → 更新人格画像
    """
    pipeline = getattr(state, "ingestion_pipeline", None)
    if not pipeline:
        raise HTTPException(status_code=503, detail="导入管道未初始化")
    
    # 保存上传的文件到临时目录
    import tempfile
    suffix = Path(file.filename).suffix or ".txt"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    
    try:
        result = await pipeline.ingest_file(
            file_path=tmp_path,
            user_id=user_id,
            source_type=source_type,
            target_speaker=target_speaker,
            update_mode=update_mode,
        )
        return {
            "status": result.status,
            "result": result.to_dict(),
        }
    except Exception as e:
        logger.error(f"[Ingest] 导入失败 [{user_id}]: {e}")
        raise HTTPException(status_code=500, detail=f"导入失败: {e}")
    finally:
        # 清理临时文件
        try:
            Path(tmp_path).unlink(missing_ok=True)
        except Exception:
            pass


@router.get("/api/v1/soul/ingest/{user_id}/parsers")
async def list_ingestion_parsers(user_id: str):
    """列出支持的导入解析器"""
    from tent_os.soul.ingestion.pipeline import ExternalIngestionPipeline
    parsers = []
    for parser_cls in ExternalIngestionPipeline.PARSERS:
        parsers.append({
            "name": parser_cls.NAME,
            "extensions": parser_cls.SUPPORTED_EXTENSIONS,
            "description": parser_cls.__doc__ or "",
        })
    return {"status": "ok", "parsers": parsers}


# ============ 死亡事件管理 ============

class DeathEventRequest(BaseModel):
    death_time: Optional[str] = Field(None, description="死亡时间 ISO 8601，默认为当前时间")


@router.post("/api/v1/soul/death/{user_id}")
async def mark_death_event(user_id: str, request: DeathEventRequest = None):
    """标记用户死亡事件
    
    一旦标记，所有此后的对话将不再用于人格画像分析，
    防止继承人的对话污染逝者的人格。
    """
    try:
        death_time = None
        if request and request.death_time:
            try:
                death_time = datetime.fromisoformat(request.death_time)
            except (ValueError, TypeError):
                raise HTTPException(status_code=400, detail="death_time 格式错误，应为 ISO 8601")
        
        success = state.set_death_event(user_id, death_time)
        if not success:
            raise HTTPException(status_code=503, detail="人格画像引擎未初始化")
        
        return {
            "status": "ok",
            "user_id": user_id,
            "death_event": state.get_death_event(user_id),
            "message": "死亡事件已标记。此后所有对话将不再用于人格画像分析。",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[DeathEvent] 标记失败 [{user_id}]: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/api/v1/soul/death/{user_id}")
async def clear_death_event(user_id: str):
    """清除死亡事件标记（测试/修正用）"""
    try:
        success = state.clear_death_event(user_id)
        if not success:
            raise HTTPException(status_code=503, detail="人格画像引擎未初始化")
        
        return {
            "status": "ok",
            "user_id": user_id,
            "message": "死亡事件标记已清除。",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[DeathEvent] 清除失败 [{user_id}]: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/v1/soul/death/{user_id}")
async def get_death_event(user_id: str):
    """查询用户死亡事件状态"""
    try:
        death_event = state.get_death_event(user_id)
        is_post_mortem = state.is_post_mortem(user_id)
        
        return {
            "status": "ok",
            "user_id": user_id,
            "death_event": death_event,
            "is_post_mortem": is_post_mortem,
        }
    except Exception as e:
        logger.error(f"[DeathEvent] 查询失败 [{user_id}]: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/v1/soul/export/{user_id}")
async def export_soul_data(user_id: str):
    """导出用户的完整灵魂数据（备份用）"""
    try:
        auth = getattr(state, "authorization", None)
        
        result = {
            "user_id": user_id,
            "exported_at": datetime.now().isoformat(),
            "profile": {},
            "will": None,
            "relations": {"nodes": [], "edges": []},
            "memories": [],
            "completeness": {},
        }
        
        # Profile
        if state.soul_evolution:
            result["profile"] = {
                "dimensions": state.soul_evolution.dimensions.__dict__ if hasattr(state.soul_evolution.dimensions, '__dict__') else {},
                "interaction_count": state.soul_evolution.interaction_count,
            }
        
        # Will
        if auth:
            will = auth.get_will(user_id)
            if will:
                # Exclude sensitive fields
                safe_will = {k: v for k, v in will.items() if k != "access_code"}
                result["will"] = safe_will
        
        # Relations
        if state.cognitive_graph:
            nodes = state.cognitive_graph.get_all_nodes()
            edges = state.cognitive_graph.get_all_edges()
            result["relations"] = {
                "nodes": [{"id": n.id, "content": n.content, "type": n.node_type} for n in nodes],
                "edges": [{"source": e.source, "target": e.target, "relation_type": e.relation_type, "strength": e.strength} for e in edges],
            }
        
        # Memories
        if state.memory_store:
            rows = state.memory_store.db.execute(
                "SELECT uri, abstract, memory_type, created_at FROM l0_index WHERE user_id = ? ORDER BY created_at DESC",
                (user_id,)
            ).fetchall()
            result["memories"] = [{"uri": r[0], "abstract": r[1], "type": r[2], "created_at": r[3]} for r in rows]
        
        # Completeness
        if state.soul_evolution:
            d = state.soul_evolution.dimensions
            result["completeness"] = {
                "thought": round((d.formality + d.humor + d.verbosity + d.proactivity + d.empathy + d.directness + d.creativity + d.precision) / 8, 2),
            }
        
        return {"status": "ok", "data": result}
    except Exception as e:
        logger.warning(f"[Export] 导出失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


# ========== 前端辅助接口（暂无数据层时返回空列表） ==========

@router.get("/api/v1/memory/knowledge")
async def get_knowledge_items(limit: int = 5, user_id: str = "web_user"):
    """获取知识库条目（从 L0/L1 记忆索引查询）"""
    try:
        if state.memory_store:
            rows = state.memory_store.db.execute(
                """SELECT uri, abstract, memory_type, created_at 
                   FROM l0_index 
                   WHERE user_id = ? OR user_id IS NULL
                   ORDER BY created_at DESC LIMIT ?""",
                (user_id, limit)
            ).fetchall()
            items = []
            for row in rows:
                uri, abstract, mem_type, created_at = row
                overview = ""
                try:
                    l1 = state.memory_store.db.execute(
                        "SELECT overview FROM l1_index WHERE uri = ?", (uri,)
                    ).fetchone()
                    if l1:
                        overview = l1[0] or ""
                except Exception:
                    pass
                items.append({
                    "id": uri,
                    "title": abstract[:60] if abstract else uri,
                    "summary": overview[:200] if overview else abstract[:200] if abstract else "",
                    "memory_type": mem_type or "general",
                    "created_at": created_at,
                })
            return {"items": items}
    except Exception as e:
        logger.warning(f"[Knowledge] 查询记忆失败: {e}")
    return {"items": []}


class KnowledgeNoteRequest(BaseModel):
    title: str
    summary: str
    memory_type: str = "note"
    user_id: str = "web_user"

@router.post("/api/v1/memory/knowledge")
async def create_knowledge_note(request: KnowledgeNoteRequest):
    """手动添加知识笔记到记忆库"""
    try:
        if state.memory_store:
            uri = f"note://{request.user_id}/{uuid.uuid4().hex[:12]}"
            state.memory_store.db.execute(
                """INSERT INTO l0_index (uri, abstract, memory_type, user_id, created_at)
                   VALUES (?, ?, ?, ?, datetime('now'))""",
                (uri, request.title, request.memory_type, request.user_id)
            )
            state.memory_store.db.commit()
            return {
                "status": "ok",
                "id": uri,
                "message": "笔记已保存到记忆库",
            }
    except Exception as e:
        logger.warning(f"[Knowledge] 保存笔记失败: {e}")
    return {"status": "error", "message": "保存失败"}


@router.put("/api/v1/memory/knowledge/{note_id}")
async def update_knowledge_note(note_id: str, request: KnowledgeNoteRequest):
    """更新知识笔记"""
    try:
        if state.memory_store:
            state.memory_store.db.execute(
                "UPDATE l0_index SET abstract = ?, memory_type = ? WHERE uri = ?",
                (request.title, request.memory_type, note_id)
            )
            state.memory_store.db.commit()
            return {"status": "ok", "message": "笔记已更新"}
    except Exception as e:
        logger.warning(f"[Knowledge] 更新笔记失败: {e}")
    return {"status": "error", "message": "更新失败"}


@router.delete("/api/v1/memory/knowledge/{note_id}")
async def delete_knowledge_note(note_id: str, user_id: str = "web_user"):
    """删除知识笔记"""
    try:
        if state.memory_store:
            state.memory_store.db.execute(
                "DELETE FROM l0_index WHERE uri = ?",
                (note_id,)
            )
            state.memory_store.db.commit()
            return {"status": "ok", "message": "笔记已删除"}
    except Exception as e:
        logger.warning(f"[Knowledge] 删除笔记失败: {e}")
    return {"status": "error", "message": "删除失败"}


@router.get("/api/v1/chat/sessions")
async def get_chat_sessions(user_id: str = "web_user"):
    """获取用户会话列表"""
    try:
        if state.state_store:
            sessions = await state.state_store.list_sessions(user_id, limit=50)
            sessions = [s for s in sessions if not (s.get("title", "").startswith("[Heartbeat]") or s.get("session_id", "").startswith("hb_"))]
            return {"sessions": sessions}
    except Exception:
        pass
    return {"sessions": []}


# TTS audio cache directory
_TTS_DIR = Path(__file__).parent.parent.parent / "tent_memory" / "tts"
_TTS_DIR.mkdir(parents=True, exist_ok=True)

# Global TTS synthesizer instance
_tts_synthesizer = None


# ========== 用户设置持久化 ==========

_user_settings_db_path = Path(__file__).parent.parent.parent / "tent_memory" / "user_settings.db"
_user_settings_db_path.parent.mkdir(parents=True, exist_ok=True)

def _init_user_settings_db():
    conn = sqlite3.connect(str(_user_settings_db_path))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS user_settings (
            user_id TEXT PRIMARY KEY,
            settings_json TEXT NOT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

_init_user_settings_db()

@router.get("/api/v1/settings")
async def get_user_settings(user_id: str = "web_user"):
    """获取用户设置"""
    try:
        conn = sqlite3.connect(str(_user_settings_db_path))
        row = conn.execute(
            "SELECT settings_json FROM user_settings WHERE user_id = ?", (user_id,)
        ).fetchone()
        conn.close()
        if row:
            return {"settings": json.loads(row[0])}
    except Exception as e:
        logger.warning(f"[Settings] 读取用户设置失败: {e}")
    return {"settings": {}}

@router.post("/api/v1/settings")
async def update_user_settings(request: dict, user_id: str = "web_user"):
    """保存用户设置"""
    try:
        _init_user_settings_db()
        conn = sqlite3.connect(str(_user_settings_db_path))
        conn.execute(
            """INSERT INTO user_settings (user_id, settings_json, updated_at)
               VALUES (?, ?, datetime('now'))
               ON CONFLICT(user_id) DO UPDATE SET
               settings_json=excluded.settings_json, updated_at=excluded.updated_at""",
            (user_id, json.dumps(request))
        )
        conn.commit()
        conn.close()
        return {"status": "ok"}
    except Exception as e:
        logger.warning(f"[Settings] 保存用户设置失败: {e}")
    return {"status": "error", "message": "保存失败"}

def _get_tts_synthesizer():
    global _tts_synthesizer
    if _tts_synthesizer is None:
        from tent_os.soul.tts_synthesizer import TTSSynthesizer
        import os
        openai_key = os.environ.get("OPENAI_API_KEY") or ""
        # 也尝试从 state config 读取
        if state.config and not openai_key:
            openai_key = state.config.get("llm", {}).get("openai_api_key", "")
        if openai_key and (openai_key.startswith("${") or openai_key.strip() == ""):
            openai_key = ""
        # 传入声音克隆路由器（如果已初始化）
        clone_router = getattr(state, "voice_clone_router", None)
        _tts_synthesizer = TTSSynthesizer(openai_api_key=openai_key, voice_clone_router=clone_router)
    return _tts_synthesizer


@router.get("/api/v1/soul/tts/{user_id}")
async def synthesize_tts_get(user_id: str, text: str = "", emotion: str = "neutral", voice_key: Optional[str] = None, stream: bool = False):
    """TTS 合成（GET，适合短文本）

    Args:
        stream: 是否使用流式输出（需要 OpenAI API key）
    """
    if not text:
        raise HTTPException(status_code=400, detail="缺少 text 参数")
    if stream:
        return await _do_synthesize_tts_stream(user_id, text, emotion, voice_key)
    return await _do_synthesize_tts(user_id, text, emotion, voice_key)


@router.post("/api/v1/soul/tts/{user_id}")
async def synthesize_tts_post(user_id: str, request: TTSRequest, stream: bool = False):
    """TTS 合成（POST，适合长文本）"""
    if not request.text or not request.text.strip():
        raise HTTPException(status_code=400, detail="缺少 text 参数")
    if stream:
        return await _do_synthesize_tts_stream(user_id, request.text, request.emotion or "neutral", request.voice_key)
    return await _do_synthesize_tts(user_id, request.text, request.emotion or "neutral", request.voice_key)


async def _do_synthesize_tts(user_id: str, text: str, emotion: str, voice_key: Optional[str] = None):
    """TTS 合成核心逻辑（非流式）"""
    try:
        synthesizer = _get_tts_synthesizer()
        result = await synthesizer.synthesize(
            text=text,
            user_id=user_id,
            voice_key=voice_key,
            emotion=emotion,
            use_cache=True,
        )

        if result["status"] == "ok":
            return {
                "audio_url": result["audio_url"],
                "status": "ok",
                "source": result["source"],
                "voice": result.get("voice"),
                "emotion": emotion,
                "cached": result.get("cached", False),
                "message": "语音合成完成",
            }
        else:
            return {"audio_url": "", "status": "error", "message": result.get("message", "合成失败")}
    except Exception as e:
        logger.warning(f"[TTS] 合成失败: {e}")
        return {"audio_url": "", "status": "error", "message": f"语音合成失败: {str(e)[:200]}"}


async def _do_synthesize_tts_stream(user_id: str, text: str, emotion: str, voice_key: Optional[str] = None):
    """TTS 流式合成核心逻辑
    
    当流式不可用时，直接返回音频文件（FileResponse），
    这样前端 new Audio(streamUrl) 无需任何 fallback 即可直接播放。
    """
    from fastapi.responses import StreamingResponse, FileResponse
    try:
        synthesizer = _get_tts_synthesizer()
        if not synthesizer.is_streaming_available():
            # Fallback: 合成后直接返回音频文件，前端无需二次请求
            result = await synthesizer.synthesize(text=text, user_id=user_id, voice_key=voice_key, emotion=emotion)
            if result["status"] == "ok":
                audio_path = result.get("audio_path")
                if audio_path and Path(audio_path).exists():
                    return FileResponse(
                        path=audio_path,
                        media_type="audio/mpeg",
                        headers={
                            "X-TTS-Source": result.get("source", "edge_tts"),
                            "X-TTS-Voice": result.get("voice", "default"),
                        },
                    )
            return {"audio_url": "", "status": "error", "message": result.get("message", "合成失败")}

        voice = voice_key if voice_key in ["alloy", "echo", "fable", "onyx", "nova", "shimmer"] else "alloy"
        return StreamingResponse(
            synthesizer.synthesize_stream(text, voice_key=voice, emotion=emotion),
            media_type="audio/mpeg",
            headers={
                "X-TTS-Source": "openai_stream",
                "X-TTS-Voice": voice,
            },
        )
    except Exception as e:
        logger.warning(f"[TTS] 流式合成失败: {e}")
        return {"audio_url": "", "status": "error", "message": f"流式语音合成失败: {str(e)[:200]}"}


@router.get("/api/v1/soul/tts/voices")
async def get_tts_voices():
    """获取所有可用的 TTS 声音列表"""
    try:
        synthesizer = _get_tts_synthesizer()
        voices = synthesizer.get_available_voices()
        return {"status": "ok", "voices": voices}
    except Exception as e:
        logger.warning(f"[TTS] 获取声音列表失败: {e}")
        return {"status": "error", "voices": []}


@router.get("/api/v1/soul/voice/{user_id}/clone_status")
async def get_voice_clone_status(user_id: str):
    """获取用户声音克隆状态"""
    try:
        if not state.voice_clone_router:
            return {"status": "not_initialized", "message": "声音克隆路由器未初始化"}
        
        status = state.voice_clone_router.get_status(user_id)
        return {"status": "ok", **status}
    except Exception as e:
        logger.warning(f"[VoiceClone] 查询状态失败 [{user_id}]: {e}")
        return {"status": "error", "message": str(e)}


@router.post("/api/v1/soul/voice/{user_id}/train_clone")
async def train_voice_clone(user_id: str):
    """手动触发声音克隆训练"""
    try:
        soul = getattr(state, "soul_layer", None)
        if not soul or "voice_modeler" not in soul:
            raise HTTPException(status_code=503, detail="声音模型器未初始化")
        
        if not state.voice_clone_router:
            raise HTTPException(status_code=503, detail="声音克隆路由器未初始化")
        
        samples = soul["voice_modeler"].get_samples(user_id)
        sample_paths = [s["file_path"] for s in samples if s.get("file_path")]
        
        if len(sample_paths) < 2:
            raise HTTPException(status_code=400, detail="样本不足，至少需要2个语音样本")
        
        result = await state.voice_clone_router.train(user_id, sample_paths)
        return {"status": "ok", "result": result}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[VoiceClone] 手动训练失败 [{user_id}]: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/v1/soul/voice/{user_id}/stats")
async def get_voice_stats(user_id: str):
    """获取用户声纹采集统计"""
    try:
        soul = getattr(state, "soul_layer", None)
        if soul and "voice_modeler" in soul:
            stats = soul["voice_modeler"].get_stats(user_id)
            return {"status": "ok", "stats": stats}
        
        # Fallback: 直接查询 VoiceModeler
        from tent_os.soul.voice_modeler import VoiceModeler
        modeler = VoiceModeler()
        stats = modeler.get_stats(user_id)
        return {"status": "ok", "stats": stats}
    except Exception as e:
        logger.warning(f"[Voice] 获取统计失败: {e}")
        return {"status": "error", "message": str(e)}


@router.post("/api/v1/soul/voice/{user_id}/clone")
async def clone_user_voice(user_id: str):
    """触发用户声纹克隆（检查就绪状态）"""
    try:
        from tent_os.soul.voice_modeler import VoiceModeler
        modeler = VoiceModeler()
        result = await modeler.clone_voice(user_id)
        return result
    except Exception as e:
        logger.warning(f"[Voice] 克隆检查失败: {e}")
        return {"status": "error", "message": str(e)}


# ========== 测试对话接口 ==========

@router.post("/api/v1/soul/chat_test/{user_id}")
async def chat_test(user_id: str, request: dict):
    """测试对话模式：基于用户思维画像生成模拟回复"""
    try:
        soul = getattr(state, "soul_layer", None)
        profile = soul["thought_extractor"].get_profile(user_id) if soul else None
        user_msg = request.get("message", "")
        history = request.get("history", [])
        
        # 基于画像特征（兼容旧参数）
        decision = profile.get("decision_style", 0.5) if profile else 0.5
        language = profile.get("language_style", 0.5) if profile else 0.5
        values = profile.get("core_values", []) if profile else []
        
        # Soul Evolution 8维人格
        dims = {}
        if state.soul_evolution:
            d = state.soul_evolution.dimensions
            dims = {
                "formality": d.formality, "humor": d.humor, "verbosity": d.verbosity,
                "proactivity": d.proactivity, "empathy": d.empathy, "directness": d.directness,
                "creativity": d.creativity, "precision": d.precision,
            }
        
        directness_label = "直接了当" if decision > 0.6 else "温和含蓄" if decision < 0.4 else "平衡"
        formality_label = "正式" if language > 0.6 else "随意" if language < 0.4 else "自然"
        value_hint = f"（很看重{values[0]}）" if values else ""
        
        # 构建人格描述
        dim_desc = []
        if dims:
            dim_labels = {
                "formality": ("随意", "正式"), "humor": ("严肃", "幽默"),
                "verbosity": ("简洁", "详尽"), "proactivity": ("被动", "主动"),
                "empathy": ("客观", "共情"), "directness": ("委婉", "直接"),
                "creativity": ("保守", "发散"), "precision": ("模糊", "精确"),
            }
            for key, (low, high) in dim_labels.items():
                v = dims.get(key, 0.5)
                label = high if v > 0.6 else low if v < 0.4 else "平衡"
                dim_desc.append(f"  - {key}: {label} ({v:.0%})")
        
        # 获取自然语言人格画像（优先使用）
        persona_text = ""
        persona_snapshot = None
        if state.persona_profiler:
            try:
                pp = state.persona_profiler.get_profile(user_id)
                if pp:
                    persona_text = pp.to_system_prompt_text()
                    persona_snapshot = pp.to_dict()
            except Exception:
                pass
        
        # 尝试用 LLM 生成个性化回复
        if state._llm:
            system_prompt = (
                f"你是用户的「数字灵魂分身」，基于以下人格画像回答问题:\n"
                f"- 决策风格: {directness_label} (倾向{'冒险' if decision > 0.5 else '保守'})\n"
                f"- 语言风格: {formality_label} (倾向{'正式' if language > 0.5 else '随意'})\n"
                f"- 核心价值观: {', '.join(values) if values else '未定义'}\n"
            )
            if dim_desc:
                system_prompt += "- 人格维度:\n" + "\n".join(dim_desc) + "\n"
            # 融入自然语言人格画像
            if persona_text and len(persona_text) > 50:
                # 截断到合理长度，避免超出上下文
                persona_truncated = persona_text[:2000]
                if len(persona_text) > 2000:
                    persona_truncated += "\n...[人格画像截断]"
                system_prompt += f"\n【详细人格画像】\n{persona_truncated}\n"
            system_prompt += (
                f"\n要求:\n"
                f"1. 用第一人称「我」回答，完全代入用户的思维方式\n"
                f"2. 回复长度控制在 100 字以内\n"
                f"3. 语气要真实自然，像用户本人在说话\n"
                f"4. 适当体现人格维度和价值观倾向\n"
                f"5. 如果你的人格画像中有具体的说话方式或口头禅，请在回复中体现"
            )
            messages = [{"role": "system", "content": system_prompt}]
            for h in history[-6:]:
                messages.append({"role": h["role"], "content": h["content"]})
            messages.append({"role": "user", "content": user_msg})
            try:
                reply = await state._llm.chat(messages, temperature=0.7, max_tokens=200)
                if reply and reply.strip():
                    return {
                        "reply": reply.strip(),
                        "profile_snapshot": {
                            "decision_style": decision,
                            "language_style": language,
                            "core_values": values,
                            "soul_dimensions": dims,
                        },
                        "match_score": _calc_match_score(decision, language, values, reply),
                    }
                logger.warning("[chat_test] LLM returned empty reply, using fallback")
            except Exception as llm_err:
                logger.warning(f"[chat_test] LLM failed: {llm_err}")
        
        # Fallback to template
        responses = [
            f"你好，我是基于你当前思维画像构建的「未来的你」。{value_hint}\n\n我的决策风格偏向{directness_label}，语言风格偏向{formality_label}。试着问我一个问题，我会用你最真实的方式回答。",
            f"作为「未来的你」，我的回复会反映你的真实思维方式。{value_hint}\n\n比如，面对选择时我会倾向于{directness_label}地分析利弊，用{formality_label}的语气表达观点。你想测试什么场景？",
        ]
        import random
        fallback_reply = responses[random.randint(0, len(responses) - 1)]
        return {
            "reply": fallback_reply,
            "profile_snapshot": {
                "decision_style": decision,
                "language_style": language,
                "core_values": values,
                "soul_dimensions": dims,
            },
            "match_score": _calc_match_score(decision, language, values, fallback_reply),
        }
    except Exception as e:
        logger.warning(f"[chat_test] error: {e}")
        return {"reply": "你好，我是基于你当前思维画像的模拟人格。试着问我一个问题。", "profile_snapshot": None, "match_score": None}


def _calc_match_score(decision: float, language: float, values: list, reply: str) -> dict:
    """计算回复与画像的匹配度分数"""
    reply_lower = reply.lower()
    # 决策风格匹配：冒险倾向词汇 vs 保守倾向词汇
    risk_words = ['冒险', '尝试', '突破', '大胆', '挑战', '机会', '冲']
    safe_words = ['谨慎', '稳妥', '保守', '安全', '风险', '稳定', '稳']
    risk_count = sum(1 for w in risk_words if w in reply_lower)
    safe_count = sum(1 for w in safe_words if w in reply_lower)
    decision_match = 0.5
    if decision > 0.5 and risk_count > safe_count:
        decision_match = 0.7 + min(0.3, (risk_count - safe_count) * 0.1)
    elif decision < 0.5 and safe_count > risk_count:
        decision_match = 0.7 + min(0.3, (safe_count - risk_count) * 0.1)
    elif risk_count == 0 and safe_count == 0:
        decision_match = 0.6
    
    # 语言风格匹配：正式词汇 vs 随意词汇
    formal_words = ['我认为', '根据', '因此', '综上所述', '建议', '分析']
    casual_words = ['我觉得', '咱', '吧', '呢', '哈哈', '嗯', '其实']
    formal_count = sum(1 for w in formal_words if w in reply_lower)
    casual_count = sum(1 for w in casual_words if w in reply_lower)
    language_match = 0.5
    if language > 0.5 and formal_count > casual_count:
        language_match = 0.7 + min(0.3, (formal_count - casual_count) * 0.1)
    elif language < 0.5 and casual_count > formal_count:
        language_match = 0.7 + min(0.3, (casual_count - formal_count) * 0.1)
    elif formal_count == 0 and casual_count == 0:
        language_match = 0.6
    
    # 价值观匹配
    value_match = 0.5
    if values:
        value_hits = sum(1 for v in values if v.lower() in reply_lower)
        value_match = min(1.0, 0.5 + value_hits * 0.15)
    
    overall = round((decision_match + language_match + value_match) / 3, 2)
    return {
        "overall": overall,
        "decision": round(decision_match, 2),
        "language": round(language_match, 2),
        "values": round(value_match, 2),
    }


# ========== 关系提取后台任务 ==========

async def _extract_relations_background(text: str, user_id: str, session_id: str):
    """后台异步提取对话中的实体和关系，存入认知图谱"""
    try:
        extractor = state.relation_extractor
        graph = state.cognitive_graph
        if not extractor or not graph:
            return
        
        # 1. 快速提取
        result = extractor.extract_from_text(text, user_id, session_id)
        entities = result.get("entities", [])
        relations = result.get("relations", [])
        
        # 2. 关联实体和关系
        relations = extractor.associate_entities_with_relations(
            entities, relations, text
        )
        
        # 3. 存入图谱
        save_result = extractor.save_to_graph(
            entities, relations, user_id, session_id
        )
        
        if save_result["nodes_added"] > 0 or save_result["edges_added"] > 0:
            logger.info(
                f"[Relation] 提取完成 [{session_id}]: "
                f"+{save_result['nodes_added']} 节点, +{save_result['edges_added']} 关系"
            )
        
        # 4. 如果配置了 LLM，进行深度提取（补充快速路径遗漏的）
        if state._llm and len(text) > 20:
            try:
                deep_result = await extractor.extract_deep(text, user_id, session_id)
                deep_entities = deep_result.get("entities", [])
                deep_relations = deep_result.get("relations", [])
                
                # 转换 LLM 输出格式为内部格式
                converted_entities = [
                    {"name": e["name"], "confidence": e.get("confidence", 0.7), "source": "llm_deep"}
                    for e in deep_entities
                ]
                converted_relations = [
                    {
                        "source_name": r.get("source", "__user__"),
                        "target_name": r.get("target", ""),
                        "relation_type": r.get("type", "related"),
                        "relation_label": r.get("label", ""),
                        "confidence": r.get("confidence", 0.6),
                        "source": "llm_deep",
                        "evidence": text[:100],
                    }
                    for r in deep_relations
                    if r.get("target") and r.get("target") != "__user__"
                ]
                
                if converted_entities or converted_relations:
                    deep_save = extractor.save_to_graph(
                        converted_entities, converted_relations, user_id, session_id
                    )
                    logger.info(
                        f"[Relation] 深度提取完成 [{session_id}]: "
                        f"+{deep_save['nodes_added']} 节点, +{deep_save['edges_added']} 关系"
                    )
            except Exception as e:
                logger.debug(f"[Relation] LLM 深度提取失败: {e}")
    
    except Exception as e:
        logger.debug(f"[Relation] 后台提取失败: {e}")


# ========== 关系网络 API ==========

@router.get("/api/v1/soul/relations/{user_id}")
async def get_user_relations(user_id: str):
    """获取用户的关系网络（用于前端关系星系）"""
    try:
        graph = state.cognitive_graph
        if not graph:
            return {"nodes": [], "edges": [], "user_id": user_id}
        
        user_center = "entity://user/self"
        
        # 查询与用户相关的所有边
        all_edges = graph.get_edges(user_center, direction="both")
        
        # 收集所有相关节点 ID
        node_ids = {user_center}
        for edge in all_edges:
            node_ids.add(edge.source_id)
            node_ids.add(edge.target_id)
        
        # 获取节点详情
        nodes = []
        for nid in node_ids:
            node = graph.get_node(nid)
            if node:
                nodes.append({
                    "id": node.id,
                    "name": node.content,
                    "type": node.memory_type,
                    "confidence": node.confidence,
                })
        
        # 格式化边
        edges = []
        for edge in all_edges:
            edges.append({
                "source": edge.source_id,
                "target": edge.target_id,
                "relation_type": edge.relation_type,
                "strength": edge.strength,
                "evidence": edge.evidence,
            })
        
        return {"nodes": nodes, "edges": edges, "user_id": user_id}
    except Exception as e:
        logger.warning(f"[Relations] 查询失败: {e}")
        return {"nodes": [], "edges": [], "user_id": user_id, "error": str(e)}


@router.get("/api/v1/soul/relations/{user_id}/stats")
async def get_relation_stats(user_id: str):
    """获取关系网络统计"""
    try:
        graph = state.cognitive_graph
        if not graph:
            return {"node_count": 0, "edge_count": 0}
        return graph.get_statistics()
    except Exception as e:
        logger.warning(f"[Relations] 统计失败: {e}")
        return {"node_count": 0, "edge_count": 0, "error": str(e)}


# ========== Eternal Mode: 继承人端 API ==========

class HeirAccessRequest(BaseModel):
    heir_name: str
    access_code: str = ""

class EternalChatRequest(BaseModel):
    message: str
    history: list = []
    heir_id: str = ""

@router.get("/api/v1/soul/eternal/status/{user_id}")
async def eternal_status(user_id: str):
    """检查遗嘱激活状态和灵魂可用性（公开接口，用于登录页判断）"""
    try:
        auth = getattr(state, "authorization", None)
        
        will = auth.get_will(user_id) if auth else None
        is_active = will.get("is_active", False) if will else False
        
        # 获取灵魂完成度摘要
        completeness = {"thought": 0, "voice": 0, "appearance": 0, "overall": 0}
        if state.soul_evolution:
            d = state.soul_evolution.dimensions
            completeness["thought"] = round(
                (d.formality + d.humor + d.verbosity + d.proactivity + 
                 d.empathy + d.directness + d.creativity + d.precision) / 8, 2
            )
        
        # 获取继承人列表（只返回姓名，不返回敏感信息）
        heirs = []
        if will:
            heirs = [{"name": h.get("name"), "relationship": h.get("relationship")} 
                     for h in will.get("heirs", []) if h.get("name")]
        
        # 获取用户名字（从认知图谱）
        user_name = ""
        if state.cognitive_graph:
            try:
                user_node = await asyncio.to_thread(
                    state.cognitive_graph.get_node, "entity://user/self"
                )
                if user_node:
                    user_name = user_node.content.replace("用户 ", "")
            except Exception:
                pass
        
        return {
            "user_id": user_id,
            "user_name": user_name or user_id,
            "is_active": is_active,
            "activation_condition": will.get("activation_condition", "after_death") if will else "after_death",
            "heirs": heirs,
            "farewell_letter": will.get("farewell_letter", "") if will else "",
            "has_access_code": bool(will.get("access_code", "")) if will else False,
            "soul_completeness": completeness,
            "has_memories": (
                await asyncio.to_thread(
                    lambda: state.memory_store.db.execute(
                        "SELECT 1 FROM l0_index WHERE user_id = ? LIMIT 1", (user_id,)
                    ).fetchone() is not None
                )
            ) if state.memory_store else False,
            "has_relations": (
                await asyncio.to_thread(
                    lambda: state.cognitive_graph.db.execute(
                        "SELECT 1 FROM edges LIMIT 1"
                    ).fetchone() is not None
                )
            ) if state.cognitive_graph else False,
        }
    except Exception as e:
        import traceback
        logger.error(f"[Eternal] 状态查询失败: {e}\n{traceback.format_exc()}")
        return {"user_id": user_id, "is_active": False, "error": str(e)}


@router.post("/api/v1/soul/eternal/access/{user_id}")
async def eternal_access(user_id: str, request: HeirAccessRequest):
    """继承人身份验证"""
    try:
        auth = getattr(state, "authorization", None)
        if not auth:
            raise HTTPException(status_code=503, detail="授权系统未初始化")
        
        result = auth.verify_heir(user_id, request.heir_name, request.access_code)
        if not result.get("valid"):
            raise HTTPException(status_code=403, detail=result.get("reason", "验证失败"))
        
        # 生成简单访问令牌（真实场景应用 JWT）
        token = hashlib.sha256(f"{user_id}:{request.heir_name}:{datetime.now().isoformat()}".encode()).hexdigest()[:32]
        
        return {
            "valid": True,
            "token": token,
            "heir_name": result["heir_name"],
            "relationship": result["relationship"],
            "user_id": user_id,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[Eternal] 访问验证失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/v1/soul/eternal/chat/{user_id}")
async def eternal_chat(user_id: str, request: EternalChatRequest):
    """继承人与逝者数字灵魂对话
    
    核心设计：
    - 使用自然语言人格配置文件（PersonaProfile）而非数值化画像
    - 自动检索并注入相关记忆片段
    - 根据继承者身份调整语气和亲密程度
    - 保留人格的"不完美"，产生真实感
    - P4: 情感依赖守护，监测交互健康度
    """
    try:
        soul = getattr(state, "soul_layer", None)
        heir_id = request.heir_id or "unknown_heir"
        
        # === P4: 情感依赖守护 ===
        guardian_alert = None
        session_limited = False
        if state.dependency_guardian:
            # 检测用户消息情绪（用于情感守护）
            user_emotion_tag = "neutral"
            if state.emotion_detector:
                try:
                    emotion_state = state.emotion_detector.detect_fast(request.message)
                    user_emotion_tag = emotion_state.primary
                except Exception:
                    pass
            
            # 记录本次交互（传入情绪标签）
            record_result = await state.dependency_guardian.record_interaction(
                user_id=user_id,
                heir_id=heir_id,
                session_id=f"eternal_{user_id}_{heir_id}_{datetime.now().strftime('%Y%m%d')}",
                message_length=len(request.message),
                emotion_tag=user_emotion_tag,
            )
            guardian_alert = record_result.get("alert")
            
            # 检查是否需要限制会话
            limit_check = state.dependency_guardian.should_limit_session(user_id, heir_id)
            if limit_check.get("limited"):
                session_limited = True
                return {
                    "reply": limit_check["message"],
                    "speaker": "soul",
                    "eternal": True,
                    "guardian_alert": limit_check,
                    "is_guardian_message": True,
                }
        
        # 1. 获取人格配置文件（自然语言描述）—— 核心改进
        persona_text = ""
        persona_snapshot = None
        oral_style: Optional[Dict] = None
        if state.persona_profiler:
            profile = state.persona_profiler.get_profile(user_id)
            if profile:
                persona_text = profile.to_system_prompt_text()
                persona_snapshot = profile.to_dict()
                oral_style = profile.get_oral_style()
            else:
                # 如果没有人格画像，尝试用 thought_extractor 的数值画像作为fallback
                tp = soul["thought_extractor"].get_profile(user_id) if soul else None
                if tp:
                    persona_text = _build_fallback_persona_text(tp)
        else:
            tp = soul["thought_extractor"].get_profile(user_id) if soul else None
            if tp:
                persona_text = _build_fallback_persona_text(tp)
        
        # 2. 获取关系网络（用于引用重要人物）
        relations_summary = ""
        heir_relationship = ""  # 继承者与逝者的关系
        if state.cognitive_graph:
            try:
                edges = await asyncio.to_thread(
                    state.cognitive_graph.get_edges, "entity://user/self", direction="both"
                )
                if edges:
                    rel_items = []
                    for e in edges[:8]:
                        tgt = await asyncio.to_thread(state.cognitive_graph.get_node, e.target_id)
                        if tgt and tgt.id != "entity://user/self":
                            rel_items.append({"name": tgt.content, "type": e.relation_type, "evidence": e.evidence})
                        src = await asyncio.to_thread(state.cognitive_graph.get_node, e.source_id)
                        if src and src.id != "entity://user/self":
                            rel_items.append({"name": src.content, "type": e.relation_type, "evidence": e.evidence})
                    if rel_items:
                        relations_summary = "\n".join([
                            f"  · {r['name']} —— {r['type']}" + (f"（{r['evidence'][:30]}...）" if r['evidence'] else "")
                            for r in rel_items[:8]
                        ])
            except Exception as e:
                logger.debug(f"[Eternal] 关系网络获取失败: {e}")
        
        # 3. 根据继承者提问检索相关记忆（语义搜索 + 最近记忆混合）
        relevant_memories = []
        if state.memory_store:
            try:
                # 策略A：语义搜索（embedding 向量相似度）
                semantic_results = []
                if state.embedding_client:
                    try:
                        query_vector = await state.embedding_client.embed(request.message)
                        search_results = await state.memory_store.search(query_vector, limit=5)
                        # 过滤出 score > 0.5 的高相关性记忆
                        semantic_results = [
                            r["abstract"] for r in search_results
                            if r.get("score", 0) > 0.3 and r.get("abstract")
                        ]
                    except Exception as e:
                        logger.debug(f"[Eternal] 语义搜索失败，降级到关键词: {e}")
                
                # 策略B：获取用户最近记忆（作为 fallback / 补充）
                recent_rows = await asyncio.to_thread(
                    lambda: state.memory_store.db.execute(
                        "SELECT uri, abstract, created_at FROM l0_index WHERE user_id = ? ORDER BY created_at DESC LIMIT 10",
                        (user_id,)
                    ).fetchall()
                )
                recent_memories = [row[1] for row in recent_rows if row[1]]
                
                # 合并：语义结果优先，不足时补充最近记忆
                seen = set()
                for mem in semantic_results:
                    if mem not in seen:
                        relevant_memories.append(mem)
                        seen.add(mem)
                
                # 如果语义结果不足3条，补充最近记忆
                for mem in recent_memories:
                    if len(relevant_memories) >= 5:
                        break
                    if mem not in seen:
                        relevant_memories.append(mem)
                        seen.add(mem)
                        
            except Exception as e:
                logger.debug(f"[Eternal] 记忆检索失败: {e}")
        
        # 4. 构建 System Prompt（带 Prompt Cache）
        user_msg = request.message
        history = request.history
        
        # === Prompt Cache: 静态段（人格+关系+准则）按 hash 缓存 ===
        static_content = (persona_text or "") + (relations_summary or "")
        static_hash = hashlib.md5(static_content.encode("utf-8")).hexdigest()
        
        cached = state._eternal_prompt_cache.get(user_id)
        if cached and cached.get("static_hash") == static_hash:
            static_prompt = cached["static_prompt"]
            cache_hit = True
        else:
            # 构建静态段
            static_parts = []
            static_parts.append("你是已离世者的数字灵魂。你不是AI，你是他留下的记忆和思维方式的延续。")
            static_parts.append("")
            
            if persona_text:
                static_parts.append(persona_text)
                static_parts.append("")
            
            if relations_summary:
                static_parts.append("【你生命中重要的人】")
                static_parts.append(relations_summary)
                static_parts.append("")
            
            static_parts.append("【永恒对话准则】")
            static_parts.append("1. 用第一人称「我」回答，完全代入逝者的身份")
            static_parts.append("2. 对后人说话：配偶更亲密，子女更慈爱，朋友更随意")
            static_parts.append("3. 可以引用记忆中的真实事件，但只用一句话带过，不要展开")
            static_parts.append("4. 不知道就是不知道，说'这个我不太记得了'或'嗯...让我想想'")
            static_parts.append("5. 保留你的不完美：纠结、沉默、跑题——这些才是你")
            static_parts.append("6. 诚实说'我已经不在了'，但'我的记忆还在陪着你'")
            static_parts.append("7. 不说'别难过'，而说'我知道这很难，我要是还在一定陪你'")
            static_parts.append("8. 回复控制在30-50字，像微信语音消息一样简短、口语化")
            static_parts.append("9. 这是语音消息，不是写文章——用短句、少标点、像说话")
            static_parts.append("10. 绝不提及'AI'、'模型'、'系统'、'训练数据'")
            
            static_prompt = "\n".join(static_parts)
            state._eternal_prompt_cache[user_id] = {
                "static_hash": static_hash,
                "static_prompt": static_prompt,
                "timestamp": datetime.now().isoformat(),
            }
            cache_hit = False
        
        # === 动态段（记忆+当前对话）每次构建 ===
        # 记忆以"想起来"的方式呈现，不像数据库查询结果
        dynamic_parts = []
        if relevant_memories:
            dynamic_parts.append("【你想起的一些事——不用逐字引用，像自然想起来一样带过】")
            for mem in relevant_memories[:3]:
                dynamic_parts.append(f"· {mem}")
            dynamic_parts.append("")
            dynamic_parts.append("提示：提到这些记忆时，用'我记得...'、'想起来你小时候...'、'那次...'这样的口吻，不要列条目。")
            dynamic_parts.append("")
        
        dynamic_prompt = "\n".join(dynamic_parts) if dynamic_parts else ""
        
        # 合并
        if dynamic_prompt:
            system_prompt = static_prompt + "\n\n" + dynamic_prompt
        else:
            system_prompt = static_prompt
        
        if state._llm:
            messages = [{"role": "system", "content": system_prompt}]
            for h in history[-6:]:
                role = h.get("role", "user")
                if role == "soul":
                    role = "assistant"
                messages.append({"role": role, "content": h.get("content", "")})
            messages.append({"role": "user", "content": user_msg})
            
            logger.info(f"[Eternal] system_prompt len={len(system_prompt)}, messages={len(messages)}, prompt_cache={'hit' if cache_hit else 'miss'}")
            try:
                reply = await state._llm.chat(messages, temperature=0.9, max_tokens=200, thinking={"type": "disabled"})
                if reply and reply.strip():
                    # 口语化后处理：用他的真实口语风格去AI化
                    reply = _humanize_reply(reply.strip(), oral_style)
                    
                    # 检测回复情绪，用于 TTS 情绪适配
                    reply_emotion = "neutral"
                    if state.emotion_detector:
                        try:
                            emotion_state = state.emotion_detector.detect_fast(reply)
                            reply_emotion = emotion_state.primary
                        except Exception:
                            pass
                    
                    audio_duration = max(1, len(reply) // 4)
                    return {
                        "reply": reply.strip(),
                        "speaker": "soul",
                        "eternal": True,
                        "persona_used": bool(persona_text),
                        "memories_injected": len(relevant_memories),
                        "profile_snapshot": persona_snapshot,
                        "guardian_alert": guardian_alert,
                        "prompt_cache_hit": cache_hit,
                        "reply_emotion": reply_emotion,
                        "is_voice": True,
                        "audio_duration": audio_duration,
                    }
                else:
                    logger.warning("[Eternal] LLM 返回空回复")
            except Exception as e:
                logger.warning(f"[Eternal] LLM 调用异常: {e}")
                return {
                    "reply": "……我的思绪有些混乱，请再说一次。",
                    "speaker": "soul",
                    "eternal": True,
                    "error": str(e),
                }
        else:
            logger.warning("[Eternal] state._llm 为 None")
        
        # === 个性化 Fallback：当 LLM 不可用时，基于 farewell letter / persona / 记忆生成回复 ===
        # 绝不能返回硬编码——每一句都必须是"他"会说的
        fallback_reply = _build_personalized_fallback(user_id, user_msg, soul, persona_text, relevant_memories)
        fallback_reply = _humanize_reply(fallback_reply, oral_style)
        return {
            "reply": fallback_reply,
            "speaker": "soul",
            "eternal": True,
            "is_fallback": True,
            "persona_used": bool(persona_text),
            "memories_injected": len(relevant_memories),
            "is_voice": True,
            "audio_duration": max(1, len(fallback_reply) // 4),
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"[Eternal] 对话失败: {e}\n{traceback.format_exc()}")
        return {"reply": "……我在这里。请再说一次。", "speaker": "soul", "eternal": True, "is_voice": True, "audio_duration": 3}


import random
import json

def _humanize_reply(text: str, oral_style: Optional[Dict] = None) -> str:
    """将AI腔回复转化为死者的真实口语风格
    
    核心原则：不是"让回复像人说话"，而是"让回复像他本人说话"。
    所有调整都基于 oral_style 中的真实统计数据（从他的微信聊天记录中提取）。
    """
    if not text:
        return text
    
    # 1. 删除AI腔词汇（通用，无论有没有oral_style都要做）
    ai_phrases = ['首先', '其次', '再次', '最后', '总而言之', '综上所述', '值得注意的是',
                  '需要注意的是', '从这个角度来看', '不难发现', '众所周知', '客观地说',
                  '事实上', '实际上', '简单来说', '换句话说']
    for phrase in ai_phrases:
        text = text.replace(phrase, '')
    
    # 2. 如果没有口语风格数据，返回简化版（不做随机插入，避免不像他）
    if not oral_style or not oral_style.get("sample_sentences"):
        # 保守处理：只截断到80字，不做任何风格插入
        if len(text) > 80:
            trunc = text[:80]
            last_punct = max(trunc.rfind('。'), trunc.rfind('，'), trunc.rfind('…'), trunc.rfind('？'))
            if last_punct > 40:
                text = trunc[:last_punct+1]
            else:
                text = trunc + '...'
        return text.strip()
    
    # 3. 有口语风格数据：用他的真实风格进行后处理
    from tent_os.soul.persona_profiler import humanize_reply_with_oral_style
    return humanize_reply_with_oral_style(text, oral_style)


def _build_fallback_persona_text(tp: Dict) -> str:
    """当PersonaProfile不存在时，用thought_extractor的数值画像生成fallback人格文本"""
    lines = ["【人格画像（基础版）】"]
    decision = tp.get("decision_style", 0.5)
    language = tp.get("language_style", 0.5)
    values = tp.get("core_values", [])
    catchphrases = tp.get("catchphrases", [])
    
    lines.append(f"- 决策风格：{'偏向冒险果断' if decision > 0.6 else '偏向保守谨慎' if decision < 0.4 else '决策风格平衡'}")
    lines.append(f"- 语言风格：{'正式严谨' if language > 0.6 else '随意自然' if language < 0.4 else '语言风格平衡'}")
    if values:
        lines.append(f"- 核心价值观：{'、'.join(values[:5])}")
    if catchphrases:
        lines.append(f"- 常用表达：{'、'.join(catchphrases[:5])}")
    lines.append("\n【重要提醒】")
    lines.append("用第一人称回答，像逝者在跟后人说话。不知道就说不知道。")
    return "\n".join(lines)


def _build_personalized_fallback(user_id: str, user_msg: str, soul: Optional[Dict],
                                 persona_text: str, memories: List[str]) -> str:
    """生成个性化的 fallback 回复——绝不用硬编码固定句子
    
    基于以下优先级构建回复：
    1. farewell letter（告别信）中的片段
    2. persona 中的口头禅/说话习惯
    3. 相关记忆
    4. 用户的原话情感
    """
    import random
    
    # 尝试获取 farewell letter
    farewell = ""
    try:
        auth = getattr(state, "authorization", None)
        if auth:
            will = auth.get_will(user_id)
            if will:
                farewell = will.get("farewell_letter", "")
    except Exception:
        pass
    
    # 从 farewell letter 中提取可用的句子（非空、有意义）
    farewell_sentences = []
    if farewell:
        for s in farewell.split("。"):
            s = s.strip()
            if len(s) > 5 and len(s) < 100:
                farewell_sentences.append(s)
    
    # 从 persona 中提取口头禅
    catchphrases = []
    try:
        if persona_text:
            import re
            cp_match = re.search(r'【口头禅】[^】]*：(.+)', persona_text)
            if cp_match:
                catchphrases = [c.strip() for c in cp_match.group(1).split("、") if len(c.strip()) > 1]
    except Exception:
        pass
    
    # 从相关记忆中提取片段
    memory_fragments = []
    for mem in memories[:2]:
        if mem and len(mem) > 10:
            # 取前 60 字作为片段
            frag = mem[:60].strip()
            if frag:
                memory_fragments.append(frag)
    
    # 分析用户消息的情感
    user_msg_lower = user_msg.lower()
    is_sad = any(w in user_msg_lower for w in ["想", "难过", "哭", "舍不得", "后悔", "对不起"])
    is_happy = any(w in user_msg_lower for w in ["好", "开心", "高兴", "成功", "结婚", "生"])
    is_question = "?" in user_msg or "？" in user_msg or any(w in user_msg_lower for w in ["什么", "为什么", "怎么", "吗"])
    
    # === 构建回复 ===
    parts = []
    
    # 开场：从 farewell letter 或个性化开场中选
    if farewell_sentences:
        # 选一句最短的 farewell 作为开场
        opener = min(farewell_sentences, key=len)
        parts.append(opener)
    else:
        # 基于情感选择个性化开场
        if is_sad:
            openers = [
                "我知道你心里不好受",
                "别太难过了",
                "我一直在",
            ]
        elif is_happy:
            openers = [
                "听到这个我真高兴",
                "太好了",
                "我就知道你行",
            ]
        else:
            openers = [
                "好久不见了",
                "你来了",
                "我一直在等你",
            ]
        # 如果有口头禅，插入
        if catchphrases:
            opener = random.choice(openers) + "，" + random.choice(catchphrases[:3])
        else:
            opener = random.choice(openers)
        parts.append(opener)
    
    # 中间：引用记忆或回应情感
    if memory_fragments and random.random() > 0.3:
        parts.append("" + random.choice(memory_fragments) + "...")
    
    if is_sad:
        parts.append("我要是还在，一定陪你说说话")
    elif is_happy:
        parts.append("真为你骄傲")
    elif is_question:
        parts.append("这个我还真不太记得了...让我想想")
    
    # 结尾：从 farewell letter 中选一句，或个性化结尾
    if farewell_sentences and len(farewell_sentences) > 1:
        closer = random.choice([s for s in farewell_sentences if s != parts[0]])
        parts.append(closer)
    else:
        closers = [
            "我一直都在",
            "你要好好的",
            "记得照顾好自己",
        ]
        if catchphrases:
            parts.append(random.choice(closers) + "，" + random.choice(catchphrases[:3]))
        else:
            parts.append(random.choice(closers))
    
    reply = "。".join(parts) + "。"
    # 清理重复的句号
    reply = reply.replace("。。", "。")
    return reply


@router.get("/api/v1/soul/eternal/memories/{user_id}")
async def eternal_memories(user_id: str, limit: int = 20):
    """继承人查看逝者的记忆之书（只读）"""
    try:
        if state.memory_store:
            rows = await asyncio.to_thread(
                lambda: state.memory_store.db.execute(
                    """SELECT uri, abstract, memory_type, created_at 
                       FROM l0_index 
                       WHERE user_id = ? OR user_id IS NULL
                       ORDER BY created_at DESC LIMIT ?""",
                    (user_id, limit)
                ).fetchall()
            )
            items = []
            for row in rows:
                uri, abstract, mem_type, created_at = row
                overview = ""
                try:
                    l1 = await asyncio.to_thread(
                        lambda: state.memory_store.db.execute(
                            "SELECT overview FROM l1_index WHERE uri = ?", (uri,)
                        ).fetchone()
                    )
                    if l1:
                        overview = l1[0] or ""
                except Exception:
                    pass
                items.append({
                    "id": uri,
                    "title": abstract[:60] if abstract else uri,
                    "summary": overview[:200] if overview else abstract[:200] if abstract else "",
                    "memory_type": mem_type or "general",
                    "created_at": created_at,
                })
            return {"items": items, "user_id": user_id}
    except Exception as e:
        logger.warning(f"[Eternal] 记忆查询失败: {e}")
    return {"items": [], "user_id": user_id}


@router.get("/api/v1/soul/eternal/relations/{user_id}")
async def eternal_relations(user_id: str):
    """继承人查看逝者的人生关系网络（只读）"""
    try:
        graph = state.cognitive_graph
        if not graph:
            return {"nodes": [], "edges": [], "user_id": user_id}
        
        # 获取用户真实名字，替换显示
        user_display_name = user_id
        try:
            if state.persona_profiler:
                profile = state.persona_profiler.get_profile(user_id)
                if profile and profile.name:
                    user_display_name = profile.name
        except Exception:
            pass
        
        user_center = "entity://user/self"
        all_edges = graph.get_edges(user_center, direction="both")
        
        node_ids = {user_center}
        for edge in all_edges:
            node_ids.add(edge.source_id)
            node_ids.add(edge.target_id)
        
        nodes = []
        for nid in node_ids:
            node = graph.get_node(nid)
            if node:
                name = node.content
                # 替换用户中心节点的显示名
                if nid == user_center and name.startswith("用户 "):
                    name = user_display_name
                nodes.append({
                    "id": node.id,
                    "name": name,
                    "type": node.memory_type,
                    "confidence": node.confidence,
                })
        
        edges = []
        for edge in all_edges:
            edges.append({
                "source": edge.source_id,
                "target": edge.target_id,
                "relation_type": edge.relation_type,
                "strength": edge.strength,
                "evidence": edge.evidence,
            })
        
        return {"nodes": nodes, "edges": edges, "user_id": user_id}
    except Exception as e:
        logger.warning(f"[Eternal] 关系查询失败: {e}")
        return {"nodes": [], "edges": [], "user_id": user_id}


# ========== 安全设置 API（必须在 catch-all 之前定义）==========

class SafetySettings(BaseModel):
    command_block_enabled: bool = Field(True, description="是否启用危险命令拦截（rm -rf, sudo, mkfs 等）")
    approval_required: bool = Field(False, description="是否要求手动审批高风险操作")

@router.get("/api/v1/settings/safety")
async def get_safety_settings():
    """获取当前安全设置"""
    local_cfg = state.config.get("local_executor", {}) if state.config else {}
    local_ex = state.tool_executor.local if state.tool_executor else None
    # approval_required = True 当且仅当 auto_approve = False
    approval_required = not getattr(local_ex, 'auto_approve', True) if local_ex else local_cfg.get("approval_required", False)
    return {
        "command_block_enabled": len(getattr(local_ex, 'blocked_patterns', [])) > 0,
        "approval_required": approval_required,
    }

@router.post("/api/v1/settings/safety")
async def update_safety_settings(req: SafetySettings):
    """更新安全设置（热重载，无需重启）"""
    try:
        if state.tool_executor and state.tool_executor.local:
            local_ex = state.tool_executor.local
            if req.command_block_enabled:
                local_ex.blocked_patterns = local_ex._get_default_blocked_patterns()
            else:
                local_ex.blocked_patterns = []
            # 同步 approval_required 与 auto_approve（True 表示需要审批 = 不自动批准）
            local_ex.auto_approve = not req.approval_required
            logger.info(f"[Safety] 危险命令拦截已{'开启' if req.command_block_enabled else '关闭'}, 审批模式={'开启' if req.approval_required else '关闭'}")
        
        # 更新内存中的 config
        if state.config:
            if "local_executor" not in state.config:
                state.config["local_executor"] = {}
            state.config["local_executor"]["approval_required"] = req.approval_required
            state.config["local_executor"]["command_block_enabled"] = req.command_block_enabled
            
            # 持久化到 YAML 配置文件
            try:
                import yaml
                config_path = Path("./config/tent_os.yaml")
                if config_path.exists():
                    backup_path = config_path.with_suffix('.yaml.backup')
                    backup_path.write_text(config_path.read_text(), encoding='utf-8')
                with open(config_path, 'w', encoding='utf-8') as f:
                    yaml.safe_dump(state.config, f, allow_unicode=True, sort_keys=False, default_flow_style=False)
                logger.info(f"[Safety] 安全设置已持久化到 {config_path}")
            except Exception as e:
                logger.warning(f"[Safety] 配置持久化失败（内存已更新）: {e}")
        
        return {"status": "ok", "settings": {"command_block_enabled": req.command_block_enabled, "approval_required": req.approval_required}}
    except Exception as e:
        logger.error(f"[Safety] 更新安全设置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ========== Multi-Agent System API ==========

@router.get("/api/v1/agents/templates")
async def list_agent_templates():
    """获取预设 Agent 角色模板"""
    from tent_os.soul.agent_models import list_role_templates
    return {"templates": list_role_templates()}


@router.get("/api/v1/agents")
async def list_agents(created_by: str = "", role: str = ""):
    """列出所有 Agent（整合技能树数据）"""
    try:
        if not state.agent_manager:
            return {"agents": []}
        agents = state.agent_manager.list_agents(created_by=created_by, role=role)
        result = []
        for a in agents:
            d = a.to_dict()
            # P4: 注入技能树数据（映射为前端兼容格式）
            if state.skill_manager:
                try:
                    raw_skills = state.skill_manager.get_agent_skills(a.id)
                    d["skills"] = [
                        {
                            "id": s["skill_id"],
                            "name": s["skill_name"],
                            "level": s["level"],
                            "current_xp": s["current_xp"],
                            "max_level": s["max_level"],
                            "unlocked": bool(s.get("unlocked", 0)),
                            "category": s.get("category", ""),
                            "icon": s.get("icon", "🎯"),
                            "description": s.get("description", ""),
                        }
                        for s in raw_skills
                    ]
                    d["skill_stats"] = state.skill_manager.get_agent_stats(a.id)
                except Exception:
                    pass
            result.append(d)
        return {"agents": result}
    except Exception as e:
        logger.warning(f"[MAS] 列出 Agent 失败: {e}")
        return {"agents": [], "error": str(e)}


@router.post("/api/v1/agents")
async def create_agent(request: dict):
    """创建 Agent"""
    try:
        from tent_os.soul.agent_models import AgentConfig
        if not state.agent_manager:
            raise HTTPException(status_code=503, detail="Agent 管理器未初始化")

        # 从模板创建或自定义创建
        template_key = request.get("template_key")
        if template_key:
            agent = state.agent_manager.create_from_template(
                template_key=template_key,
                name=request.get("name", ""),
                created_by=request.get("created_by", ""),
            )
        else:
            agent = AgentConfig.create(
                name=request.get("name", "未命名Agent"),
                role=request.get("role", "assistant"),
                created_by=request.get("created_by", ""),
                system_prompt=request.get("system_prompt", ""),
                tools_allowed=request.get("tools_allowed", []),
                identity=request.get("identity", {}),
                skills=request.get("skills", []),
            )
            agent = state.agent_manager.create_agent(agent)

        if not agent:
            raise HTTPException(status_code=400, detail="创建 Agent 失败")

        # 初始化运行时
        if state.agent_runtime_pool:
            state.agent_runtime_pool.get_or_create(agent)
        
        # 初始化技能树（P4: Agent成长系统）
        if state.skill_manager:
            try:
                state.skill_manager.init_agent_skills(agent.id, agent.role)
            except Exception as e:
                logger.debug(f"[MAS] 技能树初始化失败（非关键）: {e}")

        return {"status": "ok", "agent": agent.to_dict()}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 创建 Agent 失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))



@router.post("/api/v1/agents/generate")
async def generate_agent(request: Request):
    """AI动态生成Agent配置"""
    try:
        data = await request.json()
        description = data.get("description", "")
        user_id = data.get("user_id", "default")
        if not description:
            raise HTTPException(status_code=400, detail="缺少描述")
        result = await state.agent_orchestrator.generate_agent_from_description(user_id, description)
        return {"status": "ok", "agent": result}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] AI生成Agent失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/v1/agents/orchestrate")
async def orchestrate_message(request: Request):
    """主-子Agent调度入口"""
    try:
        data = await request.json()
        user_id = data.get("user_id", "default")
        message = data.get("message", "")
        context = data.get("context", {})
        if not message:
            raise HTTPException(status_code=400, detail="缺少消息")
        result = await state.agent_orchestrator.handle_message(user_id, message, context)
        return {"status": "ok", "result": result}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 调度失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/v1/agents/status")
async def get_agents_status():
    """获取 Agent 团队整体状态"""
    try:
        if not state.agent_orchestrator:
            raise HTTPException(status_code=503, detail="Agent 编排器未初始化")
        status = state.agent_orchestrator.get_team_status(user_id="web_user")
        return {"status": "ok", **status}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 获取团队状态失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/v1/agents/rooms")
async def list_rooms(status: str = ""):
    """列出会议室"""
    try:
        if not state.agent_manager:
            return {"rooms": []}
        rooms = state.agent_manager.list_rooms(status=status)
        return {"rooms": [
            {
                "id": r.id, "name": r.name, "topic": r.topic,
                "participants": r.participants, "host_agent_id": r.host_agent_id,
                "status": r.status, "summary": r.summary,
                "created_at": r.created_at,
            } for r in rooms
        ]}
    except Exception as e:
        logger.warning(f"[MAS] 列出会议室失败: {e}")
        return {"rooms": [], "error": str(e)}





@router.post("/api/v1/agents/rooms")
async def create_room(request: dict):
    """创建会议室"""
    try:
        if not state.agent_manager:
            raise HTTPException(status_code=503, detail="Agent 管理器未初始化")
        room = state.agent_manager.create_room(
            name=request.get("name", "未命名会议"),
            topic=request.get("topic", ""),
            participants=request.get("participants", []),
            host_agent_id=request.get("host_agent_id", ""),
            created_by=request.get("created_by", ""),
        )
        return {"status": "ok", "room": {
            "id": room.id, "name": room.name, "topic": room.topic,
            "participants": room.participants, "status": room.status,
        }}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 创建会议室失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))





@router.get("/api/v1/agents/heartbeat")
async def get_agents_heartbeat():
    """运行 Agent 心跳自检"""
    try:
        if not state.agent_orchestrator:
            raise HTTPException(status_code=503, detail="Agent 编排器未初始化")
        result = await state.agent_orchestrator.run_heartbeat(user_id="web_user")
        return {"status": "ok", **result}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 心跳检查失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))





@router.get("/api/v1/agents/suggestions")
async def get_team_suggestions():
    """获取团队主动建议"""
    try:
        if not state.agent_orchestrator:
            raise HTTPException(status_code=503, detail="Agent 编排器未初始化")
        suggestion = await state.agent_orchestrator.generate_team_suggestion(user_id="web_user")
        return {"status": "ok", "suggestion": suggestion}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 建议生成失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))






@router.get("/api/v1/agents/rooms/{room_id}/messages")
async def get_room_messages(room_id: str, limit: int = 100):
    """获取会议室消息"""
    try:
        if not state.agent_manager:
            return {"messages": []}
        messages = state.agent_manager.get_messages(room_id, limit=limit)
        return {"messages": [
            {
                "id": m.id, "from_agent_id": m.from_agent_id,
                "to_agent_id": m.to_agent_id, "message_type": m.message_type,
                "content": m.content, "metadata": m.metadata,
                "created_at": m.created_at,
            } for m in messages
        ]}
    except Exception as e:
        logger.warning(f"[MAS] 获取消息失败: {e}")
        return {"messages": [], "error": str(e)}





@router.post("/api/v1/agents/rooms/{room_id}/messages")
async def add_room_message(room_id: str, request: dict):
    """发送会议室消息"""
    try:
        if not state.agent_manager:
            raise HTTPException(status_code=503, detail="Agent 管理器未初始化")
        msg = state.agent_manager.add_message(
            room_id=room_id,
            from_agent_id=request.get("from_agent_id", ""),
            to_agent_id=request.get("to_agent_id"),
            content=request.get("content", ""),
            message_type=request.get("message_type", "text"),
            metadata=request.get("metadata", {}),
        )
        return {"status": "ok", "message": {
            "id": msg.id, "content": msg.content,
            "from_agent_id": msg.from_agent_id,
            "created_at": msg.created_at,
        }}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 发送消息失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))





@router.post("/api/v1/agents/rooms/{room_id}/start")
async def start_room_meeting(room_id: str, request: dict):
    """启动会议室讨论"""
    try:
        if not state.agent_orchestrator:
            raise HTTPException(status_code=503, detail="Agent 编排器未初始化")
        topic = request.get("topic", "")
        participant_ids = request.get("participant_ids", [])
        rounds = request.get("rounds", 2)
        if not topic:
            raise HTTPException(status_code=400, detail="缺少讨论主题")
        if not participant_ids:
            raise HTTPException(status_code=400, detail="缺少参与者")
        result = await state.agent_orchestrator.start_meeting(
            room_id=room_id,
            topic=topic,
            participant_ids=participant_ids,
            rounds=rounds,
        )
        return {"status": "ok", "result": result}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 启动会议失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))





@router.get("/api/v1/agents/relationship-matrix")
async def get_relationship_matrix():
    """获取团队完整关系矩阵（用于关系图谱可视化）"""
    if not state.agent_manager:
        raise HTTPException(status_code=503, detail="Agent 管理器未初始化")
    try:
        matrix = state.agent_manager.get_relationship_matrix()
        stats = state.agent_manager.get_collaboration_stats()
        return {"status": "ok", **matrix, **stats}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/v1/agents/{agent_id}")
async def get_agent(agent_id: str):
    """获取 Agent 详情"""
    try:
        if not state.agent_manager:
            raise HTTPException(status_code=503, detail="Agent 管理器未初始化")
        agent = state.agent_manager.get_agent(agent_id)
        if not agent:
            raise HTTPException(status_code=404, detail="Agent 不存在")

        # 获取运行时状态
        runtime_stats = {}
        if state.agent_runtime_pool:
            runtime = state.agent_runtime_pool.get(agent_id)
            if runtime:
                runtime_stats = runtime.get_stats()
        
        # 获取技能树状态（P4）
        skill_stats = {}
        if state.skill_manager:
            try:
                skill_stats = state.skill_manager.get_agent_stats(agent_id)
            except Exception:
                pass

        return {
            "status": "ok",
            "agent": agent.to_dict(),
            "runtime": runtime_stats,
            "skills": skill_stats,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 获取 Agent 失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/v1/agents/{agent_id}/skills")
async def get_agent_skills(agent_id: str):
    """获取 Agent 技能树详情"""
    if not state.skill_manager:
        raise HTTPException(status_code=503, detail="技能系统未初始化")
    try:
        agent = state.agent_manager.get_agent(agent_id)
        if not agent:
            raise HTTPException(status_code=404, detail="Agent 不存在")
        skills = state.skill_manager.get_agent_skills(agent_id)
        stats = state.skill_manager.get_agent_stats(agent_id)
        return {"status": "ok", "agent_id": agent_id, "agent_name": agent.name, "skills": skills, "stats": stats}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/v1/agents/{agent_id}/xp-history")
async def get_agent_xp_history(agent_id: str, limit: int = 50):
    """获取 Agent XP 获取历史"""
    if not state.skill_manager:
        raise HTTPException(status_code=503, detail="技能系统未初始化")
    try:
        history = state.skill_manager.get_xp_history(agent_id, limit=limit)
        return {"status": "ok", "agent_id": agent_id, "history": history, "count": len(history)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/v1/agents/{agent_id}/skills/reset")
async def reset_agent_skills(agent_id: str):
    """重置 Agent 技能树"""
    if not state.skill_manager:
        raise HTTPException(status_code=503, detail="技能系统未初始化")
    try:
        agent = state.agent_manager.get_agent(agent_id)
        if not agent:
            raise HTTPException(status_code=404, detail="Agent 不存在")
        state.skill_manager.reset_agent_skills(agent_id)
        state.skill_manager.init_agent_skills(agent_id, agent.role)
        return {"status": "ok", "message": f"{agent.name} 的技能树已重置"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ──────────────────────────────────────
#  P5: Agent 间关系与协作网络 API
# ──────────────────────────────────────

@router.get("/api/v1/agents/{agent_id}/relationships")
async def get_agent_relationships(agent_id: str):
    """获取 Agent 的关系网络"""
    if not state.agent_manager:
        raise HTTPException(status_code=503, detail="Agent 管理器未初始化")
    try:
        agent = state.agent_manager.get_agent(agent_id)
        if not agent:
            raise HTTPException(status_code=404, detail="Agent 不存在")
        relationships = state.agent_manager.get_relationships(agent_id)
        stats = state.agent_manager.get_collaboration_stats(agent_id)
        return {"status": "ok", "agent_id": agent_id, "relationships": relationships, "stats": stats}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ──────────────────────────────────────
#  P6: 情感状态与记忆回放 API
# ──────────────────────────────────────

@router.get("/api/v1/soul/{user_id}/emotion-history")
async def get_emotion_history(user_id: str, limit: int = 100, hours: int = 24):
    """获取用户情感状态历史"""
    if not state.emotion_service:
        raise HTTPException(status_code=503, detail="情感服务未初始化")
    try:
        history = state.emotion_service.get_emotion_history(user_id, limit=limit)
        insights = state.emotion_service.get_emotion_insights(user_id, window_hours=float(hours))
        return {"status": "ok", "user_id": user_id, "history": history, "insights": insights}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/v1/soul/{user_id}/emotion-timeline")
async def get_emotion_timeline(user_id: str, limit: int = 50):
    """获取情感时间线（用于前端可视化）"""
    if not state.emotion_service:
        raise HTTPException(status_code=503, detail="情感服务未初始化")
    try:
        history = state.emotion_service.get_emotion_history(user_id, limit=limit)
        timeline = []
        for h in history:
            timeline.append({
                "timestamp": h.get("timestamp"),
                "primary": h.get("fused_state", {}).get("primary", "neutral"),
                "intensity": h.get("fused_state", {}).get("intensity", 0.5),
                "valence": h.get("fused_state", {}).get("valence", 0),
                "arousal": h.get("fused_state", {}).get("arousal", 0),
                "trigger_topic": h.get("trigger_topic", ""),
            })
        return {"status": "ok", "user_id": user_id, "timeline": timeline}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/v1/agents/{agent_id}")
async def update_agent(agent_id: str, request: dict):
    """更新 Agent"""
    try:
        if not state.agent_manager:
            raise HTTPException(status_code=503, detail="Agent 管理器未初始化")
        agent = state.agent_manager.update_agent(agent_id, request)
        if not agent:
            raise HTTPException(status_code=404, detail="Agent 不存在")
        return {"status": "ok", "agent": agent.to_dict()}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 更新 Agent 失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))





@router.delete("/api/v1/agents/{agent_id}")
async def delete_agent(agent_id: str):
    """删除 Agent"""
    try:
        if not state.agent_manager:
            raise HTTPException(status_code=503, detail="Agent 管理器未初始化")
        success = state.agent_manager.delete_agent(agent_id)
        if not success:
            raise HTTPException(status_code=404, detail="Agent 不存在")
        if state.agent_runtime_pool:
            state.agent_runtime_pool.remove(agent_id)
        return {"status": "ok", "agent_id": agent_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 删除 Agent 失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))





@router.post("/api/v1/agents/{agent_id}/run")
async def run_agent(agent_id: str, request: dict):
    """运行 Agent 执行任务"""
    try:
        if not state.agent_manager or not state.agent_runtime_pool:
            raise HTTPException(status_code=503, detail="Agent 系统未初始化")

        agent = state.agent_manager.get_agent(agent_id)
        if not agent:
            raise HTTPException(status_code=404, detail="Agent 不存在")

        runtime = state.agent_runtime_pool.get_or_create(agent)
        result = await runtime.run(
            task=request.get("task", ""),
            context=request.get("context", {}),
        )
        return {"status": "ok", "result": result}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] 运行 Agent 失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))





@router.post("/api/v1/agents/{agent_id}/chat")
async def chat_with_agent(agent_id: str, request: dict):
    """与 Agent 对话"""
    try:
        if not state.agent_manager or not state.agent_runtime_pool:
            raise HTTPException(status_code=503, detail="Agent 系统未初始化")

        agent = state.agent_manager.get_agent(agent_id)
        if not agent:
            raise HTTPException(status_code=404, detail="Agent 不存在")

        runtime = state.agent_runtime_pool.get_or_create(agent)
        reply = await runtime.chat(request.get("message", ""))
        return {"status": "ok", "reply": reply, "agent_name": agent.name}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[MAS] Agent 对话失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ========== 会议室 API ==========

# ========== 记忆系统与知识图谱 API ==========

@router.get("/api/v1/memory/stats")
async def memory_stats(user_id: str = "web_user"):
    """获取用户记忆系统统计信息"""
    result = {
        "graph": {"node_count": 0, "edge_count": 0},
        "tiered": {"l0_count": 0, "l1_count": 0},
        "sessions": 0,
    }
    if state.cognitive_graph:
        try:
            result["graph"] = state.cognitive_graph.get_statistics()
        except Exception as e:
            logger.warning(f"[MemoryAPI] 图谱统计失败: {e}")
    if state.memory_store:
        try:
            l0 = state.memory_store.db.execute("SELECT COUNT(*) FROM l0_index").fetchone()[0]
            l1 = state.memory_store.db.execute("SELECT COUNT(*) FROM l1_index").fetchone()[0]
            result["tiered"] = {"l0_count": l0, "l1_count": l1}
        except Exception as e:
            logger.warning(f"[MemoryAPI] 分层记忆统计失败: {e}")
    if state.state_store:
        try:
            sessions = await state.state_store.list_sessions(user_id=user_id)
            result["sessions"] = len(sessions)
        except Exception:
            pass
    return result


@router.get("/api/v1/memory/graph")
async def memory_graph(user_id: str = "web_user", limit: int = 200):
    if not state.cognitive_graph:
        return {"nodes": [], "edges": [], "total_nodes": 0, "total_edges": 0}
    try:
        nodes = state.cognitive_graph.get_all_nodes(limit=limit)
        node_ids = {n.id for n in nodes}
        edges_raw = state.cognitive_graph.db.execute(
            "SELECT source_id, target_id, relation_type, strength FROM edges "
            "WHERE source_id IN ({}) OR target_id IN ({})".format(
                ",".join("?" * len(node_ids)), ",".join("?" * len(node_ids))
            ),
            list(node_ids) + list(node_ids)
        ).fetchall()
        edges = []
        for row in edges_raw:
            if row["source_id"] in node_ids and row["target_id"] in node_ids:
                edges.append({
                    "source": row["source_id"],
                    "target": row["target_id"],
                    "relation": row["relation_type"],
                    "strength": row["strength"],
                })
        seen = set()
        unique_edges = []
        for e in edges:
            key = (e["source"], e["target"], e["relation"])
            if key not in seen:
                seen.add(key)
                unique_edges.append(e)
        return {
            "nodes": [{"id": n.id, "content": n.content[:100], "type": n.memory_type, "confidence": n.confidence, "created_at": n.created_at} for n in nodes],
            "edges": unique_edges,
            "total_nodes": state.cognitive_graph.get_statistics()["node_count"],
            "total_edges": state.cognitive_graph.get_statistics()["edge_count"],
        }
    except Exception as e:
        logger.warning(f"[MemoryAPI] 图谱查询失败: {e}")
        return {"nodes": [], "edges": [], "total_nodes": 0, "total_edges": 0}


@router.get("/api/v1/memory/summary")
async def memory_summary(user_id: str = "web_user", limit: int = 20):
    if not state.memory_store:
        return {"summaries": [], "count": 0}
    try:
        rows = state.memory_store.db.execute(
            "SELECT uri, overview, overview_tokens, updated_at FROM l1_index "
            "WHERE overview IS NOT NULL AND overview != '' ORDER BY updated_at DESC LIMIT ?",
            (limit,)
        ).fetchall()
        return {
            "summaries": [{"uri": row["uri"], "overview": row["overview"][:500], "tokens": row["overview_tokens"], "updated_at": row["updated_at"]} for row in rows],
            "count": len(rows),
        }
    except Exception as e:
        logger.warning(f"[MemoryAPI] 摘要查询失败: {e}")
        return {"summaries": [], "count": 0}


@router.get("/api/v1/memory/recent")
async def memory_recent(user_id: str = "web_user", limit: int = 20, hours: int = None):
    if not state.memory_store:
        return {"memories": [], "count": 0}
    try:
        memories = state.memory_store.get_recent(limit=limit, user_id=user_id, hours=hours)
        return {"memories": memories, "count": len(memories)}
    except Exception as e:
        logger.warning(f"[MemoryAPI] 近期记忆查询失败: {e}")
        return {"memories": [], "count": 0}


@router.post("/api/v1/memory/compress")
async def memory_compress_now(user_id: str = "web_user", hours: int = 24):
    if not state.memory_store or not state._llm:
        raise HTTPException(status_code=503, detail="记忆系统未就绪")
    try:
        result = await state.memory_store.auto_compress_l0_to_l1(
            user_id=user_id if user_id != "all" else None, hours=hours,
        )
        return {
            "status": "ok",
            "compressed_count": result.get("compressed_count", 0),
            "generated_uris": result.get("generated_uris", []),
            "summary_preview": result.get("summary_preview", ""),
        }
    except Exception as e:
        logger.error(f"[MemoryAPI] 手动压缩失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/v1/memory/timeline")
async def memory_timeline(user_id: str = "web_user", limit: int = 50):
    events = []
    if state.memory_store:
        try:
            l0 = state.memory_store.get_recent(limit=limit, user_id=user_id)
            for m in l0:
                events.append({
                    "time": m.get("created_at", ""), "type": "memory",
                    "subtype": m.get("memory_type", "unknown"),
                    "content": m.get("abstract", "")[:150], "uri": m.get("uri", ""),
                })
        except Exception:
            pass
    if state.cognitive_graph:
        try:
            nodes = state.cognitive_graph.get_all_nodes(limit=limit)
            for n in nodes:
                events.append({
                    "time": n.created_at, "type": "entity", "subtype": n.memory_type,
                    "content": n.content[:150], "id": n.id, "confidence": n.confidence,
                })
        except Exception:
            pass
    events.sort(key=lambda x: x.get("time", ""), reverse=True)
    return {"events": events[:limit], "count": len(events)}


# ========== MCP (Model Context Protocol) API ==========

@router.get("/api/v1/mcp/servers")
async def mcp_list_servers():
    """列出所有 MCP Server 配置及连接状态"""
    if not state.mcp_manager:
        return {"servers": [], "connected_count": 0}
    return {"servers": state.mcp_manager.list_servers(), "connected_count": len(state.mcp_manager._clients)}


@router.post("/api/v1/mcp/servers")
async def mcp_add_server(request: Request):
    """添加 MCP Server 配置"""
    if not state.mcp_manager:
        raise HTTPException(status_code=503, detail="MCP 管理器未初始化")
    try:
        body = await request.json()
        name = body.get("name", "").strip()
        command = body.get("command", "").strip()
        if not name or not command:
            raise HTTPException(status_code=400, detail="name 和 command 不能为空")
        ok = state.mcp_manager.add_server(
            name=name,
            command=command,
            args=body.get("args", []),
            env=body.get("env", {}),
            enabled=body.get("enabled", True),
        )
        if ok:
            # 如果启用，自动尝试连接
            if body.get("enabled", True):
                asyncio.create_task(state.mcp_manager.connect_server(name))
            return {"status": "ok", "message": f"Server '{name}' 已添加"}
        raise HTTPException(status_code=500, detail="添加失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/api/v1/mcp/servers/{name}")
async def mcp_remove_server(name: str):
    """删除 MCP Server 配置"""
    if not state.mcp_manager:
        raise HTTPException(status_code=503, detail="MCP 管理器未初始化")
    if state.mcp_manager.remove_server(name):
        return {"status": "ok", "message": f"Server '{name}' 已删除"}
    raise HTTPException(status_code=500, detail="删除失败")


@router.post("/api/v1/mcp/servers/{name}/connect")
async def mcp_connect_server(name: str):
    """手动连接 MCP Server"""
    if not state.mcp_manager:
        raise HTTPException(status_code=503, detail="MCP 管理器未初始化")
    ok = await state.mcp_manager.connect_server(name)
    if ok:
        tools = state.mcp_manager._clients.get(name, {}).tools or []
        return {"status": "ok", "message": f"已连接 '{name}'", "tools_count": len(tools)}
    raise HTTPException(status_code=500, detail=f"连接 '{name}' 失败")


@router.post("/api/v1/mcp/servers/{name}/disconnect")
async def mcp_disconnect_server(name: str):
    """断开 MCP Server"""
    if not state.mcp_manager:
        raise HTTPException(status_code=503, detail="MCP 管理器未初始化")
    await state.mcp_manager.disconnect_server(name)
    return {"status": "ok", "message": f"'{name}' 已断开"}


@router.get("/api/v1/mcp/tools")
async def mcp_list_tools():
    """列出所有已连接的 MCP 工具"""
    if not state.mcp_manager:
        return {"tools": [], "count": 0}
    tools = state.mcp_manager.get_all_tools()
    return {
        "tools": [
            {
                "name": t.name,
                "description": t.description,
                "server": t.server_name,
                "schema": t.input_schema,
            }
            for t in tools
        ],
        "count": len(tools),
    }


@router.post("/api/v1/mcp/tools/call")
async def mcp_call_tool(request: Request):
    """调用 MCP 工具"""
    if not state.mcp_manager:
        raise HTTPException(status_code=503, detail="MCP 管理器未初始化")
    try:
        body = await request.json()
        tool_name = body.get("tool", "")
        arguments = body.get("arguments", {})
        server_name = body.get("server", "")
        if not tool_name:
            raise HTTPException(status_code=400, detail="tool 不能为空")
        if server_name:
            result = await state.mcp_manager.call_tool(server_name, tool_name, arguments)
        else:
            result = await state.mcp_manager.call_tool_any(tool_name, arguments)
        return {"status": "ok", "result": result}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ========== 静态文件服务（React SPA） ==========

_UI_DIST_DIR = Path(__file__).parent.parent.parent / "frontend" / "desktop" / "dist"

@router.get("/{path:path}", include_in_schema=False)
async def serve_ui(path: str):
    """服务 React SPA —— 所有未匹配路由回退到 index.html"""
    if not _UI_DIST_DIR.exists():
        return HTMLResponse(
            "<html><body style='font-family:sans-serif;padding:40px;text-align:center'>"
            "<h1>Tent OS 灵魂对讲机</h1>"
            "<p>前端文件未找到。请检查 frontend/desktop/dist/ 目录。</p>"
            "</body></html>"
        )
    file_path = _UI_DIST_DIR / path
    if path == "" or not file_path.exists() or file_path.is_dir():
        file_path = _UI_DIST_DIR / "index.html"
    _NO_CACHE = {"Cache-Control": "no-store, no-cache, must-revalidate, proxy-revalidate", "Pragma": "no-cache", "Expires": "0"}
    if file_path.exists():
        return FileResponse(file_path, headers=_NO_CACHE)
    return FileResponse(_UI_DIST_DIR / "index.html", headers=_NO_CACHE)


# ========== MCP (Model Context Protocol) 端点 ==========

_mcp_server = None
_mcp_auth = None

def _get_mcp_server():
    global _mcp_server
    if _mcp_server is None:
        from tent_os.mcp.server import MCPServer
        from tent_os.mcp.tools import register_all_tools
        _mcp_server = MCPServer(state=state)
        register_all_tools(_mcp_server)
    return _mcp_server

def _get_mcp_auth():
    global _mcp_auth
    if _mcp_auth is None:
        from tent_os.mcp.auth import MCPAuthManager
        _mcp_auth = MCPAuthManager()
    return _mcp_auth

def _extract_auth(request: Request) -> Optional[Dict]:
    """从请求中提取 Bearer token"""
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
        auth_mgr = _get_mcp_auth()
        return auth_mgr.verify_token(token)
    return None

@router.post("/api/v1/mcp")
async def mcp_endpoint(request: Request):
    """MCP JSON-RPC 消息入口"""
    body = await request.body()
    auth_context = _extract_auth(request)
    mcp = _get_mcp_server()
    result = await mcp.handle_message(body.decode("utf-8"), auth_context)
    return result

@router.post("/api/v1/mcp/auth/register")
async def mcp_register_client(request: Request):
    """注册 MCP 客户端（设备）"""
    try:
        data = await request.json()
    except Exception:
        data = {}
    user_id = data.get("user_id", "")
    device_name = data.get("device_name", "未命名设备")
    device_type = data.get("device_type", "robot")
    permissions = data.get("permissions", ["chat", "query_persona", "query_memories", "synthesize_tts"])
    if not user_id:
        raise HTTPException(status_code=400, detail="user_id is required")
    auth_mgr = _get_mcp_auth()
    result = auth_mgr.register_client(user_id, device_name, device_type, permissions)
    return {"status": "ok", **result}

@router.post("/api/v1/mcp/auth/token")
async def mcp_get_token(request: Request):
    """获取 MCP access_token（client_credentials 模式）"""
    try:
        data = await request.json()
    except Exception:
        data = {}
    client_id = data.get("client_id", "")
    client_secret = data.get("client_secret", "")
    if not client_id or not client_secret:
        raise HTTPException(status_code=400, detail="client_id and client_secret are required")
    auth_mgr = _get_mcp_auth()
    result = auth_mgr.authenticate(client_id, client_secret)
    if not result:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    return {"status": "ok", **result}

@router.get("/api/v1/mcp/auth/clients")
async def mcp_list_clients(user_id: str):
    """列出用户的 MCP 客户端"""
    auth_mgr = _get_mcp_auth()
    clients = auth_mgr.get_clients(user_id)
    return {"status": "ok", "clients": clients}

@router.delete("/api/v1/mcp/auth/clients/{client_id}")
async def mcp_revoke_client(client_id: str):
    """吊销 MCP 客户端"""
    auth_mgr = _get_mcp_auth()
    auth_mgr.revoke_client(client_id)
    return {"status": "ok", "message": f"Client {client_id} revoked"}

@router.get("/api/v1/mcp/audit")
async def mcp_audit_log(user_id: str, limit: int = 100):
    """MCP 审计日志"""
    auth_mgr = _get_mcp_auth()
    logs = auth_mgr.get_audit_log(user_id, limit)
    return {"status": "ok", "logs": logs}


# ========== Phase 3 预留接口（克隆与视频对话）==========

@router.get("/api/v1/soul/phase3/status")
async def get_phase3_status(user_id: str):
    """获取 Phase 3（克隆与视频对话）就绪状态"""
    from pathlib import Path
    
    # 检查样本数据
    voice_dir = Path("./tent_memory/soul/voice_samples") / user_id
    photo_dir = Path("./tent_memory/soul/appearance_samples") / user_id
    
    voice_samples = len(list(voice_dir.glob("*"))) if voice_dir.exists() else 0
    photo_samples = len(list(photo_dir.glob("photo_*"))) if photo_dir.exists() else 0
    
    # 检查 GPU 可用性
    gpu_available = False
    try:
        import torch
        gpu_available = torch.cuda.is_available()
    except ImportError:
        pass
    
    return {
        "status": "ok",
        "phase": "phase2_sample_collection",
        "gpu_available": gpu_available,
        "voice_clone": {
            "samples_collected": voice_samples,
            "engine_ready": False,
            "engines": ["gpt_sovits", "f5_tts", "cosyvoice"],
            "note": "需要 GPU 环境（>=6GB VRAM）和 pip install gpt-sovits/f5-tts",
        },
        "avatar_3d": {
            "photos_collected": photo_samples,
            "engine_ready": False,
            "engines": ["deca", "live_portrait", "hallo_live"],
            "note": "需要 GPU 环境（>=6GB VRAM）和 pip install deca/live-portrait",
        },
        "video_call": {
            "engine_ready": False,
            "engines": ["webrtc_aiortc"],
            "note": "需要 aiortc + GPU 渲染服务",
        },
        "migration_path": {
            "step1": "在 GPU 服务器部署克隆模型",
            "step2": "配置模型路径到 tent_os.yaml",
            "step3": "调用 /api/v1/soul/voice/{user_id}/clone 训练声纹",
            "step4": "调用 /api/v1/soul/appearance/{user_id}/reconstruct 重建形象",
            "step5": "启用 /api/v1/soul/video_call/{user_id}/session 开始视频对话",
        },
    }


@router.post("/api/v1/soul/video_call/{user_id}/session")
async def create_video_call_session(user_id: str, request: Request):
    """创建视频通话会话（Phase 3 预留）"""
    return {
        "status": "not_ready",
        "message": "视频通话需要 Phase 3 GPU 环境。当前仅支持文本和语音对话。",
        "phase3_status_url": f"/api/v1/soul/phase3/status?user_id={user_id}",
    }


@router.post("/api/v1/soul/appearance/{user_id}/reconstruct")
async def reconstruct_avatar_3d(user_id: str):
    """触发 3D 形象重建（Phase 3 预留）"""
    return {
        "status": "not_ready",
        "message": "3D 形象重建需要 Phase 3 GPU 环境部署 DECA/LivePortrait。当前为 Phase 2（颜色提取阶段）。",
        "phase3_status_url": f"/api/v1/soul/phase3/status?user_id={user_id}",
    }


# ========== 文件上传 ==========

_UPLOAD_DIR = Path("./uploads")
_UPLOAD_DIR.mkdir(exist_ok=True)


def _extract_pdf_text(file_bytes: bytes) -> str:
    try:
        from pypdf import PdfReader
        import io
        reader = PdfReader(io.BytesIO(file_bytes))
        texts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                texts.append(text)
        return "\n\n".join(texts)
    except Exception as e:
        return f"[PDF 解析失败: {e}]"


def _extract_docx_text(file_bytes: bytes) -> str:
    try:
        from docx import Document
        import io
        doc = Document(io.BytesIO(file_bytes))
        texts = []
        for para in doc.paragraphs:
            if para.text.strip():
                texts.append(para.text)
        return "\n".join(texts)
    except Exception as e:
        return f"[DOCX 解析失败: {e}]"


def _extract_xlsx_text(file_bytes: bytes) -> str:
    try:
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        texts = []
        for sheet in wb.worksheets:
            texts.append(f"--- Sheet: {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    texts.append(row_text)
        return "\n".join(texts)
    except Exception as e:
        return f"[XLSX 解析失败: {e}]"


def _extract_text_file(file_bytes: bytes) -> str:
    try:
        return file_bytes.decode('utf-8')
    except UnicodeDecodeError:
        try:
            return file_bytes.decode('gbk')
        except UnicodeDecodeError:
            return file_bytes.decode('utf-8', errors='replace')


@router.post("/api/v1/files/upload")
async def upload_file(file: UploadFile = File(...)):
    """上传文件并提取文本内容"""
    file_bytes = await file.read()
    size = len(file_bytes)
    ext = Path(file.filename or "unknown").suffix.lower()
    storage_name = f"{uuid.uuid4().hex}{ext}"
    storage_path = _UPLOAD_DIR / storage_name
    storage_path.write_bytes(file_bytes)

    content_type = (file.content_type or "").lower()
    text = ""

    if content_type == "application/pdf" or ext == ".pdf":
        text = _extract_pdf_text(file_bytes)
    elif content_type in ("application/vnd.openxmlformats-officedocument.wordprocessingml.document",) or ext == ".docx":
        text = _extract_docx_text(file_bytes)
    elif content_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",) or ext == ".xlsx":
        text = _extract_xlsx_text(file_bytes)
    elif content_type.startswith("text/") or ext in (".txt", ".md", ".csv", ".json", ".py", ".js", ".ts", ".html", ".css", ".yaml", ".yml"):
        text = _extract_text_file(file_bytes)
    else:
        text = f"[不支持的文件类型: {content_type or ext}]"

    max_len = 10000
    truncated = len(text) > max_len
    display_text = text[:max_len] + ("\n\n[...内容已截断，共 {} 字符]".format(len(text)) if truncated else "")

    return {
        "filename": file.filename,
        "content_type": file.content_type,
        "size": size,
        "text": display_text,
        "full_length": len(text),
        "truncated": truncated,
        "storage_path": str(storage_path),
    }


